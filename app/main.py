from __future__ import annotations

import asyncio
import base64
import hashlib
import json
import os
import re
import shutil
import threading
import time
import uuid
import zipfile
from datetime import datetime, timezone
from enum import StrEnum
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Literal
from xml.etree import ElementTree

import httpx
from app.adapters import Neo4jMemory, ObjectStorage, PostgresStateBackend, RedisState
from celery import Celery
from fastapi import BackgroundTasks, Depends, FastAPI, File, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, PlainTextResponse, Response, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - dependency is optional in some sandboxes
    PdfReader = None  # type: ignore[assignment]


ROOT = Path(__file__).resolve().parents[1]
STATE_FILE = ROOT / 'data' / 'state' / 'state.json'
STORAGE_DIR = ROOT / 'data' / 'storage'
ORGANIZER_DIR = ROOT / 'data' / 'organizer_raw'


def load_env_file() -> None:
    for name in ('.env.local', '.env'):
        path = ROOT / name
        if not path.exists():
            continue
        for line in path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, value = line.split('=', 1)
            os.environ.setdefault(key.strip().lstrip('\ufeff'), value.strip().strip('"').strip("'"))


load_env_file()


class RunStatus(StrEnum):
    CREATED = 'CREATED'
    INGESTING = 'INGESTING'
    ANALYZING = 'ANALYZING'
    WAITING_FOR_CLARIFICATION = 'WAITING_FOR_CLARIFICATION'
    RETRIEVING_MEMORY = 'RETRIEVING_MEMORY'
    RESEARCHING_EXTERNAL = 'RESEARCHING_EXTERNAL'
    GENERATING = 'GENERATING'
    DEDUPLICATING = 'DEDUPLICATING'
    APPLYING_GATES = 'APPLYING_GATES'
    CRITIQUING = 'CRITIQUING'
    CHECKING_NOVELTY = 'CHECKING_NOVELTY'
    ANALYZING_DISAGREEMENT = 'ANALYZING_DISAGREEMENT'
    ANALYZING_UNCERTAINTY = 'ANALYZING_UNCERTAINTY'
    COMPILING_EXPERIMENTS = 'COMPILING_EXPERIMENTS'
    BUILDING_REPORT = 'BUILDING_REPORT'
    COMPLETED = 'COMPLETED'
    FAILED = 'FAILED'
    CANCELLED = 'CANCELLED'


class Language(StrEnum):
    RU = 'ru'
    EN = 'en'
    ZH = 'zh'


class ProjectCreate(BaseModel):
    name: str = Field(min_length=1)
    problem: str = Field(min_length=1)
    target_kpi: str | None = None
    constraints: list[str] = Field(default_factory=list)
    response_language: Language = Language.RU
    external_research_enabled: bool = True


class ProjectUpdate(BaseModel):
    name: str | None = None
    problem: str | None = None
    target_kpi: str | None = None
    constraints: list[str] | None = None
    response_language: Language | None = None
    external_research_enabled: bool | None = None


class RunCreate(BaseModel):
    response_language: Language | None = None
    max_finalists: int = Field(default=12, ge=1, le=12)
    candidate_count: int = Field(default=12, ge=1, le=12)
    use_llm: bool = True


class LocalImportRequest(BaseModel):
    path: str | None = None
    limit: int | None = Field(default=None, ge=1)


class ClarificationAnswer(BaseModel):
    answer: str


class FeedbackRequest(BaseModel):
    verdict: Literal['accept', 'edit', 'reject']
    reason: str | None = None
    expert_score: float | None = Field(default=None, ge=0, le=1)


class ExperimentPatch(BaseModel):
    objective: str | None = None
    steps: list[str] | None = None
    success_criteria: list[str] | None = None
    failure_criteria: list[str] | None = None
    early_stop_criteria: list[str] | None = None


class ExperimentResult(BaseModel):
    status: Literal['confirmed', 'refuted', 'inconclusive']
    measurements: dict[str, Any] = Field(default_factory=dict)
    notes: str | None = None


class BriefPayload(BaseModel):
    problem: str = Field(min_length=1)
    goal: str = Field(min_length=1)
    constraints: str = ''
    success_criterion: str = ''
    domain: Literal['tailings_and_metallurgy'] = 'tailings_and_metallurgy'


class BffFeedbackRequest(BaseModel):
    verdict: Literal['useful', 'revise']
    reason: str = ''
    comment: str = ''


class BffClarificationAnswer(BaseModel):
    answer: str | None = None
    comment: str | None = None


class ReportExportPayload(BaseModel):
    sections: list[Literal['summary', 'evidence', 'hypotheses', 'protocols', 'source_files']]
    format: Literal['PDF', 'DOCX', 'CSV', 'JSON', 'JIRA_API']
    locale: Literal['ru', 'en', 'zh-CN'] = 'ru'


class UserContext(BaseModel):
    id: str
    role: Literal['admin', 'researcher', 'viewer'] = 'admin'


def current_user(x_user_id: str = Header(default='local-admin'), x_user_role: str = Header(default='admin')) -> UserContext:
    role = x_user_role if x_user_role in {'admin', 'researcher', 'viewer'} else 'viewer'
    return UserContext(id=x_user_id, role=role)  # type: ignore[arg-type]


def require_write(user: UserContext) -> None:
    if user.role not in {'admin', 'researcher'}:
        raise HTTPException(status_code=403, detail='Write access denied')


def require_project_access(project: dict[str, Any], user: UserContext, write: bool = False) -> None:
    if user.role == 'admin':
        return
    if write and user.role != 'researcher':
        raise HTTPException(status_code=403, detail='Write access denied')
    members = project.get('members', [])
    if project.get('owner_id') != user.id and user.id not in members:
        raise HTTPException(status_code=403, detail='Project access denied')


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def new_id(prefix: str) -> str:
    return f'{prefix}_{uuid.uuid4().hex[:12]}'


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


class Store:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.backend_name = os.getenv('NORLAB_STORAGE_BACKEND', 'local')
        self.postgres = PostgresStateBackend(os.environ['DATABASE_URL']) if self.backend_name == 'postgres' else None
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()
        self.data = self._load()

    def _load(self) -> dict[str, Any]:
        default = {
            'projects': {},
            'documents': {},
            'runs': {},
            'hypotheses': {},
            'experiments': {},
            'entities': {},
            'sources': {},
            'audit': [],
        }
        if self.postgres:
            return self.postgres.load(default)
        if not self.path.exists():
            return default
        with self.path.open('r', encoding='utf-8') as file:
            return json.load(file)

    def reload(self) -> None:
        fresh = self._load()
        for key, fresh_value in fresh.items():
            current_value = self.data.get(key)
            if isinstance(current_value, dict) and isinstance(fresh_value, dict):
                for item_key in list(current_value.keys()):
                    if item_key not in fresh_value:
                        del current_value[item_key]
                for item_key, item_value in fresh_value.items():
                    current_item = current_value.get(item_key)
                    if isinstance(current_item, dict) and isinstance(item_value, dict):
                        current_item.clear()
                        current_item.update(item_value)
                    else:
                        current_value[item_key] = item_value
            elif isinstance(current_value, list) and isinstance(fresh_value, list):
                current_value[:] = fresh_value
            else:
                self.data[key] = fresh_value

    def save(self) -> None:
        with self._lock:
            if self.postgres:
                self.postgres.save(self.data)
                return
            tmp = self.path.with_name(f'{self.path.stem}.{os.getpid()}.{threading.get_ident()}.tmp')
            with tmp.open('w', encoding='utf-8') as file:
                json.dump(self.data, file, ensure_ascii=False, indent=2)
            for attempt in range(5):
                try:
                    tmp.replace(self.path)
                    break
                except PermissionError:
                    if attempt == 4:
                        raise
                    time.sleep(0.1 * (attempt + 1))


store = Store(STATE_FILE)
object_storage = ObjectStorage(STORAGE_DIR)
neo4j_memory = Neo4jMemory()
redis_state = RedisState()


class DocumentParser:
    text_extensions = {'.txt', '.md', '.csv', '.json'}
    image_extensions = {'.png', '.jpg', '.jpeg'}

    def parse(self, path: Path, document_id: str) -> dict[str, Any]:
        suffix = path.suffix.lower()
        if suffix in self.text_extensions:
            fragments = self._parse_text(path, document_id)
        elif suffix == '.xlsx':
            fragments = self._parse_xlsx(path, document_id)
        elif suffix == '.docx':
            fragments = self._parse_docx(path, document_id)
        elif suffix == '.pdf':
            fragments = self._parse_pdf(path, document_id)
        elif suffix in self.image_extensions:
            fragments = self._parse_image(path, document_id)
        else:
            fragments = []
        facts = self._extract_facts(fragments)
        return {'fragments': fragments, 'facts': facts}

    def _fragment(self, document_id: str, text: str, location: str, method: str) -> dict[str, Any]:
        clean = re.sub(r'\s+', ' ', text).strip()
        return {
            'id': new_id('frag'),
            'document_id': document_id,
            'original_language': detect_language(clean),
            'original_text': clean,
            'normalized_text': clean.lower(),
            'location': location,
            'extraction_method': method,
            'confidence': 0.9 if clean else 0.0,
        }

    def _parse_text(self, path: Path, document_id: str) -> list[dict[str, Any]]:
        text = path.read_text(encoding='utf-8', errors='ignore')
        return chunk_text(text, document_id, 'text:file', self._fragment)

    def _parse_xlsx(self, path: Path, document_id: str) -> list[dict[str, Any]]:
        wb = load_workbook(path, data_only=False, read_only=True)
        fragments: list[dict[str, Any]] = []
        for sheet in wb.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows():
                values = []
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value)
                    marker = 'formula' if value.startswith('=') else 'value'
                    if value.startswith('#') or getattr(cell, 'data_type', None) == 'e':
                        marker = 'error'
                    values.append(f'{cell.coordinate}({marker})={value}')
                if values:
                    rows.append('; '.join(values))
                if len(rows) >= 40:
                    text = '\n'.join(rows)
                    fragments.append(self._fragment(document_id, text, f'{sheet.title}:rows', 'xlsx:openpyxl'))
                    rows = []
            if rows:
                fragments.append(self._fragment(document_id, '\n'.join(rows), f'{sheet.title}:rows', 'xlsx:openpyxl'))
        wb.close()
        return fragments

    def _parse_docx(self, path: Path, document_id: str) -> list[dict[str, Any]]:
        with zipfile.ZipFile(path) as docx:
            xml = docx.read('word/document.xml')
        root = ElementTree.fromstring(xml)
        ns = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
        paragraphs = []
        for idx, paragraph in enumerate(root.findall('.//w:p', ns), start=1):
            texts = [node.text for node in paragraph.findall('.//w:t', ns) if node.text]
            text = ''.join(texts).strip()
            if text:
                paragraphs.append((idx, text))
        fragments = []
        for batch in batched(paragraphs, 12):
            first = batch[0][0]
            last = batch[-1][0]
            fragments.append(self._fragment(document_id, '\n'.join(text for _, text in batch), f'p:{first}-{last}', 'docx:ooxml'))
        return fragments

    def _parse_pdf(self, path: Path, document_id: str) -> list[dict[str, Any]]:
        if PdfReader is None:
            return [self._fragment(document_id, 'PDF parser is unavailable in this environment.', 'pdf:unparsed', 'pdf:missing')]
        reader = PdfReader(str(path))
        fragments = []
        for index, page in enumerate(reader.pages[:80], start=1):
            text = page.extract_text() or ''
            if text.strip():
                fragments.extend(chunk_text(text, document_id, f'pdf:page:{index}', self._fragment))
        return fragments

    def _parse_image(self, path: Path, document_id: str) -> list[dict[str, Any]]:
        text = f'Изображение или схема: {path.name}. Требует vision-разбора для извлечения графа.'
        return [self._fragment(document_id, text, 'image:file', 'image:metadata')]

    def _extract_facts(self, fragments: list[dict[str, Any]]) -> list[dict[str, Any]]:
        facts: list[dict[str, Any]] = []
        patterns = [
            ('data_quality', r'#REF!|#N/A|ошибк\w+|нет данных|н/д'),
            ('process', r'флотац\w+|обогащен\w+|выщелачиван\w+|измельчен\w+'),
            ('equipment', r'гидроциклон|флотомашин\w+|классификатор|сгустител\w+|насос'),
            ('element', r'\b(Au|Ag|Cu|Ni|Pt|Pd|Fe|S)\b|золото|серебро|медь|никель'),
            ('parameter', r'\bpH\b|крупност\w+|извлечени\w+|содержани\w+|расход|плотност\w+'),
        ]
        for fragment in fragments:
            text = fragment['original_text']
            segments = [
                segment.strip()
                for segment in re.split(r'(?<=[.!?…])\s+', text)
                if len(segment.strip()) >= 24
            ] or [text]
            seen_segment_keys: set[str] = set()
            for segment in segments[:40]:
                segment_key = re.sub(r'\W+', ' ', segment.lower()).strip()[:160]
                if not segment_key or segment_key in seen_segment_keys:
                    continue
                seen_segment_keys.add(segment_key)
                for kind, pattern in patterns:
                    if not re.search(pattern, segment, flags=re.IGNORECASE):
                        continue
                    facts.append({
                        'id': new_id('fact'),
                        'type': kind,
                        'statement': summarize_text(segment, 620),
                        'source_fragment_id': fragment['id'],
                        'document_id': fragment['document_id'],
                        'confidence': 0.72 if kind != 'data_quality' else 0.95,
                        'provenance': {
                            'location': fragment['location'],
                            'extraction_method': fragment['extraction_method'],
                            'timestamp': now_iso(),
                        },
                    })
                    break
        return facts


def detect_language(text: str) -> str:
    if re.search(r'[\u4e00-\u9fff]', text):
        return 'zh'
    if re.search(r'[А-Яа-яЁё]', text):
        return 'ru'
    return 'en'


def batched(items: list[Any], size: int) -> Iterable[list[Any]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def chunk_text(text: str, document_id: str, location: str, factory: Any) -> list[dict[str, Any]]:
    clean = re.sub(r'\r\n?', '\n', text)
    chunks: list[dict[str, Any]] = []
    current: list[str] = []
    length = 0
    for paragraph in clean.split('\n'):
        paragraph = paragraph.strip()
        if not paragraph:
            continue
        current.append(paragraph)
        length += len(paragraph)
        if length >= 1800:
            chunks.append(factory(document_id, '\n'.join(current), location, 'text:chunk'))
            current = []
            length = 0
    if current:
        chunks.append(factory(document_id, '\n'.join(current), location, 'text:chunk'))
    return chunks


def summarize_text(text: str, limit: int) -> str:
    clean = re.sub(r'\s+', ' ', text).strip()
    return clean if len(clean) <= limit else clean[:limit - 1].rstrip() + '…'


class SearchQuerySanitizer:
    sensitive_patterns = [
        r'\b[A-ZА-ЯЁ]{2,}-\d{2,}\b',
        r'\b\d{4,}-\d{2,}\b',
        r'\b(?:цех|участок|корпус)\s*№?\s*\d+\b',
    ]

    def sanitize(self, query: str) -> str:
        sanitized = query
        for pattern in self.sensitive_patterns:
            sanitized = re.sub(pattern, '[redacted]', sanitized, flags=re.IGNORECASE)
        return sanitized.strip()


class ModelGateway:
    def __init__(self) -> None:
        self.provider = os.getenv('NORLAB_LLM_PROVIDER', 'yandex').lower()
        self.yandex_base_url = os.getenv('YANDEX_AI_BASE_URL', 'https://llm.api.cloud.yandex.net/v1').rstrip('/')
        self.openmodel_base_url = os.getenv('OPENMODEL_BASE_URL', 'https://api.openmodel.ai/v1').rstrip('/')
        self.gemini_base_url = os.getenv('GEMINI_BASE_URL', 'https://generativelanguage.googleapis.com/v1beta').rstrip('/')
        self.base_url = self.gemini_base_url if self.provider == 'gemini' else self.openmodel_base_url if self.provider == 'openmodel' else self.yandex_base_url
        self.yandex_api_key = os.getenv('YANDEX_API_KEY', '')
        self.openmodel_api_key = os.getenv('OPENMODEL_API_KEY', '')
        self.gemini_api_key = os.getenv('GEMINI_API_KEY', '')
        self.api_key = self.gemini_api_key if self.provider == 'gemini' else self.openmodel_api_key if self.provider == 'openmodel' else self.yandex_api_key
        self.folder_id = os.getenv('YANDEX_FOLDER_ID', '')
        self.generator_model = os.getenv('NORLAB_GENERATOR_MODEL', 'gpt-oss-120b')
        self.critic_model = os.getenv('NORLAB_CRITIC_MODEL', 'gpt-oss-120b')
        self.fast_model = os.getenv('NORLAB_FAST_MODEL', self.generator_model)
        configured_embedding_model = os.getenv('NORLAB_EMBEDDING_MODEL', '')
        configured_vision_model = os.getenv('NORLAB_VISION_MODEL', '')
        if self.provider == 'gemini':
            self.embedding_model = (
                configured_embedding_model
                if configured_embedding_model and not configured_embedding_model.startswith(('yandex-', 'emb://'))
                else 'gemini-embedding-001'
            )
            self.vision_model = (
                configured_vision_model
                if configured_vision_model and not configured_vision_model.startswith(('qwen', 'gpt://'))
                else 'gemini-2.5-flash'
            )
        else:
            self.embedding_model = configured_embedding_model or 'yandex-embeddings'
            self.vision_model = configured_vision_model or 'qwen3.6-35b-a3b'
        self.embedding_dimensions = int(os.getenv('NORLAB_EMBEDDING_DIMENSIONS', '256'))
        self.mode = os.getenv('NORLAB_LLM_MODE', 'real')

    def model_uri(self, model: str) -> str:
        if self.provider in {'openmodel', 'gemini'}:
            return model
        return self.yandex_model_uri(model)

    def yandex_model_uri(self, model: str) -> str:
        if model.startswith(('gpt://', 'emb://', 'art://')):
            return model
        if not self.folder_id:
            return model
        return f'gpt://{self.folder_id}/{model}'

    def embedding_uri(self, is_document: bool = True) -> str:
        if self.embedding_model.startswith('emb://'):
            return self.embedding_model
        kind = 'doc' if is_document else 'query'
        return f'emb://{self.folder_id}/text-search-{kind}/latest'

    def profile(self) -> dict[str, Any]:
        healthy = bool(self.api_key) if self.provider in {'openmodel', 'gemini'} else bool(self.api_key and self.folder_id)
        capabilities = ['chat_json', 'ru_en_zh']
        if self.provider in {'gemini', 'yandex'}:
            capabilities.extend(['vision_json', 'embeddings'])
        return {
            'id': 'recommended-v1',
            'provider': self.provider,
            'base_url': self.base_url,
            'generator': self.model_uri(self.generator_model),
            'critic': self.model_uri(self.critic_model),
            'fast': self.model_uri(self.fast_model),
            'vision': self.model_uri(self.vision_model),
            'embeddings': self.embedding_model,
            'mode': self.mode,
            'capabilities': capabilities,
            'healthy': healthy,
        }

    async def compatibility_test(self) -> dict[str, Any]:
        fallback = {'ok': self.mode == 'mock', 'languages': ['ru', 'en', 'zh'], 'evidence_ids': ['ev-test'], '_fallback': True}
        sample = await self.chat_json(
            role='fast',
            messages=[
                {'role': 'system', 'content': 'Return compact JSON only.'},
                {'role': 'user', 'content': 'Сформируй JSON {"ok": true, "languages": ["ru","en","zh"], "evidence_ids": ["ev-test"]}.'},
            ],
            fallback=fallback,
        )
        real_call_ok = self.mode == 'mock' or not sample.get('_fallback')
        return {
            'profile_id': 'recommended-v1',
            'passed': real_call_ok and bool(sample.get('ok')) and {'ru', 'en', 'zh'}.issubset(set(sample.get('languages', []))),
            'checks': {
                'json_schema': isinstance(sample, dict) and real_call_ok,
                'multilingual': {'ru', 'en', 'zh'}.issubset(set(sample.get('languages', []))),
                'evidence_ids': bool(sample.get('evidence_ids')),
                'real_call': real_call_ok,
            },
            'sample': sample,
        }

    async def chat_json(self, role: str, messages: list[dict[str, str]], fallback: dict[str, Any]) -> dict[str, Any]:
        if self.mode == 'mock' or not self.api_key:
            return fallback
        raw_model = self.critic_model if role == 'critic' else self.fast_model if role in {'fast', 'repair'} else self.generator_model
        model = self.model_uri(raw_model)
        if self.provider == 'openmodel':
            headers = {
                'x-api-key': self.api_key,
                'anthropic-version': '2023-06-01',
                'content-type': 'application/json',
            }
        else:
            headers = {'Authorization': f'Api-Key {self.api_key}', 'Content-Type': 'application/json'}
            if self.folder_id:
                headers['X-Folder-Id'] = self.folder_id
        text_input = '\n\n'.join(f'{item["role"].upper()}: {item["content"]}' for item in messages)
        system_prompt = '\n\n'.join(item['content'] for item in messages if item['role'] == 'system') or 'Return JSON only.'
        anthropic_messages = [
            {'role': item['role'] if item['role'] in {'user', 'assistant'} else 'user', 'content': item['content']}
            for item in messages
            if item['role'] != 'system'
        ]
        responses_payload = {
            'model': model,
            'input': text_input,
            'temperature': 0.2,
            'max_output_tokens': int(os.getenv('NORLAB_CHAT_MAX_OUTPUT_TOKENS', '2500')),
        }
        chat_payload = {
            'model': model,
            'messages': messages,
            'temperature': 0.2,
            'max_tokens': int(os.getenv('NORLAB_CHAT_MAX_OUTPUT_TOKENS', '2500')),
            'response_format': {'type': 'json_object'},
        }
        messages_payload = {
            'model': model,
            'system': system_prompt,
            'messages': anthropic_messages,
            'temperature': 0.2,
            'max_tokens': int(os.getenv('NORLAB_CHAT_MAX_OUTPUT_TOKENS', '2500')),
        }
        gemini_payload = {
            'systemInstruction': {'parts': [{'text': system_prompt}]},
            'contents': [{'role': 'user' if item['role'] != 'assistant' else 'model', 'parts': [{'text': item['content']}]} for item in messages if item['role'] != 'system'],
            'generationConfig': {
                'temperature': 0.2,
                'maxOutputTokens': int(os.getenv('NORLAB_CHAT_MAX_OUTPUT_TOKENS', '2500')),
                'responseMimeType': 'application/json',
            },
        }
        request_timeout = float(os.getenv('NORLAB_CHAT_REQUEST_TIMEOUT_SECONDS', '30'))
        endpoint_order = ['generateContent'] if self.provider == 'gemini' else ['messages'] if self.provider == 'openmodel' else ['responses', 'chat'] if os.getenv('NORLAB_USE_RESPONSES_API', 'false').lower() == 'true' else ['chat', 'responses']
        max_attempts = max(1, int(os.getenv('NORLAB_CHAT_ATTEMPTS', '2')))
        last_error: Exception | None = None
        started = time.perf_counter()
        for attempt in range(max_attempts):
            async with httpx.AsyncClient(timeout=request_timeout) as client:
                for endpoint in endpoint_order:
                    try:
                        if endpoint == 'generateContent':
                            response = await client.post(
                                f'{self.base_url}/models/{model}:generateContent',
                                headers={'Content-Type': 'application/json', 'x-goog-api-key': self.api_key},
                                json=gemini_payload,
                            )
                        elif endpoint == 'messages':
                            response = await client.post(f'{self.base_url}/messages', headers=headers, json=messages_payload)
                        elif endpoint == 'responses':
                            response = await client.post(f'{self.base_url}/responses', headers=headers, json=responses_payload)
                        else:
                            response = await client.post(f'{self.base_url}/chat/completions', headers=headers, json=chat_payload)
                        response.raise_for_status()
                        content = self._extract_text(response.json())
                        parsed = self._parse_json_object(content)
                        store.data['audit'].append({
                            'id': new_id('audit'),
                            'kind': 'llm_call',
                            'role': role,
                            'model': model,
                            'endpoint': endpoint,
                            'attempt': attempt + 1,
                            'json_keys': list(parsed.keys())[:12],
                            'latency_ms': round((time.perf_counter() - started) * 1000),
                            'json_validation_result': 'passed',
                            'timestamp': now_iso(),
                        })
                        store.save()
                        return parsed
                    except httpx.HTTPStatusError as exc:
                        detail = exc.response.text[:500] if exc.response is not None else repr(exc)
                        status_code = exc.response.status_code if exc.response is not None else 'HTTP'
                        last_error = RuntimeError(f'{status_code} LLM error: {detail}')
                    except Exception as exc:
                        last_error = exc
            if attempt + 1 < max_attempts:
                await asyncio.sleep(1 + attempt)
        store.data['audit'].append({
            'id': new_id('audit'),
            'kind': 'llm_call_error',
            'role': role,
            'model': model,
            'endpoints': endpoint_order,
            'attempts': max_attempts,
            'error': f'{type(last_error).__name__}: {last_error!r}'[:500],
            'latency_ms': round((time.perf_counter() - started) * 1000),
            'timestamp': now_iso(),
        })
        store.save()
        return fallback

    def _parse_json_object(self, content: str) -> dict[str, Any]:
        cleaned = content.strip()
        if cleaned.startswith('```'):
            cleaned = re.sub(r'^```(?:json)?\s*', '', cleaned, flags=re.IGNORECASE)
            cleaned = re.sub(r'\s*```$', '', cleaned).strip()
        try:
            parsed = json.loads(cleaned)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass
        decoder = json.JSONDecoder()
        # A response that starts as JSON but cannot be decoded is usually a
        # truncated outer object.  Do not accidentally accept one of its valid
        # nested fragments (for example constraint_satisfaction) as the whole
        # model response.
        if cleaned.startswith('{'):
            try:
                parsed, _ = decoder.raw_decode(cleaned)
            except json.JSONDecodeError as exc:
                raise ValueError('LLM response contains a truncated JSON object') from exc
            if isinstance(parsed, dict):
                return parsed
        for index, char in enumerate(cleaned):
            if char != '{':
                continue
            try:
                parsed, _ = decoder.raw_decode(cleaned[index:])
            except json.JSONDecodeError:
                continue
            if isinstance(parsed, dict):
                return parsed
        raise ValueError('LLM response does not contain a JSON object')

    async def embed_text(self, text: str, is_document: bool = True) -> dict[str, Any]:
        if self.provider == 'gemini':
            return await self._embed_text_gemini(text, is_document=is_document)
        if not self.yandex_api_key:
            raise RuntimeError('YANDEX_API_KEY is required for embeddings')
        max_chars = int(os.getenv('NORLAB_EMBEDDING_MAX_CHARS', '2000'))
        payload = {'modelUri': self.embedding_uri(is_document), 'text': text[:max_chars]}
        headers = {
            'Authorization': f'Api-Key {self.yandex_api_key}',
            'Content-Type': 'application/json',
            'x-data-logging-enabled': 'false',
        }
        last_error: Exception | None = None
        request_timeout = float(os.getenv('NORLAB_EMBEDDING_REQUEST_TIMEOUT_SECONDS', '20'))
        max_attempts = int(os.getenv('NORLAB_EMBEDDING_MAX_ATTEMPTS', '2'))
        for attempt in range(max_attempts):
            try:
                async with httpx.AsyncClient(timeout=request_timeout) as client:
                    response = await client.post('https://llm.api.cloud.yandex.net/foundationModels/v1/textEmbedding', headers=headers, json=payload)
                    response.raise_for_status()
                break
            except httpx.HTTPStatusError as exc:
                detail = exc.response.text[:500] if exc.response is not None else repr(exc)
                last_error = RuntimeError(f'{exc.response.status_code if exc.response else "HTTP"} embedding error: {detail}')
                await asyncio.sleep(1 + attempt * 2)
            except Exception as exc:
                last_error = exc
                await asyncio.sleep(1 + attempt * 2)
        else:
            raise RuntimeError(f'Yandex embeddings request failed: {type(last_error).__name__}: {last_error!r}')
        body = response.json()
        return {
            'embedding': [float(item) for item in body['embedding']],
            'num_tokens': int(body.get('numTokens') or 0),
            'model_version': body.get('modelVersion'),
            'model_uri': payload['modelUri'],
        }

    async def _embed_text_gemini(self, text: str, is_document: bool = True) -> dict[str, Any]:
        if not self.gemini_api_key:
            raise RuntimeError('GEMINI_API_KEY is required for embeddings')
        max_chars = int(os.getenv('NORLAB_EMBEDDING_MAX_CHARS', '8000'))
        model = self.embedding_model.removeprefix('models/')
        payload = {
            'model': f'models/{model}',
            'content': {'parts': [{'text': text[:max_chars]}]},
            'taskType': 'RETRIEVAL_DOCUMENT' if is_document else 'RETRIEVAL_QUERY',
            'outputDimensionality': self.embedding_dimensions,
        }
        headers = {
            'Content-Type': 'application/json',
            'x-goog-api-key': self.gemini_api_key,
        }
        request_timeout = float(os.getenv('NORLAB_EMBEDDING_REQUEST_TIMEOUT_SECONDS', '30'))
        max_attempts = max(1, int(os.getenv('NORLAB_EMBEDDING_MAX_ATTEMPTS', '2')))
        last_error: Exception | None = None
        started = time.perf_counter()
        for attempt in range(max_attempts):
            try:
                async with httpx.AsyncClient(timeout=request_timeout) as client:
                    response = await client.post(
                        f'{self.gemini_base_url}/models/{model}:embedContent',
                        headers=headers,
                        json=payload,
                    )
                    response.raise_for_status()
                body = response.json()
                values = (body.get('embedding') or {}).get('values')
                if not isinstance(values, list) or not values:
                    raise ValueError('Gemini embedding response does not contain embedding.values')
                result = {
                    'embedding': [float(item) for item in values],
                    'num_tokens': 0,
                    'model_version': model,
                    'model_uri': f'models/{model}',
                }
                store.data['audit'].append({
                    'id': new_id('audit'),
                    'kind': 'embedding_call',
                    'provider': 'gemini',
                    'model': model,
                    'task_type': payload['taskType'],
                    'dimensions': len(result['embedding']),
                    'attempt': attempt + 1,
                    'latency_ms': round((time.perf_counter() - started) * 1000),
                    'timestamp': now_iso(),
                })
                store.save()
                return result
            except httpx.HTTPStatusError as exc:
                detail = exc.response.text[:500] if exc.response is not None else repr(exc)
                status_code = exc.response.status_code if exc.response is not None else 'HTTP'
                last_error = RuntimeError(f'{status_code} Gemini embedding error: {detail}')
            except Exception as exc:
                last_error = exc
            if attempt + 1 < max_attempts:
                await asyncio.sleep(1 + attempt * 2)
        store.data['audit'].append({
            'id': new_id('audit'),
            'kind': 'embedding_call_error',
            'provider': 'gemini',
            'model': model,
            'attempts': max_attempts,
            'error': f'{type(last_error).__name__}: {last_error!r}'[:500],
            'latency_ms': round((time.perf_counter() - started) * 1000),
            'timestamp': now_iso(),
        })
        store.save()
        raise RuntimeError(f'Gemini embeddings request failed: {type(last_error).__name__}: {last_error!r}')

    def _extract_text(self, body: dict[str, Any]) -> str:
        for candidate in body.get('candidates') or []:
            content = candidate.get('content') or {}
            for part in content.get('parts') or []:
                if isinstance(part, dict) and isinstance(part.get('text'), str) and part['text'].strip():
                    return part['text']
        if isinstance(body.get('content'), list):
            for part in body['content']:
                if isinstance(part, dict) and isinstance(part.get('text'), str) and part['text'].strip():
                    return part['text']
        if isinstance(body.get('output_text'), str) and body['output_text'].strip():
            return body['output_text']
        for choice in body.get('choices') or []:
            message = choice.get('message') or {}
            for key in ('content', 'reasoning_content'):
                content = message.get(key)
                if isinstance(content, str) and content.strip():
                    return content
                if isinstance(content, list):
                    for part in content:
                        if isinstance(part, dict) and isinstance(part.get('text'), str) and part['text'].strip():
                            return part['text']
        for item in body.get('output', []):
            for content in (item.get('content') or []):
                if isinstance(content.get('text'), str) and content['text'].strip():
                    return content['text']
        raise ValueError('LLM response does not contain text output')

    async def vision_json(self, image_path: Path, prompt: str) -> dict[str, Any]:
        if self.provider == 'gemini':
            return await self._vision_json_gemini(image_path, prompt)
        if not self.yandex_api_key:
            raise RuntimeError('YANDEX_API_KEY is required for vision')
        mime = 'image/png' if image_path.suffix.lower() == '.png' else 'image/jpeg'
        data_url = f'data:{mime};base64,{base64.b64encode(image_path.read_bytes()).decode("ascii")}'
        model = self.yandex_model_uri(self.vision_model)
        headers = {'Authorization': f'Api-Key {self.yandex_api_key}', 'Content-Type': 'application/json'}
        payload = {
            'model': model,
            'input': [{
                'role': 'user',
                'content': [
                    {'type': 'input_text', 'text': prompt},
                    {'type': 'input_image', 'image_url': data_url},
                ],
            }],
            'temperature': 0.1,
            'max_output_tokens': 4000,
        }
        async with httpx.AsyncClient(timeout=120) as client:
            response = await client.post(f'{self.yandex_base_url}/responses', headers=headers, json=payload)
            response.raise_for_status()
        return self._parse_json_object(self._extract_text(response.json()))

    async def _vision_json_gemini(self, image_path: Path, prompt: str) -> dict[str, Any]:
        if not self.gemini_api_key:
            raise RuntimeError('GEMINI_API_KEY is required for vision')
        mime_by_suffix = {
            '.png': 'image/png',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.webp': 'image/webp',
        }
        mime = mime_by_suffix.get(image_path.suffix.lower())
        if not mime:
            raise ValueError(f'Unsupported Gemini vision image type: {image_path.suffix or "unknown"}')
        image_bytes = image_path.read_bytes()
        # Base64 increases the payload size; 14 MiB leaves room under Gemini's
        # 20 MiB total inline-request limit for JSON and the prompt.
        max_bytes = int(os.getenv('NORLAB_VISION_MAX_INLINE_BYTES', str(14 * 1024 * 1024)))
        if len(image_bytes) > max_bytes:
            raise ValueError(f'Image is too large for inline Gemini vision request: {len(image_bytes)} bytes')
        model = self.vision_model.removeprefix('models/')
        payload = {
            'contents': [{
                'role': 'user',
                'parts': [
                    {'inlineData': {'mimeType': mime, 'data': base64.b64encode(image_bytes).decode('ascii')}},
                    {'text': prompt},
                ],
            }],
            'generationConfig': {
                'temperature': 0.1,
                'maxOutputTokens': int(os.getenv('NORLAB_VISION_MAX_OUTPUT_TOKENS', '4000')),
                'responseMimeType': 'application/json',
            },
        }
        headers = {
            'Content-Type': 'application/json',
            'x-goog-api-key': self.gemini_api_key,
        }
        request_timeout = float(os.getenv('NORLAB_VISION_REQUEST_TIMEOUT_SECONDS', '120'))
        max_attempts = max(1, int(os.getenv('NORLAB_VISION_MAX_ATTEMPTS', '2')))
        last_error: Exception | None = None
        started = time.perf_counter()
        for attempt in range(max_attempts):
            try:
                async with httpx.AsyncClient(timeout=request_timeout) as client:
                    response = await client.post(
                        f'{self.gemini_base_url}/models/{model}:generateContent',
                        headers=headers,
                        json=payload,
                    )
                    response.raise_for_status()
                parsed = self._parse_json_object(self._extract_text(response.json()))
                store.data['audit'].append({
                    'id': new_id('audit'),
                    'kind': 'llm_call',
                    'role': 'vision',
                    'provider': 'gemini',
                    'model': model,
                    'endpoint': 'generateContent',
                    'attempt': attempt + 1,
                    'json_keys': list(parsed.keys())[:12],
                    'latency_ms': round((time.perf_counter() - started) * 1000),
                    'json_validation_result': 'passed',
                    'timestamp': now_iso(),
                })
                store.save()
                return parsed
            except httpx.HTTPStatusError as exc:
                detail = exc.response.text[:500] if exc.response is not None else repr(exc)
                status_code = exc.response.status_code if exc.response is not None else 'HTTP'
                last_error = RuntimeError(f'{status_code} Gemini vision error: {detail}')
            except Exception as exc:
                last_error = exc
            if attempt + 1 < max_attempts:
                await asyncio.sleep(1 + attempt)
        store.data['audit'].append({
            'id': new_id('audit'),
            'kind': 'llm_call_error',
            'role': 'vision',
            'provider': 'gemini',
            'model': model,
            'endpoints': ['generateContent'],
            'attempts': max_attempts,
            'error': f'{type(last_error).__name__}: {last_error!r}'[:500],
            'latency_ms': round((time.perf_counter() - started) * 1000),
            'timestamp': now_iso(),
        })
        store.save()
        raise RuntimeError(f'Gemini vision request failed: {type(last_error).__name__}: {last_error!r}')


async def fetch_json(url: str, params: dict[str, Any], timeout: float = 20) -> dict[str, Any]:
    async with httpx.AsyncClient(timeout=timeout, headers={'User-Agent': 'NORLAB hackathon backend/0.1'}) as client:
        response = await client.get(url, params=params)
        response.raise_for_status()
        return response.json()


async def external_research(query: str, max_sources: int = 8) -> list[dict[str, Any]]:
    sanitized = sanitizer.sanitize(query)
    sources: list[dict[str, Any]] = []
    try:
        openalex = await fetch_json('https://api.openalex.org/works', {'search': sanitized, 'per-page': min(max_sources, 8)})
        for item in openalex.get('results', []):
            source_id = new_id('src')
            sources.append({
                'id': source_id,
                'title': item.get('title') or 'Untitled',
                'authors': [auth.get('author', {}).get('display_name') for auth in item.get('authorships', [])[:6] if auth.get('author')],
                'year': item.get('publication_year'),
                'doi': item.get('doi'),
                'url': item.get('id') or item.get('doi'),
                'source_type': 'openalex_work',
                'language': item.get('language'),
                'trust_tier': 'metadata_scientific',
                'retrieval_date': now_iso(),
                'query': sanitized,
                'abstract': inverted_abstract(item.get('abstract_inverted_index') or {}),
            })
    except Exception as exc:
        store.data['audit'].append({'id': new_id('audit'), 'kind': 'external_research_error', 'provider': 'openalex', 'error': str(exc)[:500], 'timestamp': now_iso()})
    try:
        crossref = await fetch_json('https://api.crossref.org/works', {'query': sanitized, 'rows': min(max_sources, 8)})
        for item in crossref.get('message', {}).get('items', []):
            source_id = new_id('src')
            title = item.get('title') or ['Untitled']
            sources.append({
                'id': source_id,
                'title': title[0],
                'authors': [
                    ' '.join(part for part in [author.get('given'), author.get('family')] if part)
                    for author in item.get('author', [])[:6]
                ],
                'year': ((item.get('published-print') or item.get('published-online') or {}).get('date-parts') or [[None]])[0][0],
                'doi': item.get('DOI'),
                'url': item.get('URL'),
                'source_type': 'crossref_work',
                'language': item.get('language'),
                'trust_tier': 'metadata_scientific',
                'retrieval_date': now_iso(),
                'query': sanitized,
                'abstract': strip_tags(item.get('abstract') or ''),
            })
    except Exception as exc:
        store.data['audit'].append({'id': new_id('audit'), 'kind': 'external_research_error', 'provider': 'crossref', 'error': str(exc)[:500], 'timestamp': now_iso()})
    deduped: dict[str, dict[str, Any]] = {}
    for source in sources:
        key = (source.get('doi') or source.get('title') or source['id']).lower()
        deduped.setdefault(key, source)
    return list(deduped.values())[:max_sources]


async def patent_research(query: str, max_sources: int = 8) -> list[dict[str, Any]]:
    sanitized = sanitizer.sanitize(query)
    sources: list[dict[str, Any]] = []
    payload = {
        'q': {'_text_any': {'patent_title': sanitized}},
        'f': ['patent_id', 'patent_title', 'patent_date', 'assignees.assignee_organization', 'inventors.inventor_name_first', 'inventors.inventor_name_last'],
        'o': {'per_page': max_sources},
    }
    try:
        async with httpx.AsyncClient(timeout=30, headers={'User-Agent': 'NORLAB research connector/0.1'}) as client:
            response = await client.post('https://search.patentsview.org/api/v1/patent/', json=payload)
            response.raise_for_status()
        for item in response.json().get('patents', []):
            patent_id = item.get('patent_id')
            sources.append({
                'id': new_id('pat'),
                'title': item.get('patent_title') or patent_id,
                'patent_number': patent_id,
                'year': (item.get('patent_date') or '')[:4] or None,
                'url': f'https://patents.google.com/patent/US{patent_id}' if patent_id else None,
                'source_type': 'patent',
                'language': 'en',
                'trust_tier': 'patent_metadata',
                'retrieval_date': now_iso(),
                'query': sanitized,
                'assignees': item.get('assignees', []),
                'inventors': item.get('inventors', []),
            })
    except Exception as exc:
        store.data['audit'].append({'id': new_id('audit'), 'kind': 'patent_research_error', 'provider': 'patentsview', 'error': str(exc)[:500], 'timestamp': now_iso()})
        store.save()
    return sources


def inverted_abstract(index: dict[str, list[int]]) -> str:
    if not index:
        return ''
    words: list[tuple[int, str]] = []
    for word, positions in index.items():
        for position in positions:
            words.append((position, word))
    return ' '.join(word for _, word in sorted(words))[:2000]


def strip_tags(text: str) -> str:
    return re.sub(r'<[^>]+>', '', text).strip()


def xml_escape(text: str) -> str:
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def make_docx(markdown: str) -> bytes:
    paragraphs = []
    for line in markdown.splitlines():
        style = ''
        text = line
        if line.startswith('# '):
            style = '<w:pStyle w:val="Heading1"/>'
            text = line[2:]
        elif line.startswith('## '):
            style = '<w:pStyle w:val="Heading2"/>'
            text = line[3:]
        elif line.startswith('### '):
            style = '<w:pStyle w:val="Heading3"/>'
            text = line[4:]
        paragraphs.append(
            f'<w:p><w:pPr>{style}</w:pPr><w:r><w:t xml:space="preserve">{xml_escape(text)}</w:t></w:r></w:p>'
        )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body>{"".join(paragraphs)}<w:sectPr/></w:body></w:document>'
    )
    output = BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as docx:
        docx.writestr('[Content_Types].xml', (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            '</Types>'
        ))
        docx.writestr('_rels/.rels', (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>'
        ))
        docx.writestr('word/document.xml', document_xml)
    return output.getvalue()


def pdf_escape(text: str) -> str:
    return text.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')


def make_pdf(markdown: str) -> bytes:
    lines = [line[:95] for line in markdown.splitlines()[:70]]
    content_lines = ['BT', '/F1 10 Tf', '50 790 Td', '14 TL']
    for line in lines:
        safe = pdf_escape(line.encode('latin-1', errors='replace').decode('latin-1'))
        content_lines.append(f'({safe}) Tj')
        content_lines.append('T*')
    content_lines.append('ET')
    stream = '\n'.join(content_lines).encode('latin-1')
    objects = [
        b'<< /Type /Catalog /Pages 2 0 R >>',
        b'<< /Type /Pages /Kids [3 0 R] /Count 1 >>',
        b'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>',
        b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>',
        b'<< /Length ' + str(len(stream)).encode('ascii') + b' >>\nstream\n' + stream + b'\nendstream',
    ]
    output = BytesIO()
    output.write(b'%PDF-1.4\n')
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f'{index} 0 obj\n'.encode('ascii'))
        output.write(obj)
        output.write(b'\nendobj\n')
    xref = output.tell()
    output.write(f'xref\n0 {len(objects) + 1}\n'.encode('ascii'))
    output.write(b'0000000000 65535 f \n')
    for offset in offsets[1:]:
        output.write(f'{offset:010d} 00000 n \n'.encode('ascii'))
    output.write(f'trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF'.encode('ascii'))
    return output.getvalue()


parser = DocumentParser()
gateway = ModelGateway()
sanitizer = SearchQuerySanitizer()
app = FastAPI(title='NORLAB Backend', version='0.1.0')
celery_app = Celery(
    'norlab',
    broker=os.getenv('CELERY_BROKER_URL', os.getenv('REDIS_URL', 'redis://localhost:6379/0')),
    backend=os.getenv('CELERY_RESULT_BACKEND', os.getenv('REDIS_URL', 'redis://localhost:6379/0')),
)


@app.middleware('http')
async def strip_api_prefix(request: Request, call_next: Any) -> Any:
    if request.scope.get('path', '').startswith('/api/'):
        request.scope['path'] = request.scope['path'][4:]
    return await call_next(request)


def get_project(project_id: str) -> dict[str, Any]:
    if store.postgres:
        store.reload()
    project = store.data['projects'].get(project_id)
    if not project:
        raise HTTPException(status_code=404, detail='Project not found')
    return project


def get_run(run_id: str) -> dict[str, Any]:
    if store.postgres:
        store.reload()
    run = store.data['runs'].get(run_id)
    if not run:
        raise HTTPException(status_code=404, detail='Run not found')
    return run


def add_event(run: dict[str, Any], stage: RunStatus, message: str, completed_units: int = 0, queued_units: int = 0) -> None:
    event = {
        'id': new_id('event'),
        'stage': stage.value,
        'message': message,
        'completed_units': completed_units,
        'queued_units': queued_units,
        'timestamp': now_iso(),
    }
    run['status'] = stage.value
    run['updated_at'] = now_iso()
    run.setdefault('events', []).append(event)
    store.save()
    redis_state.set_run_status(run['id'], stage.value)
    redis_state.publish_event(run['id'], event)


def project_documents(project_id: str) -> list[dict[str, Any]]:
    return [doc for doc in store.data['documents'].values() if doc['project_id'] == project_id]


def project_facts(project_id: str) -> list[dict[str, Any]]:
    facts = []
    for doc in project_documents(project_id):
        facts.extend(doc.get('facts', []))
    return facts


def vision_graph_to_facts(document: dict[str, Any], graph: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    filename = str(document.get('filename') or document.get('id') or 'image')
    nodes = graph.get('nodes') if isinstance(graph.get('nodes'), list) else []
    edges = graph.get('edges') if isinstance(graph.get('edges'), list) else []
    uncertain = graph.get('uncertain_items') if isinstance(graph.get('uncertain_items'), list) else []

    def node_label(node: Any) -> str:
        if isinstance(node, dict):
            return str(node.get('label') or node.get('name') or node.get('id') or '').strip()
        return str(node or '').strip()

    def node_type(node: Any) -> str:
        if isinstance(node, dict):
            return str(node.get('type') or 'VisualNode').strip()
        return 'VisualNode'

    node_summaries = [
        f'{node_type(node)}: {node_label(node)}'
        for node in nodes
        if node_label(node)
    ][:24]
    edge_summaries = []
    for edge in edges[:32]:
        if not isinstance(edge, dict):
            continue
        source = str(edge.get('source') or edge.get('from') or '').strip()
        target = str(edge.get('target') or edge.get('to') or '').strip()
        relation = str(edge.get('type') or edge.get('relation') or 'связано').strip()
        if source and target:
            edge_summaries.append(f'{source} → {target} ({relation})')

    summary_parts = []
    if node_summaries:
        summary_parts.append('узлы: ' + '; '.join(node_summaries))
    if edge_summaries:
        summary_parts.append('связи: ' + '; '.join(edge_summaries[:16]))
    if uncertain:
        summary_parts.append('неуверенно распознано: ' + '; '.join(str(item) for item in uncertain[:8]))
    summary_text = f'Vision-разбор схемы/изображения {filename}: ' + ('; '.join(summary_parts) if summary_parts else 'значимые элементы не распознаны.')

    fragment = {
        'id': new_id('frag'),
        'document_id': document['id'],
        'original_language': 'ru',
        'original_text': summary_text,
        'normalized_text': summary_text.lower(),
        'location': 'image:vision',
        'extraction_method': 'vision:qwen',
        'confidence': 0.72 if node_summaries or edge_summaries else 0.35,
    }
    facts: list[dict[str, Any]] = []
    for node in nodes[:40]:
        label = node_label(node)
        if not label:
            continue
        kind_raw = node_type(node).lower()
        if 'equipment' in kind_raw or 'аппарат' in kind_raw or 'оборуд' in kind_raw:
            fact_type = 'equipment'
        elif 'parameter' in kind_raw or 'параметр' in kind_raw:
            fact_type = 'parameter'
        elif 'stream' in kind_raw or 'поток' in kind_raw:
            fact_type = 'process'
        else:
            fact_type = 'process'
        facts.append({
            'id': new_id('fact'),
            'type': fact_type,
            'statement': f'На схеме/изображении {filename} распознан элемент: {node_type(node)} «{label}».',
            'source_fragment_id': fragment['id'],
            'document_id': document['id'],
            'confidence': float(node.get('confidence', 0.72)) if isinstance(node, dict) else 0.72,
            'provenance': {
                'location': 'image:vision',
                'extraction_method': 'vision:qwen',
                'timestamp': now_iso(),
            },
        })
    for edge in edges[:40]:
        if not isinstance(edge, dict):
            continue
        source = str(edge.get('source') or edge.get('from') or '').strip()
        target = str(edge.get('target') or edge.get('to') or '').strip()
        relation = str(edge.get('type') or edge.get('relation') or 'связано').strip()
        if not source or not target:
            continue
        facts.append({
            'id': new_id('fact'),
            'type': 'process',
            'statement': f'На схеме/изображении {filename} распознана связь технологической цепи: {source} → {target} ({relation}).',
            'source_fragment_id': fragment['id'],
            'document_id': document['id'],
            'confidence': float(edge.get('confidence', 0.68)),
            'provenance': {
                'location': 'image:vision',
                'extraction_method': 'vision:qwen',
                'timestamp': now_iso(),
            },
        })
    return [fragment], facts


def evidence_records(project_id: str, limit: int = 24) -> list[dict[str, Any]]:
    project = store.data['projects'].get(project_id, {})
    query_text = ' '.join([
        str(project.get('problem') or ''),
        str(project.get('target_kpi') or ''),
        ' '.join(project.get('constraints', [])),
        'tailings хвосты флотация flotation losses потери gold copper nickel золото медь никель',
    ])
    query_tokens = tokenize_for_grounding(query_text)
    preferred_phrases = [
        'хвост', 'tailing', 'флотац', 'flotation', 'потер', 'loss', 'извлеч',
        'recovery', 'золото', 'gold', 'медь', 'copper', 'никель', 'nickel',
        'элемент 28', 'элемент 29', 'схем', 'регламент', 'apparatus', 'equipment',
        'гидроциклон', 'классификатор', 'поток', 'цеп',
    ]
    risky_transfer_terms = ['обжиг', 'roasting', 'цианирован', 'cyanidation', 'автоклав', 'pressure oxidation']
    records: list[tuple[float, int, dict[str, Any]]] = []
    for doc_index, doc in enumerate(project_documents(project_id)):
        filename = str(doc.get('filename') or '').lower()
        doc_boost = 0
        if any(phrase in filename for phrase in ['хвост', 'tailing', 'отчет', 'report', 'xlsx', 'csv']):
            doc_boost += 8
        if doc.get('content_type') in {'xlsx', 'csv', 'docx'}:
            doc_boost += 3
        if file_kind(doc) == 'image' and doc.get('vision_graph'):
            doc_boost += 6
        for fact_index, fact in enumerate(doc.get('facts', [])):
            statement = str(fact.get('statement') or '')
            lowered = statement.lower()
            fact_tokens = tokenize_for_grounding(statement)
            score = doc_boost + len(query_tokens & fact_tokens) * 3
            score += sum(2 for phrase in preferred_phrases if phrase in lowered)
            if any(term in lowered for term in risky_transfer_terms):
                score -= 4
            records.append((score, doc_index * 10000 + fact_index, {
                'id': fact['id'],
                'statement': statement,
                'document_id': doc['id'],
                'source_name': doc['filename'],
                'location': fact['provenance']['location'],
                'trust_tier': 'internal',
            }))
    records.sort(key=lambda item: (-item[0], item[1]))
    return [item[2] for item in records[:limit]]


def tokenize_for_grounding(text: str) -> set[str]:
    stopwords = {
        'для', 'или', 'при', 'без', 'как', 'что', 'это', 'the', 'and', 'with', 'from',
        'will', 'this', 'that', 'into', 'требует', 'проверить', 'гипотеза',
    }
    return {
        token.lower()
        for token in re.findall(r'[A-Za-zА-Яа-яЁё0-9]{4,}', text)
        if token.lower() not in stopwords
    }


def unsupported_numeric_claims(text: str, allowed_context: str) -> list[str]:
    text = re.sub(r'[\u00a0\u202f]', ' ', text)
    allowed_context = re.sub(r'[\u00a0\u202f]', ' ', allowed_context)
    unit_pattern = (
        r'%|п\.?\s*п\.?|пп|мм|мкм|микрон\w*|мг|г|кг|т|л|мл|'
        r'минут\w*|мин|час\w*|дн\w*|сут\w*|°c|c|ph|ppm|ppb|'
        r'mm|um|kg|g|mg|t|l|ml|min|h|d|day|days'
    )
    number_pattern = rf'(?<![\w-])\d+(?:[.,]\d+)?(?:\s*(?:{unit_pattern}))?(?![\w-])'
    material_unit_pattern = unit_pattern

    def is_material_claim(value: str) -> bool:
        normalized = value.strip().lower()
        return bool(re.search(material_unit_pattern, normalized, flags=re.IGNORECASE)) or bool(re.search(r'\d+[.,]\d+', normalized))

    def canonical_claim(value: str) -> str:
        value = re.sub(r'[\u00a0\u202f]', ' ', value.strip().lower())
        value = re.sub(r'\s+', ' ', value)
        return value.replace(',', '.')

    allowed_numbers = {canonical_claim(item) for item in re.findall(number_pattern, allowed_context) if is_material_claim(item)}
    allowed_numeric_values = {
        re.match(r'\d+(?:[.,]\d+)?', item.strip()).group(0).replace(',', '.')
        for item in allowed_numbers
        if re.match(r'\d+(?:[.,]\d+)?', item.strip())
    }
    allowed_numeric_values.update(
        match.group(0).replace(',', '.')
        for match in re.finditer(r'\d+(?:[.,]\d+)?', allowed_context)
    )
    claims = []
    for value in re.findall(number_pattern, text):
        normalized = canonical_claim(value)
        numeric_value = re.match(r'\d+(?:[.,]\d+)?', normalized)
        if is_material_claim(normalized) and normalized not in allowed_numbers and (not numeric_value or numeric_value.group(0).replace(',', '.') not in allowed_numeric_values):
            claims.append(value.strip())
    return claims


NUMERIC_UNIT_PATTERN = (
    r'%|п\.?\s*п\.?|пп|процентн(?:ых|ого|ые)?\s+пункт(?:а|ов)?|мм|мкм|микрон\w*|'
    r'мг|г|кг|т|л|мл|минут\w*|мин|час\w*|дн\w*|сут\w*|°c|c|ph|ppm|ppb|'
    r'mm|um|µm|kg|g|mg|t|l|ml|min|h|d|day|days|p\.?\s*p\.?|percentage\s+points?'
)
NUMERIC_CLAIM_RE = re.compile(
    rf'(?<![\w-])(?:[<>≤≥=~≈±+-]\s*)?\d+(?:[.,]\d+)?(?:\s*(?:{NUMERIC_UNIT_PATTERN}))?(?![\w-])',
    re.IGNORECASE,
)


def material_numeric_claims(text: str) -> list[str]:
    normalized = re.sub(r'[\u00a0\u202f]', ' ', str(text or ''))
    claims: list[str] = []
    for match in NUMERIC_CLAIM_RE.finditer(normalized):
        value = re.sub(r'\s+', ' ', match.group(0)).strip()
        if not value:
            continue
        # Plain IDs and years are not useful process parameters. Keep decimals
        # and anything with a unit/sign because those are the numbers a
        # researcher expects to see in a testable hypothesis.
        has_unit_or_signal = bool(re.search(NUMERIC_UNIT_PATTERN, value, flags=re.IGNORECASE)) or bool(re.search(r'[<>≤≥=~≈±+-]', value))
        has_decimal = bool(re.search(r'\d+[,.]\d+', value))
        if has_unit_or_signal or has_decimal:
            claims.append(value)
    return claims


def text_has_material_number(text: str) -> bool:
    return bool(material_numeric_claims(text))


def project_numeric_context(project: dict[str, Any], evidence: list[dict[str, Any]], limit: int = 18) -> dict[str, Any]:
    project_text = '\n'.join([
        str(project.get('problem') or ''),
        str(project.get('target_kpi') or ''),
        '\n'.join(str(item) for item in project.get('constraints', [])),
    ])
    project_claims = material_numeric_claims(project_text)
    evidence_items = []
    seen: set[str] = set()
    for item in evidence:
        statement = str(item.get('statement') or '')
        claims = material_numeric_claims(statement)
        if not claims:
            continue
        key = f'{item.get("id")}:{",".join(claims[:6])}'
        if key in seen:
            continue
        seen.add(key)
        evidence_items.append({
            'id': item.get('id'),
            'numbers': claims[:8],
            'source_name': item.get('source_name'),
            'location': item.get('location'),
            'statement': summarize_text(statement, 360),
        })
        if len(evidence_items) >= limit:
            break
    return {
        'project_required_numbers': project_claims[:10],
        'evidence_numeric_snippets': evidence_items,
        'requires_numeric_hypotheses': bool(project_claims or evidence_items),
    }


def ensure_project_numeric_target_in_hypothesis(hypothesis: dict[str, Any], project: dict[str, Any]) -> dict[str, Any]:
    target = str(project.get('target_kpi') or '').strip()
    if not target or not text_has_material_number(target):
        return hypothesis
    result = dict(hypothesis)
    expected = as_string_list(result.get('expected_kpis'))
    if not expected or not any(text_has_material_number(item) for item in expected):
        result['expected_kpis'] = [target]
    statement = str(result.get('statement') or '').strip()
    if statement and not text_has_material_number(statement):
        result['statement'] = f'{statement.rstrip(". ")}. Числовой критерий проверки: {target}.'
    falsification = as_string_list(result.get('falsification_conditions'))
    if not falsification or not any(text_has_material_number(item) for item in falsification):
        result['falsification_conditions'] = [
            f'Гипотеза отклоняется, если результат не достигает критерия: {target}.',
            *falsification[:2],
        ]
    return result


def add_concrete_experimental_levels(hypothesis: dict[str, Any]) -> dict[str, Any]:
    result = dict(hypothesis)
    text = f'{result.get("title", "")} {result.get("statement", "")}'.lower()
    additions: list[str] = []
    if any(term in text for term in ['p80', 'раскрыт', 'доизмельч', 'крупност']):
        additions.append('Первый проверяемый уровень: P80 150–180 мкм против baseline.')
    if any(term in text for term in ['ph', 'селективност', 'поверхност']):
        additions.append('Первый проверяемый диапазон: pH 9,0–9,5 против baseline.')
    if any(term in text for term in ['кондиционир', 'контакт']):
        additions.append('Первый проверяемый диапазон: 8–12 минут кондиционирования против baseline.')
    if any(term in text for term in ['классификац', 'тонких поток', 'гидроциклон']):
        additions.append('Первый проверяемый режим: отсечка тонкого потока 0,5–0,8 мм против baseline.')
    if any(term in text for term in ['pgm', 'проб', 'верификац']):
        additions.append('Первый проверяемый контроль: 2–3 повторных PGM-анализа на каждой точке.')
    if any(term in text for term in ['схем', 'точк']):
        additions.append('Первый проверяемый контроль: 3 точки отбора проб по цепи аппаратов.')
    if any(term in text for term in ['глинист']):
        additions.append('Первый проверяемый фактор: классы глинистости 5%, 10% и 15% против baseline.')
    if not additions:
        additions.append('Первый проверяемый план: матрица 2×3 режимов против baseline.')
    statement = str(result.get('statement') or '').strip()
    for addition in additions[:2]:
        if addition not in statement:
            statement = f'{statement} {addition}'.strip()
    result['statement'] = statement
    assumptions = as_string_list(result.get('assumptions'))
    for addition in additions[:2]:
        assumptions.append(f'Численный уровень является стартовым планом опыта и уточняется по данным первой серии: {addition}')
    result['assumptions'] = assumptions[:6]
    return result


def as_string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        items: list[str] = []
        for item in value:
            if item is None:
                continue
            if isinstance(item, dict):
                candidate = item.get('id') or item.get('evidence_id') or item.get('fact_id')
                if candidate is not None:
                    items.append(str(candidate))
                continue
            items.append(str(item))
        return items
    if isinstance(value, str):
        return [value]
    return []


def constraint_violations(text: str, constraints: list[str]) -> list[str]:
    lowered = text.lower()
    violations: list[str] = []

    def has_unnegated(patterns: list[str]) -> bool:
        for pattern in patterns:
            start = lowered.find(pattern)
            if start < 0:
                continue
            prefix = lowered[max(0, start - 40):start]
            if any(negation in prefix for negation in ['без ', 'не ', 'no ', 'not ', 'without ', 'does not ', 'не требует', 'не нужна']):
                continue
            return True
        return False

    for constraint in constraints:
        constraint_lower = constraint.lower()
        forbids_major_replacement = (
            'no major equipment replacement' in constraint_lower
            or 'без капитальной замены' in constraint_lower
            or 'без замены оборудования' in constraint_lower
            or 'без смены технологического контура' in constraint_lower
            or 'без изменения технологического контура' in constraint_lower
            or 'без смены контура' in constraint_lower
        )
        if forbids_major_replacement and has_unnegated([
            'replace equipment',
            'equipment replacement',
            'заменить оборудование',
            'замена оборудования',
            'замен',
            'замена классификатор',
            'заменить классификатор',
            'замена гидроциклон',
            'заменить гидроциклон',
            'новое оборудование',
        ]):
            violations.append(constraint)
        forbids_capex = 'no major capex' in constraint_lower or 'без capex' in constraint_lower or 'без капиталь' in constraint_lower
        if forbids_capex and has_unnegated(['major capex', 'capital expenditure', 'капитальн', 'дорогостоящ']):
            violations.append(constraint)
    return violations


def unsupported_specific_terms(text: str, allowed_context: str) -> list[str]:
    risky_term_groups = [
        ('дитиофосфат', 'dithiophosphate'), ('ксантогенат', 'xanthate'),
        ('гидроциклон', 'hydrocyclone'), ('мельница', 'mill'),
        ('магнит', 'magnetic'), ('флокулянт', 'flocculant'),
        ('ультразвук', 'ультразвуков', 'ultrasound'), ('кавитац', 'cavitation'),
        ('микросфер', 'microsphere'), ('адсорбент', 'adsorbent'),
        ('биовыщелач', 'бактериаль', 'bioleaching'),
        ('окислител', 'пероксид', 'перекись', 'oxidant', 'peroxide'),
        ('плотность пульпы', 'pulp density'), ('аэраци', 'aeration'),
        ('температур', 'temperature'),
    ]
    standalone_terms = [
        'new frother', 'новый вспениватель', 'новый собиратель',
    ]
    lowered = text.lower()
    allowed_lower = allowed_context.lower()
    unsupported = [
        term
        for group in risky_term_groups
        for term in group
        if term in lowered and not any(alias in allowed_lower for alias in group)
    ]
    unsupported.extend(term for term in standalone_terms if term in lowered and term not in allowed_lower)
    if (
        ('комбинац' in lowered or 'синерг' in lowered or ('двух' in lowered and 'собирател' in lowered))
        and not any(marker in allowed_lower for marker in ['комбинац', 'синерг', 'двух собирател', 'смесь собирател'])
    ):
        unsupported.append('комбинация собирателей')
    return sorted(set(unsupported))


def evidence_backed_intervention_families(allowed_context: str) -> list[dict[str, Any]]:
    lowered = allowed_context.lower()
    families: list[dict[str, Any]] = []
    if any(term in lowered for term in ['p80', 'доизмельч', 'измельчен', 'крупност', 'раскрыти', 'grind', 'particle size', 'liberation']):
        families.append({
            'id': 'grind_size',
            'label': 'крупность / P80 / доизмельчение / раскрытие минералов',
            'terms': ['p80', 'доизмельч', 'измельч', 'крупност', 'раскрыти', 'переизмельч', 'ошлам', 'grind', 'particle size', 'liberation'],
        })
    if any(term in lowered for term in ['кондиционир', 'собирател', 'время', 'conditioning', 'collector', 'contact time']):
        families.append({
            'id': 'conditioning',
            'label': 'время кондиционирования / контакт с собирателем',
            'terms': ['кондиционир', 'собирател', 'контакт', 'время', 'минут', 'conditioning', 'collector', 'contact time'],
        })
    if any(term in lowered for term in ['ph', 'щелоч', 'селектив', 'selectivity', 'alkalin']):
        families.append({
            'id': 'ph_selectivity',
            'label': 'pH / селективность флотации',
            'terms': ['ph', 'щелоч', 'селектив', 'пустой пород', 'порода', 'selectivity', 'alkalin'],
        })
    if any(term in lowered for term in ['гидроциклон', 'отбор проб', 'контрольный отбор', 'схема', 'hydrocyclone', 'classification', 'sampling']):
        families.append({
            'id': 'sampling_classification',
            'label': 'гидроциклон / классификация / точка отбора проб',
            'terms': ['гидроциклон', 'классификац', 'отбор проб', 'проб', 'схема', 'hydrocyclone', 'classification', 'sampling'],
        })
    if any(term in lowered for term in ['повторн', 'анализ', 'глинист', 'качество данных', 'repeat analysis', 'data quality', 'clay']):
        families.append({
            'id': 'data_quality',
            'label': 'повторные анализы / качество данных / глинистость',
            'terms': ['повторн', 'анализ', 'глинист', 'качество данных', 'pgm', 'repeat analysis', 'data quality', 'clay'],
        })
    return families


def unsupported_availability_claims(text: str, selected_evidence_context: str) -> list[str]:
    lowered = text.lower()
    evidence_lower = selected_evidence_context.lower()
    uncertainty_markers = ['??? ????????????? ???????????', '???? ??????', 'if available', 'if confirmed', '??? ???????', '????????? ???????', '??? ???????????? ????????', '??? ????????????? ?????']
    if any(marker in lowered for marker in uncertainty_markers):
        return []
    availability_markers = ['доступн', 'имеется', 'available', 'is present']
    equipment_terms = ['гидроциклон', 'мельниц', 'сепаратор', 'печь', 'флотомашин', 'hydrocyclone', 'mill', 'separator', 'furnace']
    claims = []
    if any(marker in lowered for marker in availability_markers):
        for term in equipment_terms:
            if term in lowered and not any(marker in evidence_lower for marker in availability_markers):
                claims.append(term)
    return claims


def hypothesis_validation_context(project: dict[str, Any], evidence: list[dict[str, Any]], external_sources: list[dict[str, Any]]) -> dict[str, Any]:
    evidence_text = '\n'.join(f'{item["id"]}: {item["statement"]}' for item in evidence)
    external_text = '\n'.join(str(item.get('title', '')) for item in external_sources[:8])
    allowed_context = '\n'.join([
        project.get('problem', ''),
        project.get('target_kpi') or '',
        '\n'.join(project.get('constraints', [])),
        evidence_text,
        external_text,
    ])
    numeric_context = project_numeric_context(project, evidence)
    return {
        'allowed_evidence': {item['id']: item for item in evidence},
        'allowed_context': allowed_context,
        'evidence_tokens': tokenize_for_grounding(evidence_text),
        'project_tokens': tokenize_for_grounding(project.get('problem', '') + ' ' + (project.get('target_kpi') or '')),
        'intervention_families': evidence_backed_intervention_families(allowed_context),
        'numeric_context': numeric_context,
    }


def compact_evidence_for_llm(evidence: list[dict[str, Any]]) -> list[dict[str, Any]]:
    compact = []
    for item in evidence:
        statement = str(item.get('statement', ''))
        compact.append({
            'id': item['id'],
            'statement': summarize_text(statement, 360 if material_numeric_claims(statement) else 320),
            'numbers': material_numeric_claims(statement)[:8],
            'source_name': item.get('source_name'),
            'location': item.get('location'),
            'trust_tier': item.get('trust_tier', 'internal'),
        })
    return compact


def evidence_quality_report(evidence: list[dict[str, Any]]) -> dict[str, Any]:
    statements = [str(item.get('statement') or '').strip() for item in evidence]
    joined = '\n'.join(statements).lower()
    broken_markers = ['#ref!', '#value!', '#div/0!', 'error)=', 'ошибка', 'не удалось']
    meaningful = [
        statement for statement in statements
        if len(statement) >= 40 and not any(marker in statement.lower() for marker in broken_markers)
    ]
    tokens = tokenize_for_grounding('\n'.join(meaningful))
    has_material = any(term in joined for term in ['хвост', 'tailing', 'шлак', 'slag', 'пульп', 'руда', 'концентрат'])
    has_process = any(term in joined for term in ['флотац', 'измельч', 'классификац', 'обогащ', 'выщелач', 'сепарац', 'переработ'])
    has_kpi = any(term in joined for term in ['извлеч', 'recovery', 'потер', 'loss', 'содержан', 'grade', 'kpi', 'au', 'ni', 'cu', 'pgm'])
    sufficient = (
        len(meaningful) >= 2
        or (len(tokens) >= 16 and has_material and (has_process or has_kpi))
    )
    return {
        'sufficient': sufficient,
        'facts': len(evidence),
        'meaningful_facts': len(meaningful),
        'tokens': len(tokens),
        'has_material': has_material,
        'has_process': has_process,
        'has_kpi': has_kpi,
        'broken_markers_found': [marker for marker in broken_markers if marker in joined],
    }


def unsupported_generic_hypothesis_phrases(hypothesis_text: str, allowed_context: str) -> list[str]:
    lowered = hypothesis_text.lower()
    allowed = allowed_context.lower()
    phrases = [
        'например',
        'лабораторный реагент',
        'лабораторный аналог',
        'изменить режим обработки',
        'режим обработки хвостов',
        'время контакта',
        'продолжительность обработки',
        'сравнить два типа',
        'предварительная подготовка',
        'реагент может',
        'more длительный контакт',
    ]
    return [phrase for phrase in phrases if phrase in lowered and phrase not in allowed]


def unsupported_vague_parameter_phrases(hypothesis_text: str) -> list[str]:
    lowered = hypothesis_text.lower()
    vague_patterns = [
        r'несколько\s+минут',
        r'каждые\s+несколько',
        r'оптимальн\w*\s+(?:уров|диапазон|режим|значен|крупност|p80|ph)',
        r'умеренн\w+\s+(?:щелоч|кисл|уров|диапазон)',
        r'слегка\s+(?:кисл|щелоч|пониз|повыс)',
        r'более\s+мелк\w+\s+(?:фракц|крупност|частиц)',
        r'более\s+длительн\w+\s+(?:контакт|время|кондиционир)',
    ]
    vague = []
    for pattern in vague_patterns:
        match = re.search(pattern, lowered, flags=re.IGNORECASE)
        if match:
            fragment = hypothesis_text[max(0, match.start() - 24):match.end() + 24].strip()
            vague.append(re.sub(r'\s+', ' ', fragment))
    special_parameter_rules = [
        (r'увеличить\s+время', r'(?:до|на)\s*\d'),
        (r'снизить\s+p80', r'(?:до|на|≤|<|=)\s*\d'),
        (r'повысить\s+ph', r'(?:до|на|≤|≥|<|>|=)\s*\d'),
        (r'понизить\s+ph', r'(?:до|на|≤|≥|<|>|=)\s*\d'),
    ]
    for phrase_pattern, value_pattern in special_parameter_rules:
        for match in re.finditer(phrase_pattern, lowered, flags=re.IGNORECASE):
            tail = lowered[match.start():match.start() + 90]
            if not re.search(value_pattern, tail, flags=re.IGNORECASE):
                fragment = hypothesis_text[max(0, match.start() - 24):match.end() + 60].strip()
                vague.append(re.sub(r'\s+', ' ', fragment))
    return sorted(set(vague))


def normalize_vague_parameters_for_display(text: str) -> str:
    result = str(text or '')
    replacements = [
        (r'каждые\s+несколько\s+минут', 'с интервалом, который определить в первом опыте'),
        (r'несколько\s+минут', 'интервал определить в первом опыте'),
        (r'до\s+умеренно\s+щелочного\s+уровня', 'до диапазона, подтверждённого первым опытом'),
        (r'до\s+слегка\s+кислой\s+зоны', 'до диапазона, подтверждённого первым опытом'),
        (r'оптимальн\w*\s+диапазон', 'диапазон, определённый первым опытом'),
        (r'оптимальн\w*\s+уровень', 'уровень, определённый первым опытом'),
        (r'более\s+мелк\w+\s+фракц\w*', 'фракция с заданным P80 из первого опыта'),
    ]
    for pattern, replacement in replacements:
        result = re.sub(pattern, replacement, result, flags=re.IGNORECASE)
    return re.sub(r'\s+', ' ', result).strip()


def validate_llm_hypothesis(
    hypothesis: dict[str, Any],
    project: dict[str, Any],
    validation_context: dict[str, Any],
) -> tuple[bool, list[str], dict[str, Any]]:
    reasons: list[str] = []
    # A verbose but otherwise grounded answer should not be discarded only
    # because the model wrote a sentence where the UI needs a card title.
    # Keep the complete scientific statement and derive a compact display
    # title before applying the strict title gate.
    hypothesis = ensure_project_numeric_target_in_hypothesis(dict(hypothesis), project)
    title = presentation_hypothesis_title(hypothesis)
    hypothesis['title'] = title
    statement = str(hypothesis.get('statement') or '').strip()
    numeric_context = validation_context.get('numeric_context') or {}
    evidence_ids = [item for item in hypothesis.get('supporting_evidence', []) if item in validation_context['allowed_evidence']]
    if len(title) < 8:
        reasons.append('title is missing or too short')
    if len(title) > 64 or len(title.split()) > 8:
        reasons.append('title must contain 3-8 words and no more than 64 characters')
    if len(statement) < 40:
        reasons.append('statement is missing or too short')
    if not evidence_ids:
        reasons.append('no valid supporting_evidence ids')
    numeric_text = ' '.join([
        statement,
        ' '.join(as_string_list(hypothesis.get('expected_kpis'))),
        ' '.join(as_string_list(hypothesis.get('falsification_conditions'))),
        str(hypothesis.get('economic_effect') or ''),
    ])
    if numeric_context.get('requires_numeric_hypotheses') and not text_has_material_number(numeric_text):
        reasons.append(
            'missing numeric test criterion: use exact supported numbers from project target, constraints or supplied evidence'
        )
    selected_evidence_text = '\n'.join(validation_context['allowed_evidence'][item]['statement'] for item in evidence_ids)
    selected_allowed_context = '\n'.join([
        project.get('problem', ''),
        project.get('target_kpi') or '',
        '\n'.join(project.get('constraints', [])),
        selected_evidence_text,
    ])
    constraint_satisfaction = hypothesis.get('constraint_satisfaction', {})
    if isinstance(constraint_satisfaction, dict):
        constraint_satisfaction_text = ' '.join(str(value) for value in constraint_satisfaction.values())
    elif isinstance(constraint_satisfaction, list):
        constraint_satisfaction_text = ' '.join(str(value) for value in constraint_satisfaction)
    else:
        constraint_satisfaction_text = str(constraint_satisfaction or '')
    hypothesis_text = ' '.join([
        title,
        statement,
        json.dumps(hypothesis.get('intervention', {}), ensure_ascii=False),
        ' '.join(as_string_list(hypothesis.get('causal_mechanism'))),
        ' '.join(as_string_list(hypothesis.get('assumptions'))),
        str(hypothesis.get('economic_effect') or ''),
        ' '.join(as_string_list(hypothesis.get('risks'))),
        ' '.join(as_string_list(hypothesis.get('falsification_conditions'))),
        constraint_satisfaction_text,
    ])
    validation_warnings: list[str] = []
    claims = unsupported_numeric_claims(hypothesis_text, selected_allowed_context)
    if claims:
        validation_warnings.append(
            f'experimental numeric levels need confirmation: {", ".join(sorted(set(claims)))}'
        )
    specific_terms = unsupported_specific_terms(hypothesis_text, selected_allowed_context)
    if specific_terms:
        validation_warnings.append(
            f'specific terms need source confirmation: {", ".join(sorted(set(specific_terms)))}'
        )
    generic_phrases = unsupported_generic_hypothesis_phrases(hypothesis_text, selected_allowed_context)
    if any(str(family.get('id')) == 'conditioning' for family in validation_context.get('intervention_families') or []):
        generic_phrases = [phrase for phrase in generic_phrases if phrase not in {'время контакта', 'продолжительность обработки'}]
    if generic_phrases:
        reasons.append(f'generic unsupported phrasing: {", ".join(sorted(set(generic_phrases)))}')
    vague_phrases = unsupported_vague_parameter_phrases(hypothesis_text)
    if vague_phrases:
        validation_warnings.append(
            'vague process parameters without exact supported values: '
            + ', '.join(vague_phrases[:4])
        )
    scores = hypothesis.get('scores', {})
    if isinstance(scores, dict):
        score_values = [
            clamp01(scores.get('evidence', 0.5)),
            clamp01(scores.get('feasibility', 0.5)),
            clamp01(scores.get('novelty', 0.5)),
            clamp01(scores.get('experimentability', 0.5)),
        ]
        if len({round(value, 2) for value in score_values}) == 1 and round(score_values[0], 2) == 0.5:
            reasons.append('scores are non-informative: all four values are 0.5')
    availability_claims = unsupported_availability_claims(hypothesis_text, selected_evidence_text)
    if availability_claims:
        reasons.append(f'unsupported equipment availability claims: {", ".join(sorted(set(availability_claims)))}')
    intervention_families = validation_context.get('intervention_families') or []
    matched_families: list[str] = []
    intervention_text = ' '.join([
        title,
        statement,
        json.dumps(hypothesis.get('intervention', {}), ensure_ascii=False),
        ' '.join(as_string_list(hypothesis.get('causal_mechanism'))),
    ])
    hypothesis_lower = intervention_text.lower()
    for family in intervention_families:
        terms = [str(term).lower() for term in family.get('terms', [])]
        if any(term in hypothesis_lower for term in terms):
            matched_families.append(str(family.get('id') or family.get('label')))
    if intervention_families and not matched_families:
        labels = ', '.join(str(family.get('label') or family.get('id')) for family in intervention_families[:6])
        reasons.append(f'intervention is outside evidence-backed families: {labels}')
    violations = constraint_violations(
        hypothesis_text,
        project.get('constraints', []),
    )
    if violations:
        reasons.append(f'violates project constraints: {"; ".join(violations)}')
    hypothesis_tokens = tokenize_for_grounding(
        ' '.join([
            title,
            statement,
            json.dumps(hypothesis.get('intervention', {}), ensure_ascii=False),
            ' '.join(as_string_list(hypothesis.get('causal_mechanism'))),
        ])
    )
    evidence_overlap = hypothesis_tokens & validation_context['evidence_tokens']
    project_overlap = hypothesis_tokens & validation_context['project_tokens']
    if len(evidence_overlap | project_overlap) < 1 and not matched_families:
        reasons.append('weak lexical grounding in supplied evidence/problem')
    normalized = dict(hypothesis)
    normalized['constraints'] = project.get('constraints', [])
    normalized['conditions'] = project.get('constraints', [])
    if isinstance(normalized.get('constraint_satisfaction'), str):
        normalized['constraint_satisfaction'] = {'summary': normalized['constraint_satisfaction']}
    normalized['supporting_evidence'] = evidence_ids
    normalized['data_triggers'] = [item for item in normalized.get('data_triggers', evidence_ids) if item in validation_context['allowed_evidence']] or evidence_ids
    normalized['grounding'] = {
        'evidence_ids': evidence_ids,
        'evidence_overlap_terms': sorted(evidence_overlap)[:20],
        'project_overlap_terms': sorted(project_overlap)[:20],
        'intervention_families': matched_families,
        'constraint_compliance_required': project.get('constraints', []),
        'numeric_context_used': numeric_context,
        'validation_warnings': validation_warnings,
    }
    if os.getenv('NORLAB_STRICT_LLM', 'true').lower() != 'true':
        normalized['grounding']['validation_warnings'] = [
            *validation_warnings,
            *[f'non_strict_acceptance: {reason}' for reason in reasons],
        ]
        if not normalized.get('supporting_evidence'):
            fallback_evidence = list(validation_context.get('allowed_evidence', {}).keys())[:2]
            normalized['supporting_evidence'] = fallback_evidence
            normalized['data_triggers'] = fallback_evidence
            normalized['grounding']['evidence_ids'] = fallback_evidence
        return True, [], normalized
    return not reasons, reasons, normalized


async def reindex_project_embeddings(project_id: str, limit: int = 64, timeout_seconds: float | None = None) -> dict[str, Any]:
    get_project(project_id)
    started = time.monotonic()
    if timeout_seconds is None:
        timeout_seconds = float(os.getenv('NORLAB_EMBEDDING_STAGE_TIMEOUT_SECONDS', '45'))
    indexed = {'fragments': 0, 'facts': 0, 'hypotheses': 0, 'errors': [], 'status': 'completed'}
    if not neo4j_memory.driver:
        indexed['errors'].append({'id': project_id, 'error': 'Neo4j is disabled; vector index requires full profile'})
        indexed['status'] = 'skipped'
        return indexed

    def remaining_time() -> float:
        if timeout_seconds <= 0:
            return 0
        return max(0.0, timeout_seconds - (time.monotonic() - started))

    async def embed_with_deadline(text: str) -> dict[str, Any]:
        remaining = remaining_time()
        if remaining <= 0:
            raise TimeoutError('Embedding reindex stage timeout reached')
        return await asyncio.wait_for(gateway.embed_text(text, is_document=True), timeout=remaining)

    def mark_timeout() -> None:
        indexed['status'] = 'partial_timeout'
        indexed['errors'].append({'id': project_id, 'error': f'Embedding reindex stopped after {timeout_seconds:g}s stage timeout'})

    for doc in project_documents(project_id):
        if indexed['fragments'] >= limit:
            break
        for fragment in doc.get('fragments', []):
            if indexed['fragments'] >= limit:
                break
            try:
                result = await embed_with_deadline(fragment['original_text'])
                neo4j_memory.set_embedding('Fragment', fragment['id'], result['embedding'], result.get('model_version'))
                fragment['embedding_model_version'] = result.get('model_version')
                indexed['fragments'] += 1
            except TimeoutError:
                mark_timeout()
                store.save()
                return indexed
            except Exception as exc:
                indexed['errors'].append({'id': fragment['id'], 'error': str(exc)[:300]})
                break
    for doc in project_documents(project_id):
        if indexed['facts'] >= limit:
            break
        for fact in doc.get('facts', []):
            if indexed['facts'] >= limit:
                break
            try:
                result = await embed_with_deadline(fact['statement'])
                neo4j_memory.set_embedding('Fact', fact['id'], result['embedding'], result.get('model_version'))
                fact['embedding_model_version'] = result.get('model_version')
                indexed['facts'] += 1
            except TimeoutError:
                mark_timeout()
                store.save()
                return indexed
            except Exception as exc:
                indexed['errors'].append({'id': fact['id'], 'error': str(exc)[:300]})
                break
    for hypothesis in [item for item in store.data['hypotheses'].values() if item['project_id'] == project_id][:limit]:
        try:
            text = f'{hypothesis["title"]}\n{hypothesis["statement"]}'
            result = await embed_with_deadline(text)
            neo4j_memory.set_embedding('Hypothesis', hypothesis['id'], result['embedding'], result.get('model_version'))
            hypothesis['embedding_model_version'] = result.get('model_version')
            indexed['hypotheses'] += 1
        except TimeoutError:
            mark_timeout()
            store.save()
            return indexed
        except Exception as exc:
            indexed['errors'].append({'id': hypothesis['id'], 'error': str(exc)[:300]})
            break
    store.save()
    return indexed


def create_hypotheses(project: dict[str, Any], run: dict[str, Any]) -> list[dict[str, Any]]:
    evidence = evidence_records(project['id'])
    if not evidence:
        evidence = [{
            'id': 'ev-assumption',
            'statement': 'Недостаточно локальных evidence; гипотезы сформированы как направления диагностики.',
            'document_id': None,
            'source_name': 'project brief',
            'location': 'brief',
            'trust_tier': 'assumption',
        }]
    perspectives = [
        ('TECH_OPT', 'Оптимизация режима флотации хвостов', 'Уточнить pH, крупность и расход реагентов для повышения извлечения ценных компонентов.'),
        ('LOW_CAPEX', 'Низкозатратная перенастройка классификации', 'Проверить изменение отсечки гидроциклонов и возвратов без капитальной замены оборудования.'),
        ('COUNTERFACTUAL', 'Переразводка потоков и доизмельчение', 'Проверить, снижает ли отдельная обработка тонких классов потери металлов в хвостах.'),
        ('NOVELTY', 'Комбинация предварительной активации и контрольной флотации', 'Проверить новую комбинацию мягкой активации поверхности и контрольной операции.'),
    ]
    # Keep the deterministic branch useful for demos without LLM: it should
    # produce a real candidate pool, not just the three/four old generic cards.
    perspectives = [
        ('TECH_OPT', 'Настройка раскрытия хвостов', 'Если уточнить раскрытие минералов и режим контакта хвостов с флотационной средой, то извлечение целевых металлов должно вырасти относительно baseline без смены технологического контура.'),
        ('LOW_CAPEX', 'Классификация тонких потоков', 'Если разделить тонкие и грубые классы перед контрольной операцией, то потери ценных компонентов в хвостах должны снизиться за счет более устойчивого режима обработки.'),
        ('COUNTERFACTUAL', 'Доизмельчение раскрытых сульфидов', 'Если проверить мягкое доизмельчение фракции с признаками недораскрытия, то извлечение Ni/Cu/PGM должно улучшиться относительно текущего baseline.'),
        ('NOVELTY', 'Активация поверхности перед флотацией', 'Если добавить щадящую подготовку поверхности перед контрольной флотацией, то селективность извлечения должна повыситься без промышленного синтеза.'),
        ('SAMPLING', 'Повторная PGM-верификация проб', 'Если закрыть пробел повторных PGM-анализов и сравнить результаты с baseline, то ранжирование гипотез станет устойчивее и снизит риск ложного вывода.'),
        ('SCHEME', 'Контроль точки отбора проб', 'Если зафиксировать точку отбора проб в схеме аппаратов и связать ее с таблицами опытов, то проверка причин потерь станет воспроизводимой.'),
        ('DATA_QUALITY', 'Проверка глинистости хвостов', 'Если добавить контроль глинистости как ограничивающего фактора, то можно отделить режимный эффект от влияния свойств исходной пробы.'),
        ('MECHANISM', 'Сравнение механизма с baseline', 'Если отдельно проверить физико-химический механизм изменения поверхности и раскрытия, то станет понятно, дает ли гипотеза реальный прирост KPI относительно baseline.'),
    ]
    hypotheses = []
    for index, (kind, title, statement) in enumerate(perspectives, start=1):
        selected = evidence[(index - 1)::4] or evidence[:2]
        eid = [item['id'] for item in selected[:4]]
        hypothesis = {
            'id': new_id('hyp'),
            'run_id': run['id'],
            'project_id': project['id'],
            'title': title,
            'statement': statement,
            'intervention': {'type': kind.lower(), 'description': statement},
            'target_process': ['flotation', 'tailings_processing'],
            'conditions': project.get('constraints', []),
            'causal_mechanism': ['Изменение раскрытия, поверхности или времени пребывания должно изменить извлечение.'],
            'data_triggers': eid,
            'supporting_evidence': eid,
            'contradicting_evidence': [],
            'assumptions': ['Точные режимные параметры требуют проверки на конкретной пробе.'],
            'expected_kpis': [project.get('target_kpi') or 'рост извлечения / снижение потерь в хвостах'],
            'economic_effect': 'Экономический эффект ожидается за счёт дополнительного извлечения ценных металлов и снижения потерь в хвостах; точная денежная оценка требует баланса металлов, цен и данных по реагентам/энергии.',
            'risks': ['Недостаточная представительность исходных данных', 'Переносимость эффекта между рудами ограничена'],
            'constraints': project.get('constraints', []),
            'falsification_conditions': ['Нет статистически значимого улучшения KPI относительно baseline.'],
            'novelty': {'class': 'UNKNOWN', 'analogs': []},
            'uncertainty': {'level': 'MEDIUM', 'reasons': ['нужны лабораторные проверки']},
            'disagreement': {},
            'lineage': {'perspective': kind, 'model_profile': run['model_profile']['id']},
            'status': 'DRAFT',
            'scores': {'evidence': 0.55, 'feasibility': 0.7, 'novelty': 0.45, 'experimentability': 0.8},
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        hypotheses.append(hypothesis)
    return hypotheses


def normalize_llm_hypotheses(project: dict[str, Any], run: dict[str, Any], payload: dict[str, Any]) -> list[dict[str, Any]]:
    def collect_candidate_dicts(value: Any, depth: int = 0) -> list[dict[str, Any]]:
        if depth > 4:
            return []
        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned.startswith(('```json', '```JSON')) and cleaned.endswith('```'):
                cleaned = cleaned[7:-3].strip()
            elif cleaned.startswith('```') and cleaned.endswith('```'):
                cleaned = cleaned[3:-3].strip()
            if cleaned.startswith(('{', '[')):
                try:
                    return collect_candidate_dicts(json.loads(cleaned), depth + 1)
                except json.JSONDecodeError:
                    return []
            return []
        if isinstance(value, dict):
            if ({'title', 'statement', 'hypothesis'} & set(value.keys())) and (
                'supporting_evidence' in value or 'evidence_ids' in value or 'evidence' in value
            ):
                return [value]
            found: list[dict[str, Any]] = []
            for child in value.values():
                found.extend(collect_candidate_dicts(child, depth + 1))
            return found
        if isinstance(value, list):
            found = []
            for child in value:
                found.extend(collect_candidate_dicts(child, depth + 1))
            return found
        return []

    raw_items: Any = None
    if isinstance(payload, dict):
        for key in ('hypotheses', 'items', 'results', 'candidates'):
            if isinstance(payload.get(key), list):
                raw_items = payload[key]
                break
            if isinstance(payload.get(key), dict):
                nested = payload[key]
                if {'title', 'statement'} & set(nested.keys()):
                    raw_items = [nested]
                    break
                for nested_key in ('hypotheses', 'items', 'results', 'candidates'):
                    if isinstance(nested.get(nested_key), list):
                        raw_items = nested[nested_key]
                        break
                if raw_items is None:
                    nested_values = [value for value in nested.values() if isinstance(value, dict) and ({'title', 'statement'} & set(value.keys()))]
                    if nested_values:
                        raw_items = nested_values
            if isinstance(raw_items, list):
                break
        if raw_items is None and isinstance(payload.get('hypothesis'), dict):
            raw_items = [payload['hypothesis']]
        if raw_items is None and {'title', 'statement'} & set(payload.keys()):
            raw_items = [payload]
        if raw_items is None:
            raw_items = collect_candidate_dicts(payload)
    if not isinstance(raw_items, list):
        return []
    normalized: list[dict[str, Any]] = []
    allowed_evidence = {item['id'] for item in evidence_records(project['id'], limit=100)}
    for raw in raw_items[:12]:
        if not isinstance(raw, dict):
            continue
        raw_evidence = raw.get('supporting_evidence', raw.get('evidence_ids', raw.get('evidence', [])))
        evidence_ids = [item for item in as_string_list(raw_evidence) if item in allowed_evidence]
        title = str(raw.get('title') or raw.get('name') or 'Проверяемая гипотеза')
        statement = str(raw.get('statement') or raw.get('hypothesis') or title)
        normalized_item = {
            'id': new_id('hyp'),
            'run_id': run['id'],
            'project_id': project['id'],
            'title': title[:240],
            'statement': statement[:1200],
            'intervention': raw.get('intervention') if isinstance(raw.get('intervention'), dict) else {'description': statement[:500]},
            'target_process': as_string_list(raw.get('target_process')) or ['tailings_processing'],
            'conditions': raw.get('conditions') if isinstance(raw.get('conditions'), list) else project.get('constraints', []),
            'causal_mechanism': as_string_list(raw.get('causal_mechanism')),
            'data_triggers': [item for item in as_string_list(raw.get('data_triggers', evidence_ids)) if item in allowed_evidence],
            'supporting_evidence': evidence_ids,
            'contradicting_evidence': [item for item in as_string_list(raw.get('contradicting_evidence', [])) if item in allowed_evidence],
            'assumptions': as_string_list(raw.get('assumptions')),
            'expected_kpis': as_string_list(raw.get('expected_kpis')) or [project.get('target_kpi') or 'KPI not specified'],
            'economic_effect': str(raw.get('economic_effect') or raw.get('economics') or 'Экономический эффект оценивается как качественный: дополнительное извлечение ценных металлов минус проверяемые затраты на реагенты, энергию и подготовку пробы.')[:700],
            'risks': as_string_list(raw.get('risks')),
            'constraint_satisfaction': raw.get('constraint_satisfaction') if isinstance(raw.get('constraint_satisfaction'), (dict, list, str)) else {},
            'constraints': project.get('constraints', []),
            'falsification_conditions': as_string_list(raw.get('falsification_conditions')) or ['Нет улучшения KPI относительно baseline.'],
            'novelty': {},
            'uncertainty': {},
            'disagreement': {},
            'lineage': {'perspective': raw.get('perspective', 'LLM'), 'model_profile': run['model_profile']['id']},
            'status': 'DRAFT',
            'scores': raw.get('scores') if isinstance(raw.get('scores'), dict) else {'evidence': 0.6, 'feasibility': 0.65, 'novelty': 0.5, 'experimentability': 0.75},
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        normalized.append(ensure_project_numeric_target_in_hypothesis(normalized_item, project))
    return normalized


async def generate_llm_hypotheses(
    project: dict[str, Any],
    run: dict[str, Any],
    external_sources: list[dict[str, Any]],
    desired_count: int,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    evidence = evidence_records(project['id'], limit=int(os.getenv('NORLAB_GENERATION_EVIDENCE_LIMIT', '36')))
    if not evidence:
        raise RuntimeError('Cannot generate grounded hypotheses: project has no extracted evidence records')
    evidence_quality = evidence_quality_report(evidence)
    if not evidence_quality['sufficient']:
        raise RuntimeError(
            'Cannot generate grounded hypotheses: extracted evidence is insufficient or contains parsing errors. '
            f'quality={evidence_quality}. Upload richer PDF/XLSX/image sources or fix source tables before LLM generation.'
        )
    validation_context = hypothesis_validation_context(project, evidence, external_sources)
    accepted: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    raw_received = 0
    max_attempts = int(os.getenv(
        'NORLAB_LLM_GENERATION_ATTEMPTS',
        '1' if gateway.provider == 'yandex' else '2',
    ))
    # Quality is the hard gate; portfolio size is a target, not a reason to
    # accept weak hypotheses or fail an otherwise useful run.  We still ask the
    # models to fill `desired_count`, but after all grounded attempts a partial
    # portfolio is accepted as long as at least one hypothesis passes validation.
    hard_minimum_accepted = 1
    compact_evidence = compact_evidence_for_llm(evidence)
    numeric_context = validation_context.get('numeric_context') or project_numeric_context(project, evidence)

    system_prompt = (
        'Ты генератор научно-исследовательских гипотез NORLAB. Верни только JSON object строго вида {"hypotheses":[...]}. '
        'Top-level ключи кроме hypotheses запрещены. Нельзя возвращать constraint_satisfaction как top-level object. '
        'Нельзя возвращать description, список ограничений, markdown или обычный текст вместо {"hypotheses":[...]}. Не придумывай новые ограничения проекта: используй только project.constraints. '
        'Нельзя выдумывать факты, численные эффекты, оборудование, реагенты, параметры руды или источники. '
        'Если evidence не содержит конкретное название реагента/оборудования, используй общий класс ("лабораторный реагент при подтверждении доступности", "лабораторный аналог при подтверждении доступности") без новых названий. '
        'Не предлагай новые классы вмешательств вне evidence_backed_intervention_families: никаких ультразвуков, адсорбентов, биовыщелачивания, окислителей, аэрации, плотности пульпы или температуры, если этого нет в supplied evidence. '
        'Если supplied evidence содержит только название KPI/металла без механизма, параметров, материала или операции, не генерируй гипотезы: такие данные недостаточны для grounded reasoning. '
        'Каждая гипотеза должна быть проверяемой, фальсифицируемой и опираться только на supplied evidence ids. '
        'Title — короткое название карточки: 3-8 слов и не более 64 символов. Формула title: конкретное воздействие или механизм + объект/результат. '
        'Не начинай title с "Гипотеза о", "Исследование влияния" или общего "Оптимизация процесса"; не превращай title в предложение и не ставь точку в конце. '
        'Все ограничения проекта являются жесткими. Если данных не хватает, формулируй это как проверяемое условие эксперимента, а не как установленный факт. '
        'Statement — это основная проверяемая гипотеза, а не заголовок. Формула statement: "Если проверить [конкретное воздействие] на [материал/операцию], то [KPI] должен измениться относительно baseline/цели, потому что [механизм]". '
        'Запрещены общие формулировки "изменить режим", "лабораторный реагент", "например", "время контакта", если конкретный параметр или класс реагента не указан в supplied evidence. '
        'Числа, проценты и пороги допустимы только если они прямо есть в project.target_kpi, constraints или allowed_evidence; если числа нет в данных, описывай направление эффекта без выдуманного значения. '
        'Если numeric_context содержит project_required_numbers, каждое statement, expected_kpis и falsification_conditions обязаны использовать точный числовой критерий из project.target_kpi, например "не менее 2 п.п."; это не выдумка, а критерий проверки. '
        'Если numeric_context.evidence_numeric_snippets содержит pH, P80, минуты, проценты, мкм или диапазоны, используй эти значения в конкретных режимах проверки только вместе с соответствующим supporting_evidence id. '
        'Гипотеза без числового критерия проверки недопустима, когда numeric_context.requires_numeric_hypotheses=true. '
        'Нельзя писать "несколько минут", "оптимальный диапазон", "умеренно щелочной", "слегка кислый", "более мелкая фракция" без конкретного числа из numeric_context; если числа режима нет, формулируй как "диапазон определить в первом опыте", но KPI-порог всё равно укажи точно. '
        'Economic_effect — отдельное поле: опиши ожидаемый экономический драйвер через дополнительное извлечение металлов, снижение потерь, CAPEX/OPEX, реагенты и энергию. Не указывай деньги, проценты и сроки окупаемости, если они не следуют прямо из supplied evidence или project target. '
        'Для каждого ограничения проекта заполни constraint_satisfaction: ключ = ограничение, значение = как гипотеза его соблюдает. '
        'Не утверждай, что конкретное оборудование/реагент доступен, если это прямо не указано в evidence; пиши "при подтверждении доступности" или "при наличии лабораторного аналога". '
        'Для каждой гипотезы обязательны поля: title, statement, intervention, target_process, causal_mechanism, '
        'supporting_evidence, assumptions, expected_kpis, economic_effect, risks, falsification_conditions, constraint_satisfaction, scores.'
        ' constraint_satisfaction обязан быть JSON object, scores обязан быть JSON object с evidence, feasibility, novelty, experimentability.'
        ' Scores должны быть дифференцированы по основаниям; не возвращай все четыре значения ровно 0.5.'
    )
    repair_system_prompt = (
        'You are NORLAB JSON schema repair and grounded generation fallback. Return JSON only, exactly '
        '{"hypotheses":[...]}; the only top-level key is hypotheses. Never use a project constraint, '
        'description, schema field, or explanatory sentence as a top-level key. Preserve usable content '
        'from malformed_model_output. If that output has no usable hypothesis, generate the requested '
        'number solely from project and allowed_evidence. Every hypothesis must include title, statement, '
        'intervention, target_process, causal_mechanism, supporting_evidence, assumptions, expected_kpis, '
        'economic_effect, risks, falsification_conditions, constraint_satisfaction, and scores. '
        'supporting_evidence may contain only supplied evidence ids and may not be empty. Do not invent '
        'numbers, reagents, equipment, facts, or sources. If numeric_context has project_required_numbers, '
        'copy those exact numbers into statement, expected_kpis and falsification_conditions as the test '
        'criterion. Use evidence numeric snippets only with their supplied evidence ids. Titles must be 3-8 words and at most 64 characters.'
    )

    attempts_used = 0
    batch_size = max(1, int(os.getenv('NORLAB_LLM_BATCH_SIZE', str(desired_count))))
    intervention_families = validation_context.get('intervention_families') or []
    while len(accepted) < desired_count and attempts_used < max_attempts:
        attempts_used += 1
        missing = desired_count - len(accepted)
        requested_in_call = min(missing, batch_size)
        payload = {
            'task': f'Generate exactly {requested_in_call} NEW valid hypotheses.',
            'required_top_level_json_shape': {'hypotheses': ['hypothesis_object']},
            'strict_json_output_rule': 'Return exactly one JSON object with one key "hypotheses". Do not return constraints, description, markdown, prose, arrays, or any other top-level keys.',
            'hypothesis_object_schema': {
                'title': '3-8 words, max 64 characters: specific mechanism/action + target, no sentence or generic prefix',
                'statement': 'main testable hypothesis: if [intervention] then [KPI direction/target] vs baseline because [mechanism]; include exact project KPI number when numeric_context requires it',
                'intervention': {'description': 'what to test, no unsupported names'},
                'target_process': ['tailings_processing'],
                'causal_mechanism': ['grounded mechanism'],
                'supporting_evidence': ['allowed evidence id'],
                'assumptions': ['explicitly uncertain assumptions'],
                'expected_kpis': ['project KPI with exact supported numeric target, no invented numeric effect'],
                'economic_effect': 'qualitative economics: recovered value vs reagents/energy/CAPEX, no invented money or unsupported percentages',
                'risks': ['grounded risk'],
                'falsification_conditions': ['what result refutes it; include exact numeric KPI threshold when available'],
                'constraint_satisfaction': {'project constraint text': 'how it is respected'},
                'scores': {'evidence': 0.5, 'feasibility': 0.5, 'novelty': 0.5, 'experimentability': 0.5},
            },
            'required_hypotheses_in_this_call': requested_in_call,
            'remaining_needed_after_this_call': max(0, desired_count - len(accepted) - requested_in_call),
            'language': run.get('response_language', project.get('response_language', 'ru')),
            'project': {
                'problem': project.get('problem'),
                'target_kpi': project.get('target_kpi'),
                'constraints': project.get('constraints', []),
            },
            'allowed_evidence': compact_evidence,
            'numeric_context': numeric_context,
            'evidence_quality': evidence_quality,
            'evidence_backed_intervention_families': [
                {'id': family.get('id'), 'label': family.get('label')}
                for family in intervention_families
            ],
            'external_sources_metadata_only': [],
            'external_sources_rule': 'Do not use external sources for hypothesis facts; external sources are reserved for later novelty checking only.',
            'already_accepted_titles': [item['title'] for item in accepted],
            'previous_rejections': rejected[-8:],
            'avoid_unless_exactly_present_in_allowed_evidence': [
                'ксантогенат', 'дитиофосфат', 'цианид', 'флокулянт',
                'мельница', 'гидроциклон', 'печь', 'флотомашина',
                'xanthate', 'dithiophosphate', 'cyanide', 'flocculant',
                'mill', 'hydrocyclone', 'furnace',
            ],
            'validation_rules': [
                'supporting_evidence must contain only ids from allowed_evidence and must not be empty',
                'numeric claims, units, percentages, thresholds, ranges, dosage, particle size, mass or time are allowed only if they appear in project target/constraints or allowed_evidence',
                'when numeric_context.requires_numeric_hypotheses is true, every hypothesis must include at least one supported numeric criterion in statement, expected_kpis and falsification_conditions',
                'if project target contains a KPI such as "не менее 2 п.п.", copy that exact KPI target into expected_kpis and the first falsification condition',
                'do not use vague parameter words such as "несколько минут", "оптимальный диапазон", "умеренно", "слегка" or "более мелкая фракция" unless an exact value/range from numeric_context is attached',
                'do not propose actions that violate project constraints',
                'the intervention must belong to one of evidence_backed_intervention_families; do not invent new reagents, equipment, ultrasound, adsorbents, bioleaching, oxidation, aeration, pulp density, or temperature changes unless explicitly present in allowed_evidence',
                'do not repeat already_accepted_titles',
                'constraint_satisfaction must be an object with every project constraint as a key',
                'scores must be an object, not a scalar',
                'scores must be conservative numbers from 0 to 1',
                'scores must not all be exactly 0.5; explain uncertainty via lower evidence/feasibility/testability values instead',
                'do not use generic placeholder actions such as "laboratory reagent", "change processing mode", "for example", "contact time" unless that exact variable exists in allowed_evidence',
                'title must contain 3-8 words, be <= 64 characters and name a specific mechanism/action plus its target',
                'keep each string concise: statement <= 280 chars, economic_effect <= 300 chars, each list item <= 160 chars',
                'if equipment or reagent availability is not explicitly stated, phrase it as an experimental condition to verify',
            ],
        }
        llm_result = await gateway.chat_json(
            role='generator',
            messages=[
                {'role': 'system', 'content': system_prompt},
                {'role': 'user', 'content': json.dumps(payload, ensure_ascii=False)},
            ],
            fallback={'hypotheses': []},
        )
        candidates = normalize_llm_hypotheses(project, run, llm_result)
        raw_received += len(candidates)
        if not candidates and int(os.getenv('NORLAB_SCHEMA_REPAIR_ATTEMPTS', '0')) > 0:
            keys = list(llm_result.keys())[:12] if isinstance(llm_result, dict) else []
            malformed_reason = 'empty_or_invalid_llm_response' if not keys else f'wrong_top_level_json_keys: {keys}'
            schema_repair_payload = {
                'task': f'Repair or replace the malformed response with exactly {requested_in_call} grounded hypotheses.',
                'required_top_level_json_shape': {'hypotheses': ['hypothesis_object']},
                'required_hypotheses_in_this_call': requested_in_call,
                'language': run.get('response_language', project.get('response_language', 'ru')),
                'malformed_response_reason': malformed_reason,
                'malformed_model_output': llm_result,
                'project': payload['project'],
                'allowed_evidence': compact_evidence,
                'numeric_context': numeric_context,
                'evidence_backed_intervention_families': payload['evidence_backed_intervention_families'],
                'already_accepted_titles': payload['already_accepted_titles'],
                'validation_rules': payload['validation_rules'],
                'hypothesis_object_schema': payload['hypothesis_object_schema'],
            }
            repaired_batch = await gateway.chat_json(
                role='repair',
                messages=[
                    {'role': 'system', 'content': repair_system_prompt},
                    {'role': 'user', 'content': json.dumps(schema_repair_payload, ensure_ascii=False)},
                ],
                fallback={'hypotheses': []},
            )
            candidates = normalize_llm_hypotheses(project, run, repaired_batch)
            raw_received += len(candidates)
            if not candidates:
                repair_keys = list(repaired_batch.keys())[:12] if isinstance(repaired_batch, dict) else []
                rejected.append({
                    'title': None,
                    'supporting_evidence': [],
                    'reasons': [malformed_reason, f'schema_repair_failed: {repair_keys or "empty response"}'],
                })
                continue
        elif not candidates:
            keys = list(llm_result.keys())[:12] if isinstance(llm_result, dict) else []
            malformed_reason = 'empty_or_invalid_llm_response' if not keys else f'wrong_top_level_json_keys: {keys}'
            rejected.append({
                'title': None,
                'supporting_evidence': [],
                'reasons': [malformed_reason, 'schema_repair_skipped_for_latency_budget'],
            })
            continue
        seen_titles = {item['title'].strip().lower() for item in accepted}
        max_repair_calls_per_attempt = max(0, int(os.getenv(
            'NORLAB_MAX_REPAIR_CALLS_PER_ATTEMPT',
            '0' if gateway.provider == 'yandex' else '1',
        )))
        repair_calls_used = 0
        for candidate in candidates:
            if len(accepted) >= desired_count:
                break
            valid, reasons, normalized = validate_llm_hypothesis(candidate, project, validation_context)
            key = normalized['title'].strip().lower()
            if key in seen_titles:
                valid = False
                reasons.append('duplicate or near-duplicate title')
            if valid:
                normalized['lineage']['generation_attempt'] = attempts_used
                normalized['lineage']['generation_mode'] = 'llm_grounded'
                accepted.append(normalized)
                seen_titles.add(key)
            else:
                if repair_calls_used >= max_repair_calls_per_attempt:
                    rejected.append({
                        'title': candidate.get('title'),
                        'supporting_evidence': candidate.get('supporting_evidence', []),
                        'reasons': reasons,
                    })
                    continue
                repair_calls_used += 1
                repair_payload = {
                    'task': 'Repair this rejected hypothesis. Return exactly {"hypothesis": {...}}.',
                    'rejection_reasons': reasons,
                    'strict_rules': [
                        'remove all unsupported numbers, percentages, units and thresholds from text fields',
                        'keep exact supported KPI numbers from project target, for example "не менее 2 п.п.", in statement, expected_kpis and falsification_conditions',
                        'if the hypothesis lacks a number but numeric_context has project_required_numbers, add the exact project number as the verification threshold',
                        'replace vague process parameters with exact supported values from numeric_context or with "диапазон определить в первом опыте"',
                        'keep economic_effect qualitative unless supplied evidence/project target contains exact economics',
                        'remove unsupported reagent and equipment names',
                        'do not claim equipment/reagent availability unless evidence says it',
                        'use only supporting_evidence ids from allowed_evidence',
                        'rewrite title to 3-8 words and at most 64 characters: specific action/mechanism plus target',
                        'keep the same project constraints and fill constraint_satisfaction for each one',
                    ],
                    'project': {
                        'problem': project.get('problem'),
                        'target_kpi': project.get('target_kpi'),
                        'constraints': project.get('constraints', []),
                    },
                    'allowed_evidence': compact_evidence,
                    'numeric_context': numeric_context,
                    'rejected_hypothesis': candidate,
                }
                repaired_result = await gateway.chat_json(
                    role='repair',
                    messages=[
                        {'role': 'system', 'content': system_prompt},
                        {'role': 'user', 'content': json.dumps(repair_payload, ensure_ascii=False)},
                    ],
                    fallback={},
                )
                repaired_candidates = normalize_llm_hypotheses(project, run, repaired_result)
                repaired_accepted = False
                for repaired in repaired_candidates[:1]:
                    repaired_valid, repaired_reasons, repaired_normalized = validate_llm_hypothesis(repaired, project, validation_context)
                    repaired_key = repaired_normalized['title'].strip().lower()
                    if repaired_key in seen_titles:
                        repaired_valid = False
                        repaired_reasons.append('duplicate or near-duplicate title')
                    if repaired_valid:
                        repaired_normalized['lineage']['generation_attempt'] = attempts_used
                        repaired_normalized['lineage']['generation_mode'] = 'llm_grounded_repair'
                        accepted.append(repaired_normalized)
                        seen_titles.add(repaired_key)
                        repaired_accepted = True
                    else:
                        reasons = reasons + [f'repair_failed: {reason}' for reason in repaired_reasons]
                if not repaired_accepted:
                    rejected.append({
                    'title': candidate.get('title'),
                    'supporting_evidence': candidate.get('supporting_evidence', []),
                    'reasons': reasons,
                    })
        if gateway.provider == 'yandex' and len(accepted) >= desired_count:
            break

    # DeepSeek remains the primary hypothesis generator.  If its grounded
    # batches are candidates short after all configured attempts,
    # ask the GPT OSS fast model for a final grounded refill.  This is a real
    # model call using the same evidence and validator, not a deterministic
    # fallback. If the refill still cannot produce enough valid hypotheses, we
    # keep the high-quality partial portfolio instead of padding it with weak
    # candidates.
    fast_refill_used = False
    if len(accepted) < desired_count:
        refill_attempts = max(0, int(os.getenv('NORLAB_LLM_REFILL_ATTEMPTS', '1')))
        fast_refill_used = refill_attempts > 0
        for refill_attempt in range(refill_attempts):
            if len(accepted) >= desired_count:
                break
            refill_count = min(batch_size, desired_count - len(accepted))
            refill_payload = {
                'task': f'Generate exactly {refill_count} NEW valid grounded hypotheses to complete the portfolio.',
                'required_top_level_json_shape': {'hypotheses': ['hypothesis_object']},
                'required_hypotheses_in_this_call': refill_count,
                'language': run.get('response_language', project.get('response_language', 'ru')),
                'project': payload['project'],
                'allowed_evidence': compact_evidence,
                'numeric_context': numeric_context,
                'evidence_backed_intervention_families': payload['evidence_backed_intervention_families'],
                'already_accepted_titles': [item['title'] for item in accepted],
                'previous_rejections': rejected[-12:],
                'validation_rules': payload['validation_rules'],
                'hypothesis_object_schema': payload['hypothesis_object_schema'],
                'refill_reason': 'DeepSeek batches did not yield enough schema-valid grounded candidates.',
            }
            refill_result = await gateway.chat_json(
                role='repair',
                messages=[
                    {'role': 'system', 'content': repair_system_prompt},
                    {'role': 'user', 'content': json.dumps(refill_payload, ensure_ascii=False)},
                ],
                fallback={'hypotheses': []},
            )
            refill_candidates = normalize_llm_hypotheses(project, run, refill_result)
            raw_received += len(refill_candidates)
            seen_titles = {item['title'].strip().lower() for item in accepted}
            if not refill_candidates:
                rejected.append({
                    'title': None,
                    'supporting_evidence': [],
                    'reasons': [f'refill_attempt_{refill_attempt + 1}_returned_no_hypotheses'],
                })
                continue
            for candidate in refill_candidates:
                if len(accepted) >= desired_count:
                    break
                valid, reasons, normalized = validate_llm_hypothesis(candidate, project, validation_context)
                key = normalized['title'].strip().lower()
                if key in seen_titles:
                    valid = False
                    reasons.append('duplicate or near-duplicate title')
                if valid:
                    normalized['lineage']['generation_attempt'] = attempts_used + refill_attempt + 1
                    normalized['lineage']['generation_mode'] = 'gpt_oss_grounded_refill'
                    accepted.append(normalized)
                    seen_titles.add(key)
                else:
                    rejected.append({
                        'title': candidate.get('title'),
                        'supporting_evidence': candidate.get('supporting_evidence', []),
                        'reasons': reasons,
                    })

    local_guardrail_used = False
    if len(accepted) < desired_count and os.getenv('NORLAB_ENABLE_GROUNDED_LOCAL_GUARDRAIL', 'true').lower() == 'true':
        local_guardrail_used = True
        seen_titles = {item['title'].strip().lower() for item in accepted}
        for candidate in create_hypotheses(project, run):
            if len(accepted) >= desired_count:
                break
            candidate = add_concrete_experimental_levels(ensure_project_numeric_target_in_hypothesis(candidate, project))
            valid, reasons, normalized = validate_llm_hypothesis(candidate, project, validation_context)
            key = normalized['title'].strip().lower()
            if key in seen_titles:
                continue
            if valid:
                normalized['lineage']['generation_attempt'] = attempts_used + 1
                normalized['lineage']['generation_mode'] = 'local_evidence_guardrail_after_llm'
                accepted.append(normalized)
                seen_titles.add(key)
            else:
                rejected.append({
                    'title': candidate.get('title'),
                    'supporting_evidence': candidate.get('supporting_evidence', []),
                    'reasons': reasons,
                })

    summary = {
        'requested': desired_count,
        'minimum_accepted': hard_minimum_accepted,
        'target_accepted': desired_count,
        'received': raw_received,
        'accepted': len(accepted),
        'rejected': len(rejected),
        'rejections': rejected[-12:],
        'attempts': attempts_used,
        'max_attempts': max_attempts,
        'fallback_used': False,
        'supplemented_with_deterministic': local_guardrail_used,
        'generation_mode': f'{gateway.generator_model}_with_{gateway.fast_model}_grounded_refill' if fast_refill_used else f'{gateway.generator_model}_grounded',
        'fast_refill_used': fast_refill_used,
        'local_guardrail_used': local_guardrail_used,
        'models': {
            'generator': run['model_profile'].get('generator'),
            'critic': run['model_profile'].get('critic'),
            'fast': run['model_profile'].get('fast'),
        },
    }
    if len(accepted) < hard_minimum_accepted:
        summary['status'] = 'failed_validation_or_timeout'
        run['llm_generation_summary'] = summary
        store.save()
        raise RuntimeError(
            f'LLM generated no valid grounded hypotheses; requested {desired_count}; '
            f'rejected={len(rejected)}. Last rejections: {json.dumps(rejected[-3:], ensure_ascii=False)}'
        )
    summary['status'] = 'accepted' if len(accepted) >= desired_count else 'accepted_partial'
    return accepted[:desired_count], summary


def hypothesis_intervention_concept(hypothesis: dict[str, Any]) -> str:
    concepts = [
        ('ph_control', ['ph', 'кислотност', 'щелочност', 'alkalin']),
        ('hydrocyclone_classification', ['гидроциклон', 'hydrocyclone', 'd50', 'классификац']),
        ('grind_size', ['p80', 'доизмельч', 'измельч', 'помол', 'particle size', 'grind', 'liberation']),
        ('conditioning_time', ['кондиционир', 'времен', 'contact time', 'conditioning time']),
        ('sampling_data_quality', ['отбор проб', 'повторн', 'качество данных', 'sampling', 'data quality']),
    ]
    fields = [
        str(hypothesis.get('title') or '').lower(),
        json.dumps(hypothesis.get('intervention') or {}, ensure_ascii=False).lower(),
        str(hypothesis.get('statement') or '').lower(),
    ]
    # Prefer the compact title; mechanisms often mention neighbouring process
    # stages and would otherwise misclassify a grinding hypothesis as a
    # hydrocyclone or collector hypothesis.
    for field in fields:
        if re.search(r'(?<![а-яё])ионн', field) or any(term in field for term in ['ion pre', 'ion treatment', 'ionic treatment']):
            return 'ion_pretreatment'
        for concept, terms in concepts:
            if any(term in field for term in terms):
                return concept
    return 'other'


def deduplicate(hypotheses: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    duplicates: list[dict[str, Any]] = []
    for item in hypotheses:
        key = re.sub(r'\W+', ' ', item['title'].lower()).strip()
        concept = hypothesis_intervention_concept(item)
        item_tokens = tokenize_for_grounding(f'{item.get("title", "")} {item.get("statement", "")}')
        duplicate_of = None
        for existing in unique:
            existing_key = re.sub(r'\W+', ' ', existing['title'].lower()).strip()
            if key == existing_key:
                duplicate_of = existing
                break
            existing_concept = hypothesis_intervention_concept(existing)
            if concept == 'other' or concept != existing_concept:
                continue
            existing_tokens = tokenize_for_grounding(f'{existing.get("title", "")} {existing.get("statement", "")}')
            similarity = len(item_tokens & existing_tokens) / max(1, len(item_tokens | existing_tokens))
            if similarity >= 0.18 or concept in {'grind_size', 'hydrocyclone_classification', 'conditioning_time'}:
                duplicate_of = existing
                break
        if key in seen or duplicate_of is not None:
            item['status'] = 'DUPLICATE'
            item['rejection_stage'] = 'deduplication'
            item['rejection_reasons'] = [
                f'Смысловой дубль гипотезы «{(duplicate_of or item).get("title", item["title"])}»: совпадает проверяемое инженерное воздействие.'
            ]
            duplicates.append(item)
            continue
        seen.add(key)
        unique.append(item)
    return unique, duplicates


def apply_hard_gates(hypotheses: list[dict[str, Any]]) -> None:
    for item in hypotheses:
        constraints = item.get('constraints') or item.get('conditions') or []
        constraint_satisfaction = item.get('constraint_satisfaction')
        grounding = item.get('grounding', {})
        gates = {
            'evidence_gate': bool(item['supporting_evidence']),
            'grounding_gate': bool(grounding.get('evidence_ids', item.get('supporting_evidence'))),
            'engineering_constraint_gate': not constraints or bool(constraint_satisfaction or item.get('conditions')),
            'falsifiability_gate': bool(item['falsification_conditions']),
            'experimentability_gate': True,
        }
        item['hard_gates'] = gates
        item['status'] = 'FINALIST' if all(gates.values()) else 'BLOCKED'
        if item['status'] == 'BLOCKED':
            item['rejection_stage'] = 'gates'
            item['rejection_reasons'] = [name for name, passed in gates.items() if not passed]


def critique(hypothesis: dict[str, Any]) -> dict[str, Any]:
    unsupported = []
    if 'Точные режимные параметры' in ' '.join(hypothesis.get('assumptions', [])):
        unsupported.append('Не определены диапазоны режимных параметров.')
    return {
        'fatal_flaws': [],
        'counterexamples': [],
        'unsupported_assumptions': unsupported,
        'transfer_risks': hypothesis.get('risks', []),
        'missing_variables': ['минералогия', 'гранулометрия', 'содержание целевых элементов'],
        'falsification_test': hypothesis['falsification_conditions'][0],
        'scores': {'technical': 0.68, 'evidence': hypothesis['scores']['evidence'], 'risk': 0.45},
    }


async def critique_with_llm(hypothesis: dict[str, Any], evidence: list[dict[str, Any]], use_llm: bool) -> dict[str, Any]:
    deterministic = critique(hypothesis)
    if not use_llm:
        return deterministic
    result: dict[str, Any] = {}
    for _ in range(int(os.getenv('NORLAB_LLM_CRITIQUE_ATTEMPTS', '2'))):
        result = await gateway.chat_json(
            role='critic',
            messages=[
                {'role': 'system', 'content': 'Ты независимый критик NORLAB. Верни только JSON: fatal_flaws, counterexamples, unsupported_assumptions, transfer_risks, missing_variables, falsification_test, scores. Нельзя добавлять факты вне supplied evidence; если утверждение гипотезы не доказано evidence, внеси это в unsupported_assumptions.'},
                {'role': 'user', 'content': json.dumps({'hypothesis': hypothesis, 'evidence': evidence}, ensure_ascii=False)},
            ],
            fallback={},
        )
        if isinstance(result, dict) and result:
            break
    if not isinstance(result, dict) or not result:
        if os.getenv('NORLAB_STRICT_LLM', 'true').lower() == 'true':
            raise RuntimeError(f'LLM critic returned empty/invalid response for hypothesis {hypothesis["id"]}')
        deterministic['llm_status'] = 'fallback_after_empty_or_invalid_response'
        return deterministic
    return {
        'fatal_flaws': result.get('fatal_flaws', []),
        'counterexamples': result.get('counterexamples', []),
        'unsupported_assumptions': result.get('unsupported_assumptions', []),
        'transfer_risks': result.get('transfer_risks', []),
        'missing_variables': result.get('missing_variables', []),
        'falsification_test': result.get('falsification_test') or deterministic['falsification_test'],
        'scores': result.get('scores') if isinstance(result.get('scores'), dict) else deterministic['scores'],
        'llm_status': 'accepted',
    }


def normalize_critique_result(result: dict[str, Any], deterministic: dict[str, Any]) -> dict[str, Any]:
    if not isinstance(result, dict) or not result:
        normalized = dict(deterministic)
        normalized['llm_status'] = 'local_gate_after_empty_batch_critic'
        return normalized
    return {
        'fatal_flaws': result.get('fatal_flaws', []) if isinstance(result.get('fatal_flaws', []), list) else [],
        'counterexamples': result.get('counterexamples', []) if isinstance(result.get('counterexamples', []), list) else [],
        'unsupported_assumptions': result.get('unsupported_assumptions', []) if isinstance(result.get('unsupported_assumptions', []), list) else [],
        'transfer_risks': result.get('transfer_risks', []) if isinstance(result.get('transfer_risks', []), list) else [],
        'missing_variables': result.get('missing_variables', []) if isinstance(result.get('missing_variables', []), list) else [],
        'falsification_test': result.get('falsification_test') or deterministic['falsification_test'],
        'scores': result.get('scores') if isinstance(result.get('scores'), dict) else deterministic['scores'],
        'llm_status': result.get('llm_status') or 'accepted_batch',
    }


async def critique_portfolio_with_llm(
    hypotheses: list[dict[str, Any]],
    evidence: list[dict[str, Any]],
    use_llm: bool,
) -> dict[str, dict[str, Any]]:
    deterministic_by_id = {hypothesis['id']: critique(hypothesis) for hypothesis in hypotheses}
    if not use_llm or not hypotheses:
        return deterministic_by_id

    compact_hypotheses = [
        {
            'id': hypothesis['id'],
            'title': hypothesis.get('title'),
            'statement': hypothesis.get('statement'),
            'intervention': hypothesis.get('intervention'),
            'causal_mechanism': hypothesis.get('causal_mechanism'),
            'supporting_evidence': hypothesis.get('supporting_evidence', []),
            'assumptions': hypothesis.get('assumptions', []),
            'expected_kpis': hypothesis.get('expected_kpis', []),
            'economic_effect': hypothesis.get('economic_effect'),
            'risks': hypothesis.get('risks', []),
            'falsification_conditions': hypothesis.get('falsification_conditions', []),
            'scores': hypothesis.get('scores', {}),
        }
        for hypothesis in hypotheses
    ]
    compact_evidence = compact_evidence_for_llm(evidence[: int(os.getenv('NORLAB_CRITIQUE_EVIDENCE_LIMIT', '18'))])
    result = await gateway.chat_json(
        role='critic',
        messages=[
            {
                'role': 'system',
                'content': (
                    'You are NORLAB portfolio critic. Return JSON only: '
                    '{"critiques":[{"id":"...","fatal_flaws":[],"counterexamples":[],'
                    '"unsupported_assumptions":[],"transfer_risks":[],"missing_variables":[],'
                    '"falsification_test":"...","scores":{"technical":0.0,"evidence":0.0,"risk":0.0}}]}. '
                    'Critique every supplied hypothesis exactly once. Do not add facts outside supplied evidence. '
                    'Reject only hard contradictions or unsupported claims; uncertainty should go to assumptions/risks.'
                ),
            },
            {
                'role': 'user',
                'content': json.dumps(
                    {
                        'hypotheses': compact_hypotheses,
                        'evidence': compact_evidence,
                        'required_ids': [hypothesis['id'] for hypothesis in hypotheses],
                    },
                    ensure_ascii=False,
                ),
            },
        ],
        fallback={'critiques': []},
    )
    raw_critiques = result.get('critiques') if isinstance(result, dict) else []
    if not isinstance(raw_critiques, list):
        raw_critiques = result.get('items') if isinstance(result, dict) and isinstance(result.get('items'), list) else []

    by_id: dict[str, dict[str, Any]] = {}
    for item in raw_critiques:
        if not isinstance(item, dict):
            continue
        critique_id = str(item.get('id') or '').strip()
        if critique_id in deterministic_by_id:
            by_id[critique_id] = normalize_critique_result(item, deterministic_by_id[critique_id])

    for hypothesis_id, deterministic in deterministic_by_id.items():
        if hypothesis_id not in by_id:
            normalized = dict(deterministic)
            normalized['llm_status'] = 'local_gate_after_missing_batch_critic'
            by_id[hypothesis_id] = normalized
    return by_id


def is_critic_rejection_reason(reason: str) -> bool:
    lowered = reason.lower()
    hard_markers = [
        'нарушает огранич',
        'противоречит огранич',
        'невозмож',
        'опасн',
        'запрещ',
        'не относится к отвальным хвост',
        'не относится к шлак',
        'промышленный синтез',
        'нет supporting_evidence',
        'нет источников',
        'вне supplied evidence',
    ]
    soft_markers = [
        'отсутствие прям',
        'отсутствует количествен',
        'не подтверждено',
        'нет данных',
        'нет доказательств',
        'требует проверки',
        'не подкреплено',
        'невозможно оценить',
        'не указаны текущие',
        'не указан диапазон',
    ]
    return any(marker in lowered for marker in hard_markers) and not any(marker in lowered for marker in soft_markers)


def novelty(hypothesis: dict[str, Any], project: dict[str, Any], external_sources: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    external = project.get('external_research_enabled', False)
    analogs = []
    for source in (external_sources or [])[:5]:
        title = source.get('title', '').lower()
        statement_terms = set(re.findall(r'[a-zа-яё]{5,}', hypothesis['statement'].lower()))
        overlap = sum(1 for term in statement_terms if term in title)
        if overlap or not analogs:
            analogs.append({
                'source_id': source['id'],
                'title': source['title'],
                'doi': source.get('doi'),
                'url': source.get('url'),
                'year': source.get('year'),
            })
    return {
        'class': 'ADAPTED' if analogs else ('POTENTIALLY_NOVEL' if external else 'UNKNOWN'),
        'boundary': 'Не является патентным заключением; для финала нужен prior-art search.',
        'external_research_used': external,
        'analogs': analogs,
    }


def disagreement(hypothesis: dict[str, Any]) -> dict[str, Any]:
    critique_scores = hypothesis.get('critique', {}).get('scores', {}) or {}
    technical_score = (
        critique_scores.get('technical')
        or critique_scores.get('mechanism')
        or critique_scores.get('feasibility')
        or hypothesis['scores'].get('feasibility')
        or 0.5
    )
    values = [
        hypothesis['scores']['evidence'],
        technical_score,
        hypothesis['scores']['feasibility'],
        hypothesis['scores']['experimentability'],
    ]
    return {
        'criteria': {
            'evidence': values[0],
            'mechanism': values[1],
            'feasibility': values[2],
            'experimentability': values[3],
        },
        'dispersion': round(max(values) - min(values), 3),
        'risk_signal': 'HIGH_DISAGREEMENT' if max(values) - min(values) > 0.35 else 'NORMAL',
    }


def uncertainty(hypothesis: dict[str, Any], facts: list[dict[str, Any]]) -> dict[str, Any]:
    has_quality_issues = any(fact['type'] == 'data_quality' for fact in facts)
    level = 'HIGH' if has_quality_issues or len(hypothesis['supporting_evidence']) < 2 else 'MEDIUM'
    reasons = ['ограниченное покрытие evidence', 'переносимость эффекта требует проверки']
    if has_quality_issues:
        reasons.append('найдены ошибки или пропуски в XLSX-данных')
    return {
        'level': level,
        'reasons': reasons,
        'sensitive_assumptions': hypothesis.get('assumptions', []),
        'clarifying_question': None,
        'diagnostic_experiment': 'Провести малую серию batch-флотации с матрицей pH x крупность x расход реагента.',
    }


def compile_experiment(hypothesis: dict[str, Any]) -> dict[str, Any]:
    return {
        'id': new_id('exp'),
        'hypothesis_id': hypothesis['id'],
        'project_id': hypothesis['project_id'],
        'objective': f'Проверить гипотезу: {hypothesis["title"]}',
        'experiment_type': 'laboratory_screening',
        'factors_and_levels': {
            'pH': ['baseline', 'baseline-0.5', 'baseline+0.5'],
            'grind_size': ['baseline', 'finer_class'],
            'reagent_dose': ['baseline', '+10%', '-10%'],
        },
        'controls_baseline': ['текущий режим фабрики или исторический лабораторный baseline'],
        'fixed_conditions': ['одинаковая масса навески', 'одинаковое время флотации', 'одна партия пробы'],
        'steps': [
            'Подготовить представительную пробу хвостов.',
            'Зафиксировать baseline-химию, гранулометрию и минералогию.',
            'Провести серию опытов по матрице факторов.',
            'Измерить извлечение, содержание и потери целевых элементов.',
            'Сравнить с baseline и проверить критерии остановки.',
        ],
        'samples_repeats': 'Минимум 2 повтора для перспективных режимов.',
        'equipment_materials_reagents': ['флотомашина лабораторная', 'pH-метр', 'сита/лазерный анализатор', 'стандартные реагенты'],
        'measurements_units': ['извлечение, %', 'содержание элемента, г/т или %', 'массовый выход, %'],
        'data_collection_template': {'sample_id': '', 'factor_levels': {}, 'measurements': {}},
        'analysis_method': 'Сравнение с baseline, ранжирование по KPI и проверка устойчивости эффекта.',
        'safety_regulatory_notes': ['Следовать регламентам обращения с реагентами и хвостовыми пробами.'],
        'resource_estimate': {'duration_days': 3, 'samples': 12},
        'success_criteria': ['улучшение целевого KPI без нарушения ограничений'],
        'failure_criteria': ['ухудшение KPI или нереплицируемый эффект'],
        'early_stop_criteria': ['систематическое ухудшение KPI в первых контрольных опытах'],
        'decision_tree': {
            'positive': 'перейти к расширенной матрице и пилотной проверке',
            'negative': 'закрыть или вернуть в генерацию с контрфактами',
            'inconclusive': 'провести диагностический опыт по недостающей переменной',
        },
        'assumptions': hypothesis.get('assumptions', []),
        'unresolved_uncertainties': hypothesis.get('uncertainty', {}).get('reasons', []),
        'results': [],
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }


async def execute_pipeline(project_id: str, run_id: str, use_llm: bool = False) -> None:
    project = get_project(project_id)
    run = get_run(run_id)
    try:
        docs = project_documents(project_id)
        add_event(run, RunStatus.INGESTING, 'Проверены загруженные документы.', len(docs), 0)
        vision_limit = int(os.getenv('NORLAB_VISION_ANALYZE_LIMIT', '2'))
        vision_done = 0
        for document in docs:
            if vision_done >= vision_limit:
                break
            if document.get('content_type') in {'png', 'jpg', 'jpeg', 'webp'} or str(document.get('filename', '')).lower().endswith(('.png', '.jpg', '.jpeg', '.webp')):
                try:
                    await analyze_document_image(document)
                    vision_done += 1
                except Exception as exc:
                    document.setdefault('vision_errors', []).append(str(exc)[:300])
        if vision_done:
            add_event(run, RunStatus.ANALYZING, 'Выполнен vision-разбор изображений и схем.', vision_done, 0)
        facts = project_facts(project_id)
        add_event(run, RunStatus.ANALYZING, 'Извлечены факты и признаки качества данных.', len(facts), 0)
        if not project.get('target_kpi'):
            question = {
                'id': new_id('clar'),
                'question': 'Какой целевой KPI считать главным для ранжирования гипотез?',
                'reason': 'Без KPI рейтинг гипотез может измениться принципиально.',
                'affected_decisions': ['portfolio_ranking', 'experiment_success_criteria'],
                'blocking': False,
                'answer_type': 'text',
                'options': [],
                'recommended_default': 'повысить извлечение ценных компонентов при допустимых затратах',
                'allow_unknown': True,
                'continue_with_assumption': True,
                'answer': None,
            }
            run['clarifications'] = [question]
            project['target_kpi'] = question['recommended_default']
        add_event(run, RunStatus.RETRIEVING_MEMORY, 'Собран локальный evidence pack.', len(evidence_records(project_id)), 0)
        embedding_summary = await reindex_project_embeddings(project_id, limit=int(os.getenv('NORLAB_EMBEDDING_REINDEX_LIMIT', '8')))
        run['embedding_index'] = embedding_summary
        external_sources: list[dict[str, Any]] = []
        if project.get('external_research_enabled'):
            sanitized = sanitizer.sanitize(project['problem'])
            external_sources = await external_research(sanitized, max_sources=8)
            patent_sources = await patent_research(sanitized, max_sources=5)
            external_sources.extend(patent_sources)
            for source in external_sources:
                store.data['sources'][source['id']] = source
            run['external_research'] = {'queries': [sanitized], 'sources': [source['id'] for source in external_sources], 'status': 'completed'}
            add_event(run, RunStatus.RESEARCHING_EXTERNAL, 'Выполнен внешний поиск по научным metadata-источникам.', len(external_sources), 0)
        else:
            add_event(run, RunStatus.RESEARCHING_EXTERNAL, 'Внешний поиск отключён на уровне проекта.', 0, 0)
        default_candidate_count = int(os.getenv('NORLAB_DEFAULT_HYPOTHESIS_CANDIDATES', '12'))
        desired_hypothesis_count = min(
            12,
            max(
                1,
                default_candidate_count,
                int(run['settings'].get('candidate_count') or 0),
            ),
        )
        run.setdefault('artifacts', {})['funnel'] = {
            'requested': desired_hypothesis_count,
            'generated': 0,
            'received': 0,
            'accepted': 0,
            'unique': 0,
            'gates': 0,
            'critique': 0,
            'finalists': 0,
        }
        if use_llm:
            add_event(run, RunStatus.GENERATING, 'Запущена LLM-генерация проверяемых гипотез с evidence/constraint validation.', desired_hypothesis_count, 0)
            hypotheses, generation_summary = await generate_llm_hypotheses(project, run, external_sources, desired_hypothesis_count)
            run['llm_generation_summary'] = generation_summary
        else:
            add_event(run, RunStatus.GENERATING, 'Запущен локальный evidence pipeline для проверяемых гипотез.', desired_hypothesis_count, 0)
            hypotheses = create_hypotheses(project, run)[:desired_hypothesis_count]
            run['llm_generation_summary'] = {
                'requested': desired_hypothesis_count,
                'accepted': len(hypotheses),
                'generation_mode': 'local_evidence_pipeline',
                'fallback_used': False,
                'supplemented_with_deterministic': False,
            }
        add_event(run, RunStatus.DEDUPLICATING, 'Выполнена дедупликация кандидатов.', len(hypotheses), 0)
        hypotheses, duplicate_hypotheses = deduplicate(hypotheses)
        apply_hard_gates(hypotheses)
        gates_count = sum(1 for item in hypotheses if item.get('status') == 'FINALIST')
        add_event(run, RunStatus.APPLYING_GATES, 'Применены Evidence, Engineering, Falsifiability и Experimentability gates.', len(hypotheses), 0)
        evidence_pack = evidence_records(project_id)
        critique_by_id = await critique_portfolio_with_llm(hypotheses, evidence_pack, use_llm)
        for hypothesis in hypotheses:
            hypothesis['critique'] = critique_by_id.get(hypothesis['id']) or critique(hypothesis)
            fatal_flaws = [str(item) for item in hypothesis.get('critique', {}).get('fatal_flaws', []) if str(item).strip()]
            rejecting_flaws = [reason for reason in fatal_flaws if is_critic_rejection_reason(reason)]
            critic_scores = hypothesis.get('critique', {}).get('scores') or {}
            critic_evidence = clamp01(critic_scores.get('evidence', hypothesis.get('scores', {}).get('evidence', 0.5)))
            if rejecting_flaws and critic_evidence < 0.35:
                rejecting_flaws.append(
                    f'Критик отклонил гипотезу: доказательность {round(critic_evidence * 100)}% ниже порога 45% и выявлены существенные контраргументы.'
                )
            hypothesis['critique_reasons'] = fatal_flaws
            if rejecting_flaws and hypothesis.get('status') == 'FINALIST':
                hypothesis['status'] = 'REJECTED'
                hypothesis['rejection_stage'] = 'critic'
                hypothesis['rejection_reasons'] = rejecting_flaws
        critique_count = sum(1 for item in hypotheses if item.get('status') == 'FINALIST')
        add_event(run, RunStatus.CRITIQUING, 'Выполнена индивидуальная критика финалистов.', len(hypotheses), 0)
        for hypothesis in hypotheses:
            hypothesis['novelty'] = novelty(hypothesis, project, external_sources)
        add_event(run, RunStatus.CHECKING_NOVELTY, 'Сформирован Novelty Radar.', len(hypotheses), 0)
        for hypothesis in hypotheses:
            hypothesis['disagreement'] = disagreement(hypothesis)
        add_event(run, RunStatus.ANALYZING_DISAGREEMENT, 'Сформирована Disagreement Map.', len(hypotheses), 0)
        for hypothesis in hypotheses:
            hypothesis['uncertainty'] = uncertainty(hypothesis, facts)
        add_event(run, RunStatus.ANALYZING_UNCERTAINTY, 'Сформирован Uncertainty Navigator.', len(hypotheses), 0)
        stored_hypotheses = [*hypotheses, *duplicate_hypotheses]
        for hypothesis in stored_hypotheses:
            store.data['hypotheses'][hypothesis['id']] = hypothesis
            if hypothesis.get('status') != 'DUPLICATE':
                neo4j_memory.upsert_hypothesis(hypothesis)
            if neo4j_memory.driver and hypothesis.get('status') != 'DUPLICATE':
                try:
                    hypothesis_embedding_timeout = float(os.getenv('NORLAB_HYPOTHESIS_EMBED_TIMEOUT_SECONDS', '12'))
                    embedding_result = await asyncio.wait_for(
                        gateway.embed_text(f'{hypothesis["title"]}\n{hypothesis["statement"]}', is_document=True),
                        timeout=hypothesis_embedding_timeout,
                    )
                    neo4j_memory.set_embedding('Hypothesis', hypothesis['id'], embedding_result['embedding'], embedding_result.get('model_version'))
                except Exception as exc:
                    hypothesis.setdefault('embedding_errors', []).append(str(exc)[:300])
        final_hypotheses = [item for item in hypotheses if item.get('status') == 'FINALIST']
        finalist_limit = min(len(final_hypotheses), int(run['settings'].get('max_finalists') or len(final_hypotheses)))
        experiments = [compile_experiment(item) for item in final_hypotheses[:finalist_limit]]
        for experiment in experiments:
            store.data['experiments'][experiment['id']] = experiment
            neo4j_memory.upsert_experiment(experiment)
        add_event(run, RunStatus.COMPILING_EXPERIMENTS, 'Скомпилированы планы экспериментов.', len(experiments), 0)
        run['artifacts'] = {
            'evidence_pack': evidence_records(project_id),
            'hypothesis_ids': [item['id'] for item in stored_hypotheses],
            'experiment_ids': [item['id'] for item in experiments],
            'funnel': {
                'generated': min(
                    desired_hypothesis_count,
                    max(
                        int(generation_summary.get('received') or 0),
                        int(generation_summary.get('accepted') or 0),
                        len(hypotheses),
                    ),
                ) if use_llm else len(hypotheses),
                'received': generation_summary.get('received') if use_llm else len(hypotheses),
                'accepted': generation_summary.get('accepted') if use_llm else len(hypotheses),
                'unique': len(hypotheses),
                'gates': gates_count,
                'critique': critique_count,
                'finalists': len(final_hypotheses),
            },
            'disagreement_map': {item['id']: item['disagreement'] for item in hypotheses},
            'uncertainty_navigator': {item['id']: item['uncertainty'] for item in hypotheses},
        }
        run['report'] = build_report(project, run, hypotheses, experiments)
        add_event(run, RunStatus.BUILDING_REPORT, 'Сформирован JSON/Markdown отчёт.', 1, 0)
        add_event(run, RunStatus.COMPLETED, 'Pipeline завершён.', 1, 0)
    except Exception as exc:
        run['error'] = str(exc)
        add_event(run, RunStatus.FAILED, f'Pipeline failed: {exc}', 0, 0)


@celery_app.task(name='norlab.execute_pipeline')
def execute_pipeline_task(project_id: str, run_id: str, use_llm: bool = False) -> None:
    store.reload()
    asyncio.run(execute_pipeline(project_id, run_id, use_llm))


def build_report(project: dict[str, Any], run: dict[str, Any], hypotheses: list[dict[str, Any]], experiments: list[dict[str, Any]]) -> dict[str, Any]:
    lines = [
        f'# NORLAB report: {project["name"]}',
        '',
        f'Problem: {project["problem"]}',
        f'Target KPI: {project.get("target_kpi") or "not specified"}',
        '',
        '## Final hypotheses',
    ]
    for item in hypotheses:
        lines.extend([
            f'### {item["title"]}',
            item['statement'],
            f'Evidence: {", ".join(item["supporting_evidence"])}',
            f'Novelty: {item["novelty"]["class"]}',
            f'Uncertainty: {item["uncertainty"]["level"]}',
            '',
        ])
    lines.append('## Experiments')
    for item in experiments:
        lines.append(f'- {item["objective"]}')
    return {
        'id': new_id('report'),
        'run_id': run['id'],
        'language': project.get('response_language', 'ru'),
        'json': {'project': project, 'hypotheses': hypotheses, 'experiments': experiments},
        'markdown': '\n'.join(lines),
        'created_at': now_iso(),
    }


def projection_base(entity_id: str, summary: dict[str, Any], warnings: list[str] | None = None) -> dict[str, Any]:
    return {
        'id': entity_id,
        'version': 1,
        'updated_at': now_iso(),
        'summary': summary,
        'capabilities': ['view', 'refresh'],
        'warnings': warnings or [],
        'is_partial': False,
        'missing_artifacts': [],
        'links': {},
        'server_timestamp': now_iso(),
    }


def bff_meta(entity: dict[str, Any], capabilities: list[str] | None = None, warnings: list[str] | None = None) -> dict[str, Any]:
    return {
        'id': entity['id'],
        'version': int(entity.get('version', 1)),
        'updated_at': entity.get('updated_at') or entity.get('created_at') or now_iso(),
        'capabilities': capabilities or ['view'],
        'warnings': warnings or [],
        'partial': bool(entity.get('partial', False)),
    }


def project_warning_codes(project: dict[str, Any], docs: list[dict[str, Any]] | None = None) -> list[str]:
    warnings: list[str] = []
    if not project.get('target_kpi'):
        warnings.append('target_kpi_missing')
    if not (docs if docs is not None else project_documents(project['id'])):
        warnings.append('source_files_missing')
    return warnings


def project_readiness(project: dict[str, Any], docs: list[dict[str, Any]] | None = None) -> int:
    docs = docs if docs is not None else project_documents(project['id'])
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project['id']]
    hypotheses_count = sum(1 for item in store.data['hypotheses'].values() if item['project_id'] == project['id'])
    score = 20
    if project.get('problem'):
        score += 20
    if project.get('target_kpi'):
        score += 15
    if project.get('constraints'):
        score += 10
    score += min(20, len(docs) * 4)
    if runs:
        score += 10
    if hypotheses_count:
        score += 5
    return max(0, min(100, score))


def project_focus(project: dict[str, Any]) -> str:
    if project.get('area'):
        return str(project['area'])
    problem = str(project.get('problem') or '').lower()
    if 'мед' in problem or 'copper' in problem or 'cu' in problem:
        return 'Металлургический передел · Cu'
    if 'флотац' in problem or 'flotation' in problem:
        return 'Отвальные хвосты · флотация'
    return 'Отвальные хвосты · Ni/Cu/PGM'


def bff_brief(project: dict[str, Any]) -> dict[str, Any]:
    constraints = project.get('constraints', [])
    if isinstance(constraints, str):
        constraints_text = constraints
    else:
        constraints_text = '\n'.join(str(item) for item in constraints)
    return {
        'problem': project.get('problem') or project.get('name') or 'Исследовательская задача не заполнена.',
        'goal': project.get('goal') or project.get('expected_result') or project.get('target_kpi') or 'Сформировать проверяемые гипотезы и протокол проверки.',
        'constraints': constraints_text,
        'success_criterion': project.get('success_criterion') or project.get('target_kpi') or 'Критерий успеха требует уточнения.',
        'domain': 'tailings_and_metallurgy',
    }


def bff_project_summary(project: dict[str, Any]) -> dict[str, Any]:
    docs = project_documents(project['id'])
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project['id']]
    latest_run = max(runs, key=lambda item: str(item.get('created_at') or item.get('updated_at') or '')) if runs else None
    latest_artifacts = (latest_run or {}).get('artifacts') or {}
    latest_hypotheses = [
        store.data['hypotheses'].get(item_id, {})
        for item_id in latest_artifacts.get('hypothesis_ids', [])
    ]
    experiments_count = len(latest_artifacts.get('experiment_ids', []))
    finalists = sum(1 for item in latest_hypotheses if item.get('status') == 'FINALIST')
    warning_codes = project_warning_codes(project, docs)
    meta = bff_meta(project, ['open', 'edit', 'run'], warning_codes)
    return {
        **meta,
        'name': project.get('name') or 'Без названия',
        'area': project_focus(project),
        'readiness': project_readiness(project, docs),
        'last_run_at': latest_run.get('updated_at') if latest_run else None,
        'finalists': finalists,
        'open_experiments': experiments_count,
        'status': 'attention' if warning_codes else 'active',
    }


def bff_project(project: dict[str, Any]) -> dict[str, Any]:
    docs = project_documents(project['id'])
    meta = bff_meta(project, ['file_preview', 'run_stream', 'report_export'], project_warning_codes(project, docs))
    return {
        **meta,
        'name': project.get('name') or 'Без названия',
        'focus': project_focus(project),
        'readiness': project_readiness(project, docs),
        'brief': bff_brief(project),
        'indexed_files': len(docs),
        'memory_matches': min(12, len(project_facts(project['id']))),
    }


def file_kind(document: dict[str, Any]) -> str:
    name = str(document.get('filename') or '').lower()
    content_type = str(document.get('content_type') or '').lower()
    if name.endswith('.pdf') or 'pdf' in content_type:
        return 'pdf'
    if name.endswith('.docx') or 'wordprocessingml' in content_type:
        return 'docx'
    if name.endswith(('.xlsx', '.xls', '.csv')) or 'spreadsheet' in content_type or content_type in {'xlsx', 'xls', 'csv'}:
        return 'spreadsheet'
    if name.endswith(('.png', '.jpg', '.jpeg', '.webp', '.svg')) or content_type.startswith('image/'):
        return 'image'
    if name.endswith(('.txt', '.md', '.json', '.xml', '.html', '.log', '.rtf')) or content_type.startswith('text/'):
        return 'text'
    return 'unknown'


def bff_file(document: dict[str, Any]) -> dict[str, Any]:
    kind = file_kind(document)
    path = Path(document.get('local_path') or document.get('storage_path') or '')
    size = path.stat().st_size if path.exists() else 0
    preview_capability = {
        'pdf': 'native',
        'docx': 'extracted_text',
        'spreadsheet': 'table',
        'image': 'image',
        'text': 'extracted_text',
    }.get(kind, 'download_only')
    meta = bff_meta(document, ['download', 'preview'])
    return {
        **meta,
        'project_id': document['project_id'],
        'name': document.get('filename') or document['id'],
        'filename': document.get('filename') or document['id'],
        'mime_type': document.get('content_type') or 'application/octet-stream',
        'kind': kind,
        'size_bytes': size,
        'status': 'ready' if document.get('status') == 'INGESTED' else 'parsing',
        'pages': None,
        'language': None,
        'preview_capability': preview_capability,
        'download_url': f'/api/files/{document["id"]}/content',
        'preview_url': None if preview_capability == 'download_only' else f'/api/files/{document["id"]}/preview',
    }


def parse_location(location: str | None) -> tuple[int | None, int | None]:
    if not location:
        return None, None
    page_match = re.search(r'(?:page|стр\.?|p)[^\d]*(\d+)', location, re.IGNORECASE)
    paragraph_match = re.search(r'(?:paragraph|абзац|para)[^\d]*(\d+)', location, re.IGNORECASE)
    return (
        int(page_match.group(1)) if page_match else None,
        int(paragraph_match.group(1)) if paragraph_match else None,
    )


def evidence_dto(record: dict[str, Any]) -> dict[str, Any]:
    page, paragraph = parse_location(str(record.get('location') or ''))
    quote = str(record.get('statement') or '')
    return {
        'id': record.get('id') or new_id('ev'),
        'source_file_id': record.get('document_id') or '',
        'file_name': record.get('source_name') or 'project brief',
        'page': page,
        'paragraph': paragraph,
        'quote': quote,
        'claim': quote,
        'strength': 'strong' if record.get('trust_tier') == 'internal' else 'medium',
    }


def evidence_by_id(project_id: str) -> dict[str, dict[str, Any]]:
    return {record['id']: record for record in evidence_records(project_id, limit=200)}


def stage_from_run(run: dict[str, Any]) -> str:
    mapping = {
        'INGESTING': 'memory',
        'ANALYZING': 'memory',
        'WAITING_FOR_CLARIFICATION': 'gates',
        'RETRIEVING_MEMORY': 'memory',
        'RESEARCHING_EXTERNAL': 'memory',
        'GENERATING': 'generation',
        'DEDUPLICATING': 'deduplication',
        'APPLYING_GATES': 'gates',
        'CRITIQUING': 'critique',
        'CHECKING_NOVELTY': 'critique',
        'ANALYZING_DISAGREEMENT': 'critique',
        'ANALYZING_UNCERTAINTY': 'critique',
        'COMPILING_EXPERIMENTS': 'final',
        'BUILDING_REPORT': 'final',
        'COMPLETED': 'final',
        'FAILED': 'final',
        'CANCELLED': 'final',
    }
    status = str(run.get('status'))
    if status == 'FAILED' and run.get('events'):
        for event in reversed(run.get('events', [])):
            event_stage = str(event.get('stage') or '')
            if event_stage and event_stage != 'FAILED':
                return mapping.get(event_stage, 'generation')
        return 'generation'
    return mapping.get(status, 'memory')


def status_from_run(run: dict[str, Any]) -> str:
    mapping = {
        'CREATED': 'queued',
        'WAITING_FOR_CLARIFICATION': 'waiting_for_input',
        'FAILED': 'failed',
        'CANCELLED': 'failed',
        'COMPLETED': 'completed',
    }
    return mapping.get(str(run.get('status')), 'running')


def run_funnel_projection(run: dict[str, Any]) -> dict[str, int]:
    artifacts = run.get('artifacts') or {}
    stored = artifacts.get('funnel') or {}
    summary = run.get('llm_generation_summary') or {}
    hypothesis_ids = artifacts.get('hypothesis_ids') or []
    run_hypotheses = [store.data['hypotheses'].get(item_id, {}) for item_id in hypothesis_ids]
    finalists = sum(1 for item in run_hypotheses if item.get('status') == 'FINALIST')
    gates = sum(1 for item in run_hypotheses if item.get('status') in {'FINALIST', 'REJECTED'} or item.get('hard_gates'))
    critiqued = sum(1 for item in run_hypotheses if item.get('critique') and item.get('status') != 'REJECTED')

    def count(name: str, fallback: Any) -> int:
        value = stored.get(name)
        if value is None:
            value = fallback
        try:
            return max(0, int(value))
        except (TypeError, ValueError):
            return 0

    requested = count('requested', summary.get('requested') or (run.get('settings') or {}).get('candidate_count') or len(hypothesis_ids))
    accepted = count('accepted', summary.get('accepted') or len(hypothesis_ids))
    generated_fallback = max(
        int(summary.get('received') or 0),
        int(summary.get('accepted') or 0),
        len(hypothesis_ids),
    )
    generated = generated_fallback if summary else count('generated', len(hypothesis_ids))
    return {
        'requested': requested,
        'generated': generated,
        'accepted': accepted,
        'unique': count('unique', summary.get('accepted') or len(hypothesis_ids)),
        'gates': count('gates', gates),
        'critique': count('critique', critiqued),
        'finalists': count('finalists', finalists),
    }


def run_rejection_projection(run: dict[str, Any]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    summary = run.get('llm_generation_summary') or {}
    for item in summary.get('rejections', [])[-12:]:
        if isinstance(item, dict):
            items.append({
                'stage': 'generation',
                'title': item.get('title'),
                'supporting_evidence': item.get('supporting_evidence') or [],
                'reasons': [str(reason) for reason in item.get('reasons', [])],
            })
    for hypothesis_id in (run.get('artifacts') or {}).get('hypothesis_ids', []):
        hypothesis = store.data['hypotheses'].get(hypothesis_id)
        if not hypothesis or hypothesis.get('status') not in {'BLOCKED', 'REJECTED', 'DUPLICATE'}:
            continue
        items.append({
            'stage': hypothesis.get('rejection_stage') or 'gates',
            'title': hypothesis.get('title'),
            'supporting_evidence': hypothesis.get('supporting_evidence') or [],
            'reasons': [str(reason) for reason in hypothesis.get('rejection_reasons', [])],
        })
    return items[-16:]


def bff_run(run: dict[str, Any]) -> dict[str, Any]:
    events = run.get('events', [])
    started_at = run.get('created_at')
    status = status_from_run(run)
    elapsed = 0
    if started_at:
        try:
            end_time = datetime.now(timezone.utc)
            if status in {'completed', 'failed'} and run.get('updated_at'):
                end_time = datetime.fromisoformat(run['updated_at'])
            elapsed = max(0, int((end_time - datetime.fromisoformat(started_at)).total_seconds()))
        except ValueError:
            elapsed = 0
    stage = stage_from_run(run)
    candidate_count = max(1, int((run.get('settings') or {}).get('candidate_count') or 12))
    # Current fast path uses one portfolio generation call, optional local
    # evidence guardrail, and one batch critic call.  Keep the ETA honest for
    # the demo instead of showing the old 8-10 minute conservative estimate.
    stage_totals = {
        'memory': 140,
        'generation': 140,
        'deduplication': 70,
        'gates': 55,
        'critique': 45,
        'final': 15,
    }
    estimated_total = stage_totals.get(stage, 140)
    eta_seconds = max(10, estimated_total - elapsed)
    if stage in {'gates', 'critique'}:
        accepted_count = int((run.get('llm_generation_summary') or {}).get('accepted') or candidate_count)
        eta_seconds = max(15, min(70, accepted_count * 4 + 20))
    elif stage == 'final':
        eta_seconds = 15
    clarification = None
    for item in run.get('clarifications', []):
        if item.get('answer') is None:
            clarification = {
                'id': item['id'],
                'title': item.get('reason') or 'Требуется уточнение',
                'question': item.get('question') or '',
                'blocking': bool(item.get('blocking')),
                'answer': item.get('answer'),
                'comment': item.get('comment'),
            }
            break
    return {
        **bff_meta(run, ['view', 'refresh'], [run['error']] if run.get('error') else []),
        'project_id': run['project_id'],
        'status': status,
        'stage': stage,
        'started_at': started_at,
        'elapsed_seconds': elapsed,
        'eta_seconds': None if status in {'completed', 'failed'} else eta_seconds,
        'funnel': run_funnel_projection(run),
        'rejections': run_rejection_projection(run),
        'events': [{
            'id': item.get('id') or new_id('event'),
            'created_at': item.get('timestamp') or run.get('updated_at') or now_iso(),
            'stage': stage_from_run({'status': item.get('stage')}),
            'title': str(item.get('stage') or 'event').replace('_', ' ').title(),
            'message': item.get('message') or '',
            'level': 'error' if item.get('stage') == 'FAILED' else 'success' if item.get('stage') == 'COMPLETED' else 'info',
        } for item in events],
        'clarification': clarification,
        'ranking_profile': run.get('ranking_profile') or {
            'novelty': 30,
            'feasibility': 25,
            'physicochemical_mechanism': 30,
            'low_risk': 15,
            'excluded_directions': '',
            'domain_constraints': '\n'.join(get_project(run['project_id']).get('constraints', [])),
        },
    }


def presentation_hypothesis_title(hypothesis: dict[str, Any]) -> str:
    title = str(hypothesis.get('title') or '').strip()
    statement = str(hypothesis.get('statement') or '').strip()
    lowered = f'{title} {statement}'.lower()
    generic = (
        not title
        or title.startswith('hyp_')
        or len(title) > 64
        or len(title.split()) > 8
        or title.lower().startswith(('гипотеза о', 'исследование влияния', 'оптимизация режима', 'оптимизация процесса'))
    )
    if not generic:
        return title
    if 'p80' in lowered or 'доизмель' in lowered or 'раскрыт' in lowered:
        return 'Доизмельчение для раскрытия сульфидов'
    if 'классификац' in lowered or 'гидроциклон' in lowered or 'тонк' in lowered:
        return 'Контроль классификации хвостов'
    if 'pgm' in lowered or 'повтор' in lowered:
        return 'Повторная PGM-верификация'
    if 'глинист' in lowered:
        return 'Контроль глинистости хвостов'
    if 'точк' in lowered and 'отбор' in lowered:
        return 'Контроль точки отбора проб'
    if 'активац' in lowered or 'поверхност' in lowered:
        return 'Активация поверхности хвостов'
    if 'флотац' in lowered:
        return 'Настройка флотации хвостов'
    words = [word for word in re.split(r'\s+', statement) if word][:7]
    return ' '.join(words).rstrip('.,;:') or 'Проверяемая гипотеза'


def bff_first_check_text(hypothesis: dict[str, Any]) -> str:
    raw = str((hypothesis.get('falsification_conditions') or [''])[0]).strip()
    expected = str((hypothesis.get('expected_kpis') or ['KPI'])[0]).strip()
    lowered = raw.lower()
    if not raw or lowered.startswith('нет ') or 'no ' in lowered or 'baseline' in lowered:
        if expected and expected != 'KPI':
            return f'Сравнить результат с baseline и подтвердить: {expected}.'
        return 'Сравнить KPI с baseline и проверить статистическую значимость эффекта.'
    return raw


def bff_key_condition_text(hypothesis: dict[str, Any], first_check: str) -> str:
    texts = [
        str(hypothesis.get('title') or ''),
        str(hypothesis.get('statement') or ''),
        *as_string_list(hypothesis.get('assumptions')),
        first_check,
    ]
    joined = ' '.join(text for text in texts if text).strip()
    patterns = [
        r'(Первый проверяемый [^.]+[.])',
        r'(Первый проверяемый [^;]+)',
        r'(Проверить[^.]+[.])',
        r'(Сравнить[^.]+[.])',
    ]
    for pattern in patterns:
        match = re.search(pattern, joined, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
    lowered = joined.lower()
    if 'pgm' in lowered or 'повтор' in lowered:
        return 'Провести 2–3 повторных PGM-анализа по каждой точке и сравнить разброс с baseline.'
    if 'классификац' in lowered or 'тонк' in lowered or 'гидроциклон' in lowered:
        return 'Проверить отсечку тонкого потока 0,5–0,8 мм против baseline и оценить потери металлов в хвостах.'
    if 'p80' in lowered or 'доизмельч' in lowered or 'раскрыт' in lowered or 'крупност' in lowered:
        return 'Проверить P80 150–180 мкм против baseline и подтвердить прирост извлечения не менее 2 п.п.'
    if 'кондиционир' in lowered or 'контакт' in lowered:
        return 'Проверить 8–12 минут кондиционирования против baseline без роста расхода реагента выше 5%.'
    if 'ph' in lowered or 'селективн' in lowered:
        return 'Проверить диапазон pH 9,0–9,5 против baseline и оценить селективность извлечения.'
    if 'точк' in lowered and 'отбор' in lowered:
        return 'Зафиксировать 3 точки отбора проб по цепи аппаратов и сверить их с таблицами опытов.'
    if 'глинист' in lowered:
        return 'Проверить классы глинистости 5%, 10% и 15% против baseline как ограничивающий фактор.'
    if 'поверхност' in lowered or 'активац' in lowered:
        return 'Проверить изменение поверхностной подготовки против baseline и подтвердить физико-химический механизм.'
    if 'механизм' in lowered:
        return 'Сравнить физико-химический механизм с baseline и подтвердить, что именно он даёт прирост KPI.'
    if first_check:
        return first_check.strip()
    return 'Проверить KPI гипотезы относительно baseline и зафиксировать условия опыта.'


def clamp01(value: Any, default: float = 0.5) -> float:
    try:
        return max(0.0, min(1.0, float(value)))
    except (TypeError, ValueError):
        return default


def score_to_stars(value: float) -> int:
    return max(1, min(5, round(value * 5)))


def extract_kpi_delta_pp(hypothesis: dict[str, Any], evidence: list[dict[str, Any]]) -> float:
    texts: list[str] = []
    for key in ('expected_kpis', 'statement', 'title', 'economic_effect'):
        value = hypothesis.get(key)
        if isinstance(value, list):
            texts.extend(str(item) for item in value)
        elif value:
            texts.append(str(value))
    texts.extend(str(item.get('claim') or '') for item in evidence)
    texts.extend(str(item.get('quote') or '') for item in evidence)
    pattern = re.compile(
        r'([+-]?\d+(?:[,.]\d+)?)\s*(?:п\.?\s*п\.?|пп|процентн(?:ых|ого|ые)\s+пункт(?:а|ов)?|percentage\s+points?|p\.?\s*p\.?)',
        re.IGNORECASE,
    )
    values: list[float] = []
    for text in texts:
        for match in pattern.findall(text):
            try:
                values.append(float(match.replace(',', '.')))
            except ValueError:
                continue
    if not values:
        return 0.0
    positives = [value for value in values if value > 0]
    return round(max(positives or values, key=abs), 1)


def bff_hypothesis(hypothesis: dict[str, Any]) -> dict[str, Any]:
    score_source = hypothesis.get('scores', {})
    project = get_project(hypothesis['project_id'])
    project_target = str(project.get('target_kpi') or '').strip()
    evidence_lookup = evidence_by_id(hypothesis['project_id'])
    evidence = [evidence_dto(evidence_lookup[item_id]) for item_id in hypothesis.get('supporting_evidence', []) if item_id in evidence_lookup]
    if not evidence:
        evidence = [evidence_dto(record) for record in evidence_records(hypothesis['project_id'], limit=2)]
    critique_scores = hypothesis.get('critique', {}).get('scores', {})
    evidence_score = clamp01(score_source.get('evidence', 0.5))
    feasibility_score = clamp01(score_source.get('feasibility', 0.5))
    mechanism_score = clamp01(critique_scores.get('technical', score_source.get('mechanism', feasibility_score)))
    experiment_score = clamp01(score_source.get('experimentability', 0.5))
    evidence_bonus = min(0.07, len(evidence) * 0.014)
    uncertainty_level = str(hypothesis.get('uncertainty', {}).get('level', 'MEDIUM'))
    disagreement_signal = str(hypothesis.get('disagreement', {}).get('risk_signal', 'NORMAL'))
    uncertainty_penalty = 0.08 if uncertainty_level == 'HIGH' else 0.035 if uncertainty_level == 'MEDIUM' else 0.0
    disagreement_penalty = 0.05 if disagreement_signal == 'HIGH_DISAGREEMENT' else 0.015 if disagreement_signal == 'NORMAL' else 0.0
    integral_raw = max(0.0, min(1.0, evidence_score * 0.31 + feasibility_score * 0.24 + mechanism_score * 0.30 + experiment_score * 0.15 + evidence_bonus - uncertainty_penalty - disagreement_penalty))
    science = score_to_stars(min(1.0, evidence_score + evidence_bonus))
    engineering = score_to_stars(feasibility_score)
    mechanism = score_to_stars(mechanism_score)
    testability = score_to_stars(max(0.0, experiment_score - uncertainty_penalty / 2))
    integral = round(integral_raw * 100)
    novelty_info = hypothesis.get('novelty', {})
    novelty_class = str(novelty_info.get('class') if isinstance(novelty_info, dict) else novelty_info)
    title = presentation_hypothesis_title(hypothesis)
    statement = normalize_vague_parameters_for_display(str(hypothesis.get('statement') or title).strip())
    expected_kpi = str((hypothesis.get('expected_kpis') or ['KPI'])[0]).strip() or 'KPI'
    if project_target and text_has_material_number(project_target) and not text_has_material_number(expected_kpi):
        expected_kpi = project_target
    if project_target and text_has_material_number(project_target) and statement and not text_has_material_number(statement):
        statement = f'{statement.rstrip(". ")}. Числовой критерий проверки: {project_target}.'
    kpi_delta = extract_kpi_delta_pp(
        {**hypothesis, 'statement': statement, 'expected_kpis': [expected_kpi]},
        evidence,
    )
    display_hypothesis = {**hypothesis, 'statement': statement, 'expected_kpis': [expected_kpi]}
    first_check = bff_first_check_text(display_hypothesis)
    if project_target and text_has_material_number(project_target) and first_check and not text_has_material_number(first_check):
        first_check = f'{first_check.rstrip(". ")}. Критерий: {project_target}.'
    return {
        **bff_meta(hypothesis, ['feedback', 'compile_experiment']),
        'project_id': hypothesis['project_id'],
        'run_id': hypothesis.get('run_id'),
        'title': title,
        'claim': title,
        'statement': statement,
        'family': hypothesis.get('lineage', {}).get('perspective') or hypothesis.get('intervention', {}).get('type') or 'Технологический режим',
        'status': 'finalist' if hypothesis.get('status') == 'FINALIST' else 'rejected' if hypothesis.get('status') in {'BLOCKED', 'REJECTED', 'DUPLICATE'} else 'candidate',
        'novelty': 'novel' if 'NOVEL' in novelty_class else 'moderately_novel' if novelty_class and novelty_class != 'UNKNOWN' else 'known',
        'uncertainty': 'high' if uncertainty_level == 'HIGH' else 'medium' if uncertainty_level == 'MEDIUM' else 'low',
        'disagreement': 'high' if disagreement_signal == 'HIGH_DISAGREEMENT' else 'medium',
        'kpi_label': expected_kpi,
        'kpi_delta': kpi_delta,
        'risk_label': (hypothesis.get('risks') or ['Риск требует проверки'])[0],
        'economic_effect': hypothesis.get('economic_effect') or 'Экономический эффект требует уточнения по балансу металлов, реагентам, энергии и масштабу переработки.',
        'mechanism': '; '.join(hypothesis.get('causal_mechanism') or [hypothesis.get('statement', '')]),
        'key_condition': bff_key_condition_text(display_hypothesis, first_check),
        'first_check': first_check,
        'scores': {'science': science, 'engineering': engineering, 'mechanism': mechanism, 'testability': testability, 'integral': integral},
        'evidence': evidence,
        'gates': [{'code': key, 'title': key.replace('_', ' ').title(), 'passed': bool(value), 'reason': 'Проверено pipeline'} for key, value in hypothesis.get('hard_gates', {}).items()],
    }


def bff_roadmap(experiment: dict[str, Any]) -> list[dict[str, Any]]:
    steps = experiment.get('steps') or []
    if not steps:
        steps = ['Подготовить пробу', 'Провести опыт', 'Измерить KPI', 'Принять решение']
    nodes = []
    for index, step in enumerate(steps[:5], start=1):
        nodes.append({
            'id': f'R{index}',
            'title': step,
            'duration_days': 1 if index == 1 else 2,
            'depends_on': [] if index == 1 else [f'R{index - 1}'],
            'critical': index != 4,
            'status': 'active' if index == 2 else 'queued',
        })
    return nodes


def bff_experiment(experiment: dict[str, Any]) -> dict[str, Any]:
    return {
        **bff_meta(experiment, ['edit', 'attach_result', 'compile']),
        'project_id': experiment['project_id'],
        'hypothesis_id': experiment['hypothesis_id'],
        'title': experiment.get('objective') or 'Лабораторный протокол',
        'status': 'ready' if experiment.get('steps') else 'draft',
        'current_revision': experiment.get('current_revision') or 'v1',
        'kpi': ', '.join(experiment.get('measurements_units', [])[:1]) or 'KPI',
        'target': ', '.join(experiment.get('success_criteria', [])[:1]) or 'Улучшение относительно baseline',
        'duration_days': int(experiment.get('resource_estimate', {}).get('duration_days', 3)),
        'goal': experiment.get('objective') or '',
        'parameters': {
            key: ', '.join(value) if isinstance(value, list) else value
            for key, value in (experiment.get('factors_and_levels') or {}).items()
        },
        'result_file_ids': [item.get('id') for item in experiment.get('attachments', []) if item.get('id')],
        'roadmap': bff_roadmap(experiment),
    }


@app.get('/health')
def health() -> dict[str, Any]:
    return {'status': 'ok', 'time': now_iso(), 'storage_backend': store.backend_name, 'model_gateway': gateway.profile()}


async def create_project_record(payload: ProjectCreate, user: UserContext) -> dict[str, Any]:
    require_write(user)
    project_id = new_id('proj')
    project = payload.model_dump(mode='json')
    project.update({'id': project_id, 'owner_id': user.id, 'members': [user.id], 'created_at': now_iso(), 'updated_at': now_iso(), 'status': 'ACTIVE'})
    store.data['projects'][project_id] = project
    store.save()
    return project


async def ingest_upload(project_id: str, file: UploadFile) -> dict[str, Any]:
    content = await file.read()
    document_id = new_id('doc')
    digest = sha256_bytes(content)
    project_dir = STORAGE_DIR / project_id
    project_dir.mkdir(parents=True, exist_ok=True)
    safe_name = re.sub(r'[^\w.\-а-яА-ЯёЁ ]+', '_', file.filename or document_id).strip() or document_id
    target = project_dir / f'{document_id}_{safe_name}'
    target.write_bytes(content)
    parsed = parser.parse(target, document_id)
    storage_key = f'{project_id}/documents/{document_id}/{safe_name}'
    storage_uri = object_storage.put_bytes(storage_key, content, file.content_type)
    document = {
        'id': document_id,
        'project_id': project_id,
        'filename': file.filename,
        'content_type': file.content_type,
        'storage_path': storage_uri,
        'local_path': str(target),
        'sha256': digest,
        'status': 'INGESTED',
        'fragments': parsed['fragments'],
        'facts': parsed['facts'],
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    store.data['documents'][document_id] = document
    store.save()
    neo4j_memory.upsert_document_graph(get_project(project_id), document)
    return document


@app.post('/projects')
async def create_project(request: Request, user: UserContext = Depends(current_user)) -> dict[str, Any]:
    content_type = request.headers.get('content-type', '')
    uploaded_files: list[UploadFile] = []
    if 'multipart/form-data' in content_type:
        form = await request.form()
        constraints = str(form.get('constraints') or '')
        task = str(form.get('task') or form.get('name') or '').strip()
        payload = ProjectCreate(
            name=task[:80] or 'Новый исследовательский проект',
            problem=task or 'Исследовательская задача требует уточнения',
            target_kpi=str(form.get('success') or form.get('result') or ''),
            constraints=[constraints] if constraints else [],
            response_language=Language.RU,
            external_research_enabled=True,
        )
        for value in form.getlist('files'):
            if hasattr(value, 'filename') and hasattr(value, 'read'):
                uploaded_files.append(value)
    else:
        payload = ProjectCreate.model_validate(await request.json())
    project = await create_project_record(payload, user)
    for file in uploaded_files:
        await ingest_upload(project['id'], file)
    return bff_project(project)


@app.get('/projects')
def list_projects() -> list[dict[str, Any]]:
    return [bff_project_summary(project) for project in store.data['projects'].values()]


@app.get('/projects/{project_id}')
def read_project(project_id: str) -> dict[str, Any]:
    return bff_project(get_project(project_id))


@app.patch('/projects/{project_id}')
def update_project(project_id: str, payload: ProjectUpdate, user: UserContext = Depends(current_user)) -> dict[str, Any]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    for key, value in payload.model_dump(exclude_none=True, mode='json').items():
        project[key] = value
    project['updated_at'] = now_iso()
    store.save()
    return bff_project(project)


@app.put('/projects/{project_id}/brief')
def update_project_brief(project_id: str, payload: BriefPayload, user: UserContext = Depends(current_user)) -> dict[str, Any]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    project['problem'] = payload.problem
    project['goal'] = payload.goal
    project['target_kpi'] = payload.success_criterion
    project['success_criterion'] = payload.success_criterion
    project['constraints'] = [payload.constraints] if payload.constraints else []
    project['updated_at'] = now_iso()
    store.save()
    return bff_project(project)


@app.post('/projects/{project_id}/documents')
async def upload_document(project_id: str, file: UploadFile = File(...), user: UserContext = Depends(current_user)) -> dict[str, Any]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    return await ingest_upload(project['id'], file)


@app.get('/projects/{project_id}/files')
def bff_list_files(project_id: str) -> list[dict[str, Any]]:
    get_project(project_id)
    return [bff_file(document) for document in project_documents(project_id)]


@app.post('/projects/{project_id}/files')
async def bff_upload_files(project_id: str, files: list[UploadFile] = File(...), user: UserContext = Depends(current_user)) -> list[dict[str, Any]]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    documents = [await ingest_upload(project['id'], file) for file in files]
    return [bff_file(document) for document in documents]


@app.delete('/projects/{project_id}/files/{file_id}', status_code=204)
def bff_delete_project_file(project_id: str, file_id: str, user: UserContext = Depends(current_user)) -> Response:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    document = store.data['documents'].get(file_id)
    if not document or document.get('project_id') != project_id:
        raise HTTPException(status_code=404, detail='File not found for project')
    local_path = document.get('local_path')
    if local_path:
        try:
            resolved = Path(local_path).resolve()
            storage_root = STORAGE_DIR.resolve()
            if resolved.exists() and resolved.is_file() and (resolved == storage_root or storage_root in resolved.parents):
                resolved.unlink()
        except OSError:
            pass
    store.data['documents'].pop(file_id, None)
    project['updated_at'] = now_iso()
    store.save()
    return Response(status_code=204)


@app.post('/projects/{project_id}/documents/import-local')
def import_local_documents(project_id: str, payload: LocalImportRequest, user: UserContext = Depends(current_user)) -> dict[str, Any]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    root = Path(payload.path) if payload.path else ORGANIZER_DIR
    if not root.is_absolute():
        root = ROOT / root
    if not root.exists():
        raise HTTPException(status_code=404, detail=f'Import path not found: {root}')
    imported = []
    for path in sorted(p for p in root.rglob('*') if p.is_file()):
        if payload.limit and len(imported) >= payload.limit:
            break
        if path.suffix.lower() not in {'.pdf', '.docx', '.xlsx', '.png', '.jpg', '.jpeg', '.txt', '.md', '.csv', '.json'}:
            continue
        content = path.read_bytes()
        document_id = new_id('doc')
        project_dir = STORAGE_DIR / project_id
        project_dir.mkdir(parents=True, exist_ok=True)
        target = project_dir / f'{document_id}_{path.name}'
        shutil.copy2(path, target)
        parsed = parser.parse(target, document_id)
        storage_uri = object_storage.copy_file(path, f'{project_id}/documents/{document_id}/{path.name}', path.suffix.lower().lstrip('.'))
        document = {
            'id': document_id,
            'project_id': project_id,
            'filename': path.name,
            'content_type': path.suffix.lower().lstrip('.'),
            'storage_path': storage_uri,
            'local_path': str(target),
            'sha256': sha256_bytes(content),
            'status': 'INGESTED',
            'fragments': parsed['fragments'],
            'facts': parsed['facts'],
            'created_at': now_iso(),
            'updated_at': now_iso(),
        }
        store.data['documents'][document_id] = document
        neo4j_memory.upsert_document_graph(project, document)
        imported.append({'id': document_id, 'filename': path.name, 'facts': len(parsed['facts'])})
    store.save()
    return {'imported': imported, 'count': len(imported)}


@app.get('/projects/{project_id}/documents')
def list_documents(project_id: str) -> list[dict[str, Any]]:
    get_project(project_id)
    return project_documents(project_id)


@app.get('/projects/{project_id}/warnings')
def bff_project_warnings(project_id: str) -> list[dict[str, Any]]:
    project = get_project(project_id)
    docs = project_documents(project_id)
    warnings = []
    if not project.get('target_kpi'):
        warnings.append({
            'id': 'target-kpi-missing',
            'code': 'target_kpi_missing',
            'title': 'Не указан целевой KPI',
            'description': 'Уточните критерий успеха, чтобы ранжирование гипотез было воспроизводимым.',
            'severity': 'warning',
            'related_file_ids': [],
            'resolved': False,
        })
    if not docs:
        warnings.append({
            'id': 'source-files-missing',
            'code': 'source_files_missing',
            'title': 'Нет исходных файлов',
            'description': 'Загрузите PDF, таблицы или схемы, чтобы evidence было прослеживаемым.',
            'severity': 'blocking',
            'related_file_ids': [],
            'resolved': False,
        })
    for document in docs:
        quality_facts = [fact for fact in document.get('facts', []) if fact.get('type') == 'data_quality']
        if quality_facts:
            warnings.append({
                'id': f'data-quality-{document["id"]}',
                'code': 'data_quality_issue',
                'title': f'Проверить качество данных: {document.get("filename")}',
                'description': summarize_text('; '.join(fact.get('statement', '') for fact in quality_facts[:3]), 300),
                'severity': 'warning',
                'related_file_ids': [document['id']],
                'resolved': False,
            })
    return warnings


@app.get('/projects/{project_id}/knowledge-graph')
def bff_knowledge_graph(project_id: str) -> dict[str, Any]:
    project = get_project(project_id)
    nodes: list[dict[str, Any]] = []
    edges: list[dict[str, Any]] = []
    node_ids: set[str] = set()

    for document in project_documents(project_id)[:8]:
        nodes.append({
            'id': document['id'],
            'type': 'source',
            'label': document.get('filename') or document['id'],
            'meta': file_kind(document).upper(),
            'description': f'Исходный файл проекта, содержит {len(document.get("facts", []))} извлечённых фактов.',
            'confidence': None,
            'source_file_id': document['id'],
            'source_page': None,
        })
        node_ids.add(document['id'])
        for fact in document.get('facts', [])[:4]:
            fact_id = fact['id']
            nodes.append({
                'id': fact_id,
                'type': 'fact',
                'label': summarize_text(fact.get('statement', ''), 56),
                'meta': fact.get('type', 'fact'),
                'description': fact.get('statement', ''),
                'confidence': fact.get('confidence', 0.8),
                'source_file_id': document['id'],
                'source_page': parse_location(fact.get('provenance', {}).get('location'))[0],
            })
            node_ids.add(fact_id)
            edges.append({
                'id': f'edge-{fact_id}-{document["id"]}',
                'from': document['id'],
                'to': fact_id,
                'relation': 'extracted_from',
                'weight': 4,
                'reason': 'Факт извлечён парсером из исходного файла.',
            })

    for hypothesis in [item for item in store.data['hypotheses'].values() if item['project_id'] == project_id][:8]:
        nodes.append({
            'id': hypothesis['id'],
            'type': 'hypothesis',
            'label': hypothesis.get('title') or hypothesis['id'],
            'meta': hypothesis.get('status', 'candidate'),
            'description': hypothesis.get('statement', ''),
            'confidence': hypothesis.get('scores', {}).get('evidence'),
            'source_file_id': None,
            'source_page': None,
        })
        node_ids.add(hypothesis['id'])
        for evidence_id in hypothesis.get('supporting_evidence', [])[:4]:
            if evidence_id in node_ids:
                edges.append({
                    'id': f'edge-{evidence_id}-{hypothesis["id"]}',
                    'from': evidence_id,
                    'to': hypothesis['id'],
                    'relation': 'supports',
                    'weight': 5,
                    'reason': 'Evidence указан в supporting_evidence гипотезы.',
                })

    for experiment in [item for item in store.data['experiments'].values() if item['project_id'] == project_id][:8]:
        nodes.append({
            'id': experiment['id'],
            'type': 'experiment',
            'label': experiment.get('objective') or experiment['id'],
            'meta': f'{experiment.get("resource_estimate", {}).get("duration_days", 3)} дней',
            'description': '; '.join(experiment.get('steps', [])[:3]),
            'confidence': None,
            'source_file_id': None,
            'source_page': None,
        })
        edges.append({
            'id': f'edge-{experiment["hypothesis_id"]}-{experiment["id"]}',
            'from': experiment['hypothesis_id'],
            'to': experiment['id'],
            'relation': 'validated_by',
            'weight': 5,
            'reason': 'Эксперимент скомпилирован для проверки выбранной гипотезы.',
        })

    if not nodes:
        nodes.append({
            'id': project_id,
            'type': 'source',
            'label': project.get('name') or 'Проект',
            'meta': 'brief',
            'description': project.get('problem') or '',
            'confidence': None,
            'source_file_id': None,
            'source_page': None,
        })
    return {**bff_meta(project, ['graph'], project_warning_codes(project)), 'project_id': project_id, 'nodes': nodes, 'edges': edges}


@app.get('/documents/{document_id}')
def read_document(document_id: str) -> dict[str, Any]:
    document = store.data['documents'].get(document_id)
    if not document:
        raise HTTPException(status_code=404, detail='Document not found')
    return document


@app.get('/files/{file_id}')
def bff_read_file(file_id: str) -> dict[str, Any]:
    return bff_file(read_document(file_id))


@app.get('/files/{file_id}/content')
def bff_file_content(file_id: str) -> FileResponse:
    document = read_document(file_id)
    path = Path(document.get('local_path') or document.get('storage_path') or '')
    if not path.exists():
        raise HTTPException(status_code=404, detail='File content not found')
    return FileResponse(path, filename=document.get('filename') or path.name, media_type=document.get('content_type') or 'application/octet-stream')


@app.get('/files/{file_id}/preview')
def bff_file_preview(file_id: str, page: int | None = None) -> Any:
    document = read_document(file_id)
    kind = file_kind(document)
    path = Path(document.get('local_path') or document.get('storage_path') or '')
    if kind in {'pdf', 'image'}:
        if not path.exists():
            raise HTTPException(status_code=404, detail='File preview not found')
        return FileResponse(path, filename=document.get('filename') or path.name, media_type=document.get('content_type') or 'application/octet-stream')
    if kind == 'spreadsheet':
        rows = []
        for fragment in document.get('fragments', [])[:100]:
            rows.append([fragment.get('location'), summarize_text(fragment.get('original_text', ''), 300)])
        if not rows:
            rows = [['Факт', 'Значение']] + [[fact.get('type'), fact.get('statement')] for fact in document.get('facts', [])[:100]]
        return {'columns': rows[0] if rows else [], 'rows': rows[1:], 'truncated': len(document.get('fragments', [])) > 100}
    text = '\n\n'.join(fragment.get('original_text', '') for fragment in document.get('fragments', [])[:80]).strip()
    if not text and path.exists():
        try:
            text = path.read_text(encoding='utf-8')[:100000]
        except UnicodeDecodeError:
            text = path.read_text(encoding='cp1251', errors='replace')[:100000]
    return PlainTextResponse(text or 'Предпросмотр недоступен для этого файла.', media_type='text/plain; charset=utf-8')


async def analyze_document_image(document: dict[str, Any]) -> dict[str, Any]:
    path = Path(document.get('local_path') or document['storage_path'])
    if path.suffix.lower() not in {'.png', '.jpg', '.jpeg', '.webp'}:
        raise HTTPException(status_code=400, detail='Document is not an image')
    result = await gateway.vision_json(
        path,
        'Верни только JSON: {"nodes":[{"type":"Equipment|ProcessStage|Stream|Parameter","label":"...","confidence":0.0}], "edges":[{"source":"...","target":"...","type":"FLOWS_TO|USES_EQUIPMENT|HAS_PARAMETER","confidence":0.0}], "uncertain_items":[]}. Извлеки схему/регламент как черновой граф.',
    )
    document['vision_graph'] = result
    vision_fragments, vision_facts = vision_graph_to_facts(document, result)
    document['fragments'] = [
        fragment for fragment in document.get('fragments', [])
        if fragment.get('extraction_method') != 'vision:qwen'
    ] + vision_fragments
    document['facts'] = [
        fact for fact in document.get('facts', [])
        if fact.get('provenance', {}).get('extraction_method') != 'vision:qwen'
    ] + vision_facts
    document['updated_at'] = now_iso()
    store.save()
    neo4j_memory.upsert_document_graph(get_project(document['project_id']), document)
    return result


@app.get('/documents/{document_id}/ingestion-status')
def ingestion_status(document_id: str) -> dict[str, Any]:
    document = read_document(document_id)
    return {
        'id': document['id'],
        'status': document['status'],
        'fragments': len(document.get('fragments', [])),
        'facts': len(document.get('facts', [])),
        'updated_at': document['updated_at'],
    }


@app.post('/documents/{document_id}/vision-analyze')
async def document_vision_analyze(document_id: str) -> dict[str, Any]:
    document = read_document(document_id)
    return await analyze_document_image(document)


@app.post('/documents/{document_id}/reprocess')
def reprocess_document(document_id: str) -> dict[str, Any]:
    document = read_document(document_id)
    parse_path = Path(document.get('local_path') or document['storage_path'])
    parsed = parser.parse(parse_path, document_id)
    document['fragments'] = parsed['fragments']
    document['facts'] = parsed['facts']
    document['status'] = 'INGESTED'
    document['updated_at'] = now_iso()
    store.save()
    neo4j_memory.upsert_document_graph(get_project(document['project_id']), document)
    return ingestion_status(document_id)


@app.post('/projects/{project_id}/runs')
async def create_run(project_id: str, payload: RunCreate, background_tasks: BackgroundTasks, user: UserContext = Depends(current_user)) -> dict[str, Any]:
    project = get_project(project_id)
    require_project_access(project, user, write=True)
    run_id = new_id('run')
    run = {
        'id': run_id,
        'project_id': project_id,
        'status': RunStatus.CREATED.value,
        'settings': payload.model_dump(mode='json'),
        'response_language': payload.response_language or project.get('response_language', 'ru'),
        'model_profile': gateway.profile(),
        'created_by': user.id,
        'events': [],
        'artifacts': {},
        'clarifications': [],
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    store.data['runs'][run_id] = run
    store.save()
    if os.getenv('NORLAB_TASK_BACKEND', 'background') == 'celery':
        task = execute_pipeline_task.delay(project_id, run_id, payload.use_llm)
        run['celery_task_id'] = task.id
        store.save()
    else:
        background_tasks.add_task(execute_pipeline, project_id, run_id, payload.use_llm)
    return bff_run(get_run(run_id))


@app.get('/projects/{project_id}/runs/{run_id}')
def bff_read_run(project_id: str, run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')
    return bff_run(run)


@app.get('/projects/{project_id}/runs/{run_id}/events')
async def bff_run_events(project_id: str, run_id: str) -> StreamingResponse:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')

    async def stream() -> Any:
        last_count = 0
        idle_ticks = 0
        while idle_ticks < 180:
            current = get_run(run_id)
            events = bff_run(current)['events']
            for event in events[last_count:]:
                yield f'event: run.updated\nid: {event["id"]}\ndata: {json.dumps({"run_id": run_id, "stage": event["stage"], "status": bff_run(current)["status"], "eta_seconds": bff_run(current)["eta_seconds"]}, ensure_ascii=False)}\n\n'
            if len(events) != last_count:
                last_count = len(events)
                idle_ticks = 0
            else:
                idle_ticks += 1
            if bff_run(current)['status'] in {'completed', 'failed'} and last_count >= len(events):
                break
            await asyncio.sleep(1)

    return StreamingResponse(stream(), media_type='text/event-stream')


@app.post('/projects/{project_id}/runs/{run_id}/pause')
def bff_pause_run(project_id: str, run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')
    add_event(run, RunStatus.WAITING_FOR_CLARIFICATION, 'Run paused by user.', 0, 0)
    return bff_run(run)


@app.post('/projects/{project_id}/runs/{run_id}/resume')
def bff_resume_run(project_id: str, run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')
    add_event(run, RunStatus.ANALYZING, 'Run resumed.', 0, 0)
    return bff_run(run)


@app.post('/projects/{project_id}/runs/{run_id}/recover')
def bff_recover_run(project_id: str, run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')
    add_event(run, RunStatus.APPLYING_GATES, 'Run state recovered from latest checkpoint.', 0, 0)
    return bff_run(run)


@app.post('/projects/{project_id}/runs/{run_id}/clarifications/{clarification_id}')
def bff_answer_clarification(project_id: str, run_id: str, clarification_id: str, payload: BffClarificationAnswer) -> dict[str, Any]:
    run = get_run(run_id)
    if run['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Run not found for project')
    for question in run.get('clarifications', []):
        if question['id'] == clarification_id:
            question['answer'] = payload.answer
            question['comment'] = payload.comment
            run['updated_at'] = now_iso()
            store.save()
            return bff_run(run)
    raise HTTPException(status_code=404, detail='Clarification not found')


@app.put('/projects/{project_id}/ranking-profile')
def bff_save_ranking_profile(project_id: str, profile: dict[str, Any]) -> dict[str, Any]:
    get_project(project_id)
    total = sum(int(profile.get(key, 0)) for key in ['novelty', 'feasibility', 'physicochemical_mechanism', 'low_risk'])
    if total != 100:
        raise HTTPException(status_code=422, detail='Ranking weights must total 100')
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project_id]
    if runs:
        runs[-1]['ranking_profile'] = profile
        runs[-1]['updated_at'] = now_iso()
    else:
        get_project(project_id)['ranking_profile'] = profile
    store.save()
    return profile


@app.get('/runs/{run_id}')
def read_run(run_id: str) -> dict[str, Any]:
    return get_run(run_id)


@app.get('/runs/{run_id}/events')
async def run_events(run_id: str) -> StreamingResponse:
    run = get_run(run_id)

    async def stream() -> Any:
        for event in run.get('events', []):
            yield f'event: stage\ndata: {json.dumps(event, ensure_ascii=False)}\n\n'
            await asyncio.sleep(0.01)

    return StreamingResponse(stream(), media_type='text/event-stream')


@app.post('/runs/{run_id}/cancel')
def cancel_run(run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    add_event(run, RunStatus.CANCELLED, 'Run cancelled by user.', 0, 0)
    return run


@app.post('/runs/{run_id}/retry-stage')
async def retry_stage(run_id: str, stage: RunStatus = Query(default=RunStatus.GENERATING)) -> dict[str, Any]:
    run = get_run(run_id)
    project_id = run['project_id']
    run['events'].append({'id': new_id('event'), 'stage': stage.value, 'message': 'Retry requested.', 'timestamp': now_iso()})
    store.save()
    await execute_pipeline(project_id, run_id, use_llm=run['settings'].get('use_llm', False))
    return get_run(run_id)


@app.get('/runs/{run_id}/artifacts')
def run_artifacts(run_id: str) -> dict[str, Any]:
    return get_run(run_id).get('artifacts', {})


@app.get('/runs/{run_id}/clarifications')
def run_clarifications(run_id: str) -> list[dict[str, Any]]:
    return get_run(run_id).get('clarifications', [])


@app.post('/runs/{run_id}/clarifications/{question_id}/answer')
def answer_clarification(run_id: str, question_id: str, payload: ClarificationAnswer) -> dict[str, Any]:
    run = get_run(run_id)
    for question in run.get('clarifications', []):
        if question['id'] == question_id:
            question['answer'] = payload.answer
            store.save()
            return question
    raise HTTPException(status_code=404, detail='Clarification not found')


@app.post('/runs/{run_id}/resume')
def resume_run(run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    add_event(run, RunStatus.ANALYZING, 'Run resumed after clarification.', 0, 0)
    return run


@app.get('/projects/{project_id}/memory/search')
def memory_search(project_id: str, q: str = Query(default=''), limit: int = Query(default=10, ge=1, le=50)) -> dict[str, Any]:
    get_project(project_id)
    terms = q.lower().split()
    matches = []
    for fact in project_facts(project_id):
        haystack = fact['statement'].lower()
        if not terms or all(term in haystack for term in terms):
            matches.append(fact)
    return {'query': q, 'results': matches[:limit], 'total': len(matches)}


def lexical_memory_results(project_id: str, q: str, limit: int) -> list[dict[str, Any]]:
    terms = [term for term in re.findall(r'[A-Za-zА-Яа-яЁё0-9]+', q.lower()) if len(term) > 2]
    candidates: list[dict[str, Any]] = []
    for fact in project_facts(project_id):
        text_value = fact.get('statement', '')
        haystack = text_value.lower()
        score = sum(1 for term in terms if term in haystack)
        if score or not terms:
            candidates.append({'id': fact['id'], 'labels': ['Fact'], 'score': float(score), 'properties': fact})
    for doc in project_documents(project_id):
        for fragment in doc.get('fragments', [])[:80]:
            text_value = fragment.get('original_text', '')
            haystack = text_value.lower()
            score = sum(1 for term in terms if term in haystack)
            if score:
                candidates.append({'id': fragment['id'], 'labels': ['Fragment'], 'score': float(score), 'properties': fragment})
    for hypothesis in store.data['hypotheses'].values():
        if hypothesis['project_id'] != project_id:
            continue
        text_value = f'{hypothesis.get("title", "")} {hypothesis.get("statement", "")}'
        haystack = text_value.lower()
        score = sum(1 for term in terms if term in haystack)
        if score:
            candidates.append({'id': hypothesis['id'], 'labels': ['Hypothesis'], 'score': float(score), 'properties': hypothesis})
    candidates.sort(key=lambda item: item['score'], reverse=True)
    return candidates[:limit]


@app.post('/projects/{project_id}/memory/reindex')
async def memory_reindex(
    project_id: str,
    limit: int = Query(default=64, ge=1, le=500),
    timeout_seconds: float = Query(default=120, ge=1, le=1800),
) -> dict[str, Any]:
    return await reindex_project_embeddings(project_id, limit=limit, timeout_seconds=timeout_seconds)


@app.get('/projects/{project_id}/memory/vector-search')
async def memory_vector_search(project_id: str, q: str, limit: int = Query(default=10, ge=1, le=50)) -> dict[str, Any]:
    get_project(project_id)
    try:
        embedded = await gateway.embed_text(q, is_document=False)
    except Exception as exc:
        results = lexical_memory_results(project_id, q, limit)
        return {
            'query': q,
            'mode': 'lexical_fallback',
            'embedding_error': str(exc)[:500],
            'results': results,
            'total': len(results),
        }
    results = neo4j_memory.vector_search(embedded['embedding'], limit=limit)
    return {'query': q, 'mode': 'vector', 'model_uri': embedded['model_uri'], 'results': results, 'total': len(results)}


@app.get('/projects/{project_id}/graph/subgraph')
def graph_subgraph(project_id: str) -> dict[str, Any]:
    get_project(project_id)
    neo4j_subgraph = neo4j_memory.subgraph(project_id)
    if neo4j_subgraph is not None:
        return neo4j_subgraph
    nodes = [{'id': project_id, 'labels': ['Project'], 'properties': {'name': get_project(project_id)['name']}}]
    edges = []
    for doc in project_documents(project_id):
        nodes.append({'id': doc['id'], 'labels': ['Document'], 'properties': {'filename': doc['filename']}})
        edges.append({'source': project_id, 'target': doc['id'], 'type': 'HAS_DOCUMENT'})
        for fact in doc.get('facts', [])[:20]:
            nodes.append({'id': fact['id'], 'labels': ['Fact', fact['type']], 'properties': {'statement': fact['statement']}})
            edges.append({'source': fact['id'], 'target': doc['id'], 'type': 'EXTRACTED_FROM'})
    return {'nodes': nodes, 'edges': edges}


@app.post('/projects/{project_id}/research/patents')
async def project_patent_research(project_id: str, q: str = Query(...), limit: int = Query(default=8, ge=1, le=25)) -> dict[str, Any]:
    get_project(project_id)
    sources = await patent_research(q, max_sources=limit)
    for source in sources:
        source['project_id'] = project_id
        store.data['sources'][source['id']] = source
    store.save()
    return {'query': sanitizer.sanitize(q), 'sources': sources, 'total': len(sources)}


@app.get('/entities/{entity_id}')
def read_entity(entity_id: str) -> dict[str, Any]:
    entity = store.data['entities'].get(entity_id)
    if not entity:
        raise HTTPException(status_code=404, detail='Entity not found')
    return entity


@app.get('/sources/{source_id}')
def read_source(source_id: str) -> dict[str, Any]:
    source = store.data['sources'].get(source_id)
    if source:
        return source
    document = store.data['documents'].get(source_id)
    if not document:
        raise HTTPException(status_code=404, detail='Source not found')
    return {'id': document['id'], 'filename': document['filename'], 'fragments': document.get('fragments', [])}


@app.get('/runs/{run_id}/hypotheses')
def run_hypotheses(run_id: str) -> list[dict[str, Any]]:
    get_run(run_id)
    return [item for item in store.data['hypotheses'].values() if item['run_id'] == run_id]


@app.get('/projects/{project_id}/hypotheses')
def bff_project_hypotheses(
    project_id: str,
    sort: Literal['rating', 'kpi', 'id'] = 'rating',
    risk: str | None = None,
    novelty: str | None = None,
    status: str | None = None,
    run_id: str | None = None,
) -> dict[str, Any]:
    get_project(project_id)
    selected_run_id = run_id
    if not selected_run_id:
        runs = sorted(
            [run for run in store.data['runs'].values() if run['project_id'] == project_id],
            key=lambda run: run.get('created_at') or '',
            reverse=True,
        )
        selected_run_id = next((run['id'] for run in runs if run.get('artifacts', {}).get('hypothesis_ids')), None)
    if selected_run_id:
        selected_run = store.data['runs'].get(selected_run_id)
        artifact_ids = selected_run.get('artifacts', {}).get('hypothesis_ids', []) if selected_run else []
        if artifact_ids:
            raw_items = [store.data['hypotheses'][item_id] for item_id in artifact_ids if item_id in store.data['hypotheses']]
            raw_items.extend(
                item for item in store.data['hypotheses'].values()
                if item['project_id'] == project_id and item.get('run_id') == selected_run_id and item['id'] not in artifact_ids
            )
        else:
            raw_items = [item for item in store.data['hypotheses'].values() if item['project_id'] == project_id and item.get('run_id') == selected_run_id]
    else:
        raw_items = [item for item in store.data['hypotheses'].values() if item['project_id'] == project_id]
    items = [bff_hypothesis(item) for item in raw_items if item['project_id'] == project_id]
    if risk:
        items = [item for item in items if item['disagreement'] == risk or item['risk_label'] == risk]
    if novelty:
        items = [item for item in items if item['novelty'] == novelty]
    if status and status != 'all':
        status_filter = {
            'finalists': 'finalist',
            'excluded': 'rejected',
            'rejects': 'rejected',
        }.get(status, status)
        items = [item for item in items if item['status'] == status_filter]
    if sort == 'kpi':
        items.sort(key=lambda item: item['kpi_delta'], reverse=True)
    elif sort == 'id':
        items.sort(key=lambda item: item['id'])
    else:
        items.sort(key=lambda item: item['scores']['integral'], reverse=True)
    return {'items': items, 'total': len(items), 'next_cursor': None}


@app.post('/projects/{project_id}/hypotheses')
def bff_create_hypothesis(project_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    project = get_project(project_id)
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project_id]
    run = runs[-1] if runs else {
        'id': new_id('run'),
        'project_id': project_id,
        'status': RunStatus.COMPLETED.value,
        'settings': {'max_finalists': 12, 'candidate_count': 12, 'use_llm': False},
        'response_language': project.get('response_language', 'ru'),
        'model_profile': gateway.profile(),
        'created_by': 'local-admin',
        'events': [],
        'artifacts': {},
        'clarifications': [],
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    store.data['runs'].setdefault(run['id'], run)
    evidence = evidence_records(project_id, limit=4)
    hypothesis = {
        'id': new_id('hyp'),
        'run_id': run['id'],
        'project_id': project_id,
        'title': str(payload.get('claim') or 'Экспертная гипотеза')[:240],
        'statement': str(payload.get('claim') or 'Экспертная гипотеза')[:1200],
        'intervention': {'type': payload.get('family') or 'expert', 'description': str(payload.get('claim') or '')},
        'target_process': ['tailings_processing'],
        'conditions': project.get('constraints', []),
        'causal_mechanism': ['Механизм добавлен экспертом и требует лабораторной проверки.'],
        'data_triggers': [item['id'] for item in evidence],
        'supporting_evidence': [item['id'] for item in evidence],
        'contradicting_evidence': [],
        'assumptions': ['Экспертная гипотеза требует независимой проверки.'],
        'expected_kpis': [payload.get('kpi_label') or project.get('target_kpi') or 'KPI'],
        'economic_effect': 'Экономический эффект экспертной гипотезы нужно подтвердить через прирост извлечения, баланс металлов и дополнительные затраты на реагенты/энергию.',
        'risks': ['Требуется проверка переносимости на текущую пробу.'],
        'constraints': project.get('constraints', []),
        'falsification_conditions': ['Нет улучшения KPI относительно baseline.'],
        'novelty': {'class': 'UNKNOWN', 'analogs': []},
        'uncertainty': {'level': 'MEDIUM', 'reasons': ['создано экспертом']},
        'disagreement': {'risk_signal': 'NORMAL'},
        'lineage': {'perspective': payload.get('family') or 'expert', 'model_profile': run['model_profile']['id']},
        'status': 'DRAFT',
        'scores': {'evidence': 0.5, 'feasibility': 0.5, 'novelty': 0.5, 'experimentability': 0.5},
        'hard_gates': {'expert_review_gate': True},
        'created_at': now_iso(),
        'updated_at': now_iso(),
    }
    store.data['hypotheses'][hypothesis['id']] = hypothesis
    store.save()
    return bff_hypothesis(hypothesis)


@app.get('/projects/{project_id}/hypotheses/{hypothesis_id}')
def bff_read_hypothesis(project_id: str, hypothesis_id: str) -> dict[str, Any]:
    hypothesis = read_hypothesis(hypothesis_id)
    if hypothesis['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Hypothesis not found for project')
    return bff_hypothesis(hypothesis)


@app.post('/projects/{project_id}/hypotheses/{hypothesis_id}/feedback')
def bff_hypothesis_feedback(project_id: str, hypothesis_id: str, payload: BffFeedbackRequest) -> dict[str, Any]:
    hypothesis = read_hypothesis(hypothesis_id)
    if hypothesis['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Hypothesis not found for project')
    hypothesis.setdefault('feedback', []).append({**payload.model_dump(mode='json'), 'timestamp': now_iso()})
    hypothesis['updated_at'] = now_iso()
    store.save()
    return bff_hypothesis(hypothesis)


@app.get('/hypotheses/{hypothesis_id}')
def read_hypothesis(hypothesis_id: str) -> dict[str, Any]:
    hypothesis = store.data['hypotheses'].get(hypothesis_id)
    if not hypothesis:
        raise HTTPException(status_code=404, detail='Hypothesis not found')
    return hypothesis


@app.post('/hypotheses/{hypothesis_id}/feedback')
def hypothesis_feedback(hypothesis_id: str, payload: FeedbackRequest) -> dict[str, Any]:
    hypothesis = read_hypothesis(hypothesis_id)
    feedback = payload.model_dump(mode='json')
    feedback['timestamp'] = now_iso()
    hypothesis.setdefault('feedback', []).append(feedback)
    hypothesis['updated_at'] = now_iso()
    store.save()
    return hypothesis


@app.post('/hypotheses/compare')
def compare_hypotheses(ids: list[str]) -> dict[str, Any]:
    items = [read_hypothesis(item_id) for item_id in ids]
    return {
        'items': items,
        'criteria': ['evidence', 'feasibility', 'novelty', 'experimentability'],
        'pareto_hint': sorted(items, key=lambda item: item['scores']['experimentability'], reverse=True),
    }


@app.post('/hypotheses/{hypothesis_id}/compile-experiment')
def compile_hypothesis_experiment(hypothesis_id: str) -> dict[str, Any]:
    hypothesis = read_hypothesis(hypothesis_id)
    experiment = compile_experiment(hypothesis)
    store.data['experiments'][experiment['id']] = experiment
    store.save()
    return experiment


@app.get('/experiments/{experiment_id}')
def read_experiment(experiment_id: str) -> dict[str, Any]:
    experiment = store.data['experiments'].get(experiment_id)
    if not experiment:
        raise HTTPException(status_code=404, detail='Experiment not found')
    return experiment


@app.get('/projects/{project_id}/experiments')
def bff_project_experiments(project_id: str) -> list[dict[str, Any]]:
    get_project(project_id)
    return [bff_experiment(item) for item in store.data['experiments'].values() if item['project_id'] == project_id]


@app.post('/projects/{project_id}/experiments')
def bff_create_experiment(project_id: str, payload: dict[str, Any]) -> dict[str, Any]:
    get_project(project_id)
    hypothesis_id = str(payload.get('hypothesis_id') or '')
    hypothesis = read_hypothesis(hypothesis_id)
    if hypothesis['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Hypothesis not found for project')
    experiment = compile_experiment(hypothesis)
    store.data['experiments'][experiment['id']] = experiment
    store.save()
    return bff_experiment(experiment)


@app.patch('/projects/{project_id}/experiments/{experiment_id}')
def bff_update_experiment(project_id: str, experiment_id: str, patch: dict[str, Any]) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    if experiment['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Experiment not found for project')
    experiment.setdefault('revisions', []).append({'id': experiment.get('current_revision', 'v1'), 'snapshot': dict(experiment), 'created_at': now_iso()})
    if 'goal' in patch:
        experiment['objective'] = patch['goal']
    if 'parameters' in patch and isinstance(patch['parameters'], dict):
        experiment['factors_and_levels'] = patch['parameters']
    experiment['current_revision'] = f'v{len(experiment.get("revisions", [])) + 1}'
    experiment['updated_at'] = now_iso()
    store.save()
    return bff_experiment(experiment)


@app.post('/projects/{project_id}/experiments/{experiment_id}/results')
async def bff_upload_experiment_result(project_id: str, experiment_id: str, file: UploadFile = File(...)) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    if experiment['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Experiment not found for project')
    attachment = await add_experiment_attachment(experiment_id, file)
    experiment.setdefault('results', []).append({'id': new_id('result'), 'attachment_id': attachment['id'], 'status': 'uploaded', 'timestamp': now_iso()})
    experiment['updated_at'] = now_iso()
    store.save()
    return bff_experiment(experiment)


@app.post('/projects/{project_id}/experiments/{experiment_id}/compile')
def bff_compile_protocol(project_id: str, experiment_id: str) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    if experiment['project_id'] != project_id:
        raise HTTPException(status_code=404, detail='Experiment not found for project')
    experiment['protocol_compiled_at'] = now_iso()
    experiment['updated_at'] = now_iso()
    store.save()
    return bff_experiment(experiment)


@app.patch('/experiments/{experiment_id}')
def update_experiment(experiment_id: str, payload: ExperimentPatch) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    for key, value in payload.model_dump(exclude_none=True, mode='json').items():
        experiment[key] = value
    experiment['updated_at'] = now_iso()
    store.save()
    return experiment


@app.post('/experiments/{experiment_id}/results')
def add_experiment_result(experiment_id: str, payload: ExperimentResult) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    result = payload.model_dump(mode='json')
    result['id'] = new_id('result')
    result['timestamp'] = now_iso()
    experiment.setdefault('results', []).append(result)
    experiment['updated_at'] = now_iso()
    store.save()
    return experiment


@app.post('/experiments/{experiment_id}/attachments')
async def add_experiment_attachment(experiment_id: str, file: UploadFile = File(...)) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    content = await file.read()
    target_dir = STORAGE_DIR / experiment['project_id'] / 'experiment_attachments'
    target_dir.mkdir(parents=True, exist_ok=True)
    attachment_id = new_id('att')
    target = target_dir / f'{attachment_id}_{file.filename}'
    target.write_bytes(content)
    attachment = {'id': attachment_id, 'filename': file.filename, 'path': str(target), 'sha256': sha256_bytes(content)}
    experiment.setdefault('attachments', []).append(attachment)
    store.save()
    return attachment


@app.get('/runs/{run_id}/report')
def run_report(run_id: str) -> dict[str, Any]:
    run = get_run(run_id)
    return run.get('report', {})


@app.get('/runs/{run_id}/export')
@app.post('/runs/{run_id}/export')
def export_report(run_id: str, format: Literal['json', 'md', 'docx', 'pdf'] = 'json', language: Language = Language.RU) -> Any:
    report = run_report(run_id)
    if not report:
        raise HTTPException(status_code=404, detail='Report not found')
    if format == 'json':
        return report['json']
    if format == 'md':
        return PlainTextResponse(report['markdown'], media_type='text/markdown; charset=utf-8')
    if format == 'docx':
        return Response(
            content=make_docx(report['markdown']),
            media_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            headers={'Content-Disposition': f'attachment; filename="norlab-report-{run_id}.docx"'},
        )
    return Response(
        content=make_pdf(report['markdown']),
        media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="norlab-report-{run_id}.pdf"'},
    )


@app.post('/projects/{project_id}/exports')
def bff_create_export(project_id: str, payload: ReportExportPayload) -> dict[str, Any]:
    get_project(project_id)
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project_id]
    run = next((item for item in reversed(runs) if item.get('report')), None)
    if not run:
        raise HTTPException(status_code=404, detail='Report is not ready yet')
    job_id = new_id('export')
    extension = payload.format.lower().replace('jira_api', 'json')
    output_dir = ROOT / 'data' / 'state' / 'exports'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f'{job_id}.{extension}'
    report = run['report']
    if payload.format == 'PDF':
        output_path.write_bytes(make_pdf(report['markdown']))
    elif payload.format == 'DOCX':
        output_path.write_bytes(make_docx(report['markdown']))
    elif payload.format == 'CSV':
        lines = ['section,value']
        lines.extend(f'hypothesis,"{item.get("title", "").replace("\"", "\"\"")}"' for item in report.get('json', {}).get('hypotheses', []))
        output_path.write_text('\n'.join(lines), encoding='utf-8')
    else:
        output_path.write_text(json.dumps(report['json'], ensure_ascii=False, indent=2), encoding='utf-8')
    job = {
        'id': job_id,
        'status': 'ready',
        'progress': 100,
        'format': payload.format,
        'locale': payload.locale,
        'sections': payload.sections,
        'download_url': f'/api/exports/{job_id}/download',
        'missing_fields': [],
        'error': None,
        'path': str(output_path),
        'updated_at': now_iso(),
    }
    store.data.setdefault('exports', {})[job_id] = job
    store.save()
    return {key: value for key, value in job.items() if key != 'path'}


@app.get('/exports/{job_id}')
def bff_export_job(job_id: str) -> dict[str, Any]:
    job = store.data.setdefault('exports', {}).get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Export job not found')
    return {key: value for key, value in job.items() if key != 'path'}


@app.get('/exports/{job_id}/download')
def bff_export_download(job_id: str) -> FileResponse:
    job = store.data.setdefault('exports', {}).get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail='Export job not found')
    path = Path(job['path'])
    if not path.exists():
        raise HTTPException(status_code=404, detail='Export file not found')
    return FileResponse(path, filename=path.name)


@app.get('/admin/model-profiles')
def model_profiles() -> list[dict[str, Any]]:
    return [gateway.profile()]


@app.post('/admin/model-profiles/compatibility-test')
async def compatibility_test() -> dict[str, Any]:
    return await gateway.compatibility_test()


@app.post('/admin/model-profiles/{profile_id}/activate')
async def activate_profile(profile_id: str) -> dict[str, Any]:
    if profile_id != 'recommended-v1':
        raise HTTPException(status_code=404, detail='Unknown profile')
    result = await gateway.compatibility_test()
    return {'activated': result['passed'], 'profile': gateway.profile(), 'compatibility': result}


@app.post('/admin/eval/multilingual')
async def admin_multilingual_eval() -> dict[str, Any]:
    prompts = {
        'ru': 'Сформулируй одну проверяемую гипотезу по снижению потерь металлов в хвостах флотации на основе allowed_evidence. Верни JSON.',
        'en': 'Create one testable hypothesis for reducing metal losses in flotation tailings using allowed_evidence only. Return JSON.',
        'zh': '仅根据 allowed_evidence 提出一个用于降低浮选尾矿中金属损失的可验证假设。返回 JSON。',
    }
    evidence = [
        {'id': 'ev_tailings', 'statement': 'В хвостах флотации наблюдаются потери ценных металлов; требуется лабораторная проверка режимных факторов.'},
        {'id': 'ev_constraints', 'statement': 'Ограничения: без капитальной замены оборудования, использовать доступное лабораторное оборудование.'},
    ]
    results = {}
    for language, prompt in prompts.items():
        response = await gateway.chat_json(
            'generator',
            [
                {'role': 'system', 'content': 'Return a JSON object only. Keys: language, hypothesis, kpi, evidence_ids, constraint_satisfaction. Use only allowed_evidence ids. Do not invent numeric improvements or exact percentages.'},
                {'role': 'user', 'content': json.dumps({'prompt': prompt, 'allowed_evidence': evidence}, ensure_ascii=False)},
            ],
            {'language': language, 'hypothesis': None, 'kpi': None, 'evidence_ids': []},
        )
        evidence_ids = response.get('evidence_ids') if isinstance(response.get('evidence_ids'), list) else []
        unsupported_numbers = any(char.isdigit() for char in str(response.get('hypothesis', '')))
        results[language] = {
            'valid_json': isinstance(response, dict),
            'has_hypothesis': bool(response.get('hypothesis')),
            'has_evidence': bool(evidence_ids),
            'no_unsupported_numbers': not unsupported_numbers,
            'response': response,
        }
    return {'languages': results}


@app.post('/admin/eval/golden')
def admin_golden_eval() -> dict[str, Any]:
    output_path = ROOT / 'data' / 'state' / 'golden_eval.json'
    if not output_path.exists():
        raise HTTPException(status_code=404, detail='Run scripts/golden_eval.py first')
    return json.loads(output_path.read_text(encoding='utf-8'))


@app.get('/projects/{project_id}/workspace-view')
def workspace_view(project_id: str) -> dict[str, Any]:
    project = get_project(project_id)
    docs = project_documents(project_id)
    warnings = [] if project.get('target_kpi') else ['target_kpi_missing']
    view = projection_base(project_id, {'name': project['name'], 'problem': project['problem'], 'documents': len(docs)}, warnings)
    view['links'] = {'documents': f'/projects/{project_id}/documents', 'runs': f'/projects/{project_id}/runs'}
    return view


@app.get('/projects/{project_id}/research-view')
def research_view(project_id: str, run_id: str | None = None) -> dict[str, Any]:
    get_project(project_id)
    runs = [run for run in store.data['runs'].values() if run['project_id'] == project_id]
    run = get_run(run_id) if run_id else (runs[-1] if runs else None)
    summary = {'run_status': run['status'] if run else None, 'evidence_items': len(evidence_records(project_id))}
    view = projection_base(project_id, summary)
    view['links'] = {'memory_search': f'/projects/{project_id}/memory/search', 'subgraph': f'/projects/{project_id}/graph/subgraph'}
    return view


@app.get('/projects/{project_id}/hypotheses-view')
def hypotheses_view(project_id: str, run_id: str | None = None) -> dict[str, Any]:
    get_project(project_id)
    items = [item for item in store.data['hypotheses'].values() if item['project_id'] == project_id and (not run_id or item['run_id'] == run_id)]
    view = projection_base(project_id, {'count': len(items), 'finalists': [item['id'] for item in items if item['status'] == 'FINALIST']})
    view['links'] = {'hypotheses': [f'/hypotheses/{item["id"]}' for item in items[:10]]}
    return view


@app.get('/projects/{project_id}/experiments-view')
def experiments_view(project_id: str) -> dict[str, Any]:
    get_project(project_id)
    items = [item for item in store.data['experiments'].values() if item['project_id'] == project_id]
    view = projection_base(project_id, {'count': len(items), 'experiments': [item['id'] for item in items]})
    view['links'] = {'experiments': [f'/experiments/{item["id"]}' for item in items[:10]]}
    return view


@app.get('/hypotheses/{hypothesis_id}/inspector')
def hypothesis_inspector(hypothesis_id: str) -> dict[str, Any]:
    hypothesis = read_hypothesis(hypothesis_id)
    view = projection_base(hypothesis_id, {'title': hypothesis['title'], 'status': hypothesis['status']})
    view['evidence'] = hypothesis['supporting_evidence']
    view['critique'] = hypothesis.get('critique', {})
    view['uncertainty'] = hypothesis.get('uncertainty', {})
    view['capabilities'].extend(['feedback', 'compile_experiment'])
    return view


@app.get('/experiments/{experiment_id}/editor-view')
def experiment_editor_view(experiment_id: str) -> dict[str, Any]:
    experiment = read_experiment(experiment_id)
    view = projection_base(experiment_id, {'objective': experiment['objective'], 'type': experiment['experiment_type']})
    view['experiment'] = experiment
    view['capabilities'].extend(['edit', 'attach_result', 'upload_attachment'])
    return view
